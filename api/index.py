"""
Website Quality Audit (WQA) Generator - FastAPI Web Application

A web-based version of the WQA tool that can be deployed on Vercel.
Allows users to upload CSV files and receive Excel audit reports.
Supports Google OAuth for GA4 and GSC data integration.
"""

import io
import os
import json
import logging
import secrets
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse, urlencode, quote

import numpy as np
import pandas as pd
from fastapi import FastAPI, File, Form, UploadFile, HTTPException, Request, Response, Cookie
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware

# Google OAuth imports
try:
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GoogleRequest
    from googleapiclient.discovery import build
    import httpx
    GOOGLE_OAUTH_AVAILABLE = True
except ImportError:
    GOOGLE_OAUTH_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="WQA Generator",
    description="Website Quality Audit Generator for SEO Agencies",
    version="2.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# GOOGLE OAUTH CONFIGURATION
# =============================================================================

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REDIRECT_URI = os.environ.get("GOOGLE_REDIRECT_URI", "")
SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_hex(32))

# OAuth scopes needed for GA4 and GSC
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/analytics.readonly",
    "https://www.googleapis.com/auth/webmasters.readonly",
    "openid",
    "email",
    "profile"
]

# In-memory token storage (for serverless - tokens stored in encrypted cookies)
# In production, you'd want to use a database
token_storage: Dict[str, dict] = {}

def encrypt_token(token_data: dict) -> str:
    """Simple token encoding for cookie storage"""
    from itsdangerous import URLSafeTimedSerializer
    serializer = URLSafeTimedSerializer(SECRET_KEY)
    return serializer.dumps(token_data)

def decrypt_token(token_string: str) -> Optional[dict]:
    """Decode token from cookie"""
    from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
    serializer = URLSafeTimedSerializer(SECRET_KEY)
    try:
        return serializer.loads(token_string, max_age=86400)  # 24 hour expiry
    except (BadSignature, SignatureExpired):
        return None

def get_google_auth_url(state: str) -> str:
    """Generate Google OAuth URL"""
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": GOOGLE_REDIRECT_URI,
        "response_type": "code",
        "scope": " ".join(GOOGLE_SCOPES),
        "access_type": "offline",
        "prompt": "consent",
        "state": state
    }
    return f"https://accounts.google.com/o/oauth2/v2/auth?{urlencode(params)}"

async def exchange_code_for_tokens(code: str) -> dict:
    """Exchange authorization code for tokens"""
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "code": code,
                "grant_type": "authorization_code",
                "redirect_uri": GOOGLE_REDIRECT_URI
            }
        )
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail=f"Token exchange failed: {response.text}")
        return response.json()

async def refresh_access_token(refresh_token: str) -> dict:
    """Refresh an expired access token"""
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token"
            }
        )
        if response.status_code != 200:
            raise HTTPException(status_code=400, detail="Token refresh failed")
        return response.json()

def get_credentials_from_tokens(tokens: dict) -> Credentials:
    """Create Google credentials from token dict"""
    return Credentials(
        token=tokens.get("access_token"),
        refresh_token=tokens.get("refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=GOOGLE_SCOPES
    )

# =============================================================================
# COLUMN MAPPING CONFIGURATION
# =============================================================================

# Crawl Data Column Mapping - supports Screaming Frog, Sitebulb, JetOctopus, OnCrawl, etc.
CRAWL_COLUMN_MAP = {
    'url': ['address', 'url', 'page_url', 'page', 'page url', 'full url'],
    'status_code': ['status code', 'status_code', 'status', 'http_status', 'response_code', 'http status code'],
    'indexable': ['indexability', 'indexable', 'is_indexable', 'index_status', 'index status'],
    'meta_robots': ['meta robots 1', 'meta_robots', 'meta robots', 'robots', 'robots directive'],
    'canonical_url': ['canonical link element 1', 'canonical_url', 'canonical', 'canonical_link',
                      'canonical_tag', 'canonical url', 'canonicals'],
    'content_type': ['content type', 'content_type', 'mime type', 'type'],
    'inlinks': ['unique inlinks', 'inlinks', 'internal_inlinks', 'inlinks_count', 'internal inlinks',
                'internal links in', 'links in'],
    'outlinks': ['unique outlinks', 'outlinks', 'internal_outlinks', 'outlinks_count', 'internal outlinks',
                 'internal links out', 'links out'],
    'crawl_depth': ['crawl depth', 'crawl_depth', 'depth', 'click_depth', 'level', 'click depth',
                    'page depth', 'distance'],
    'in_sitemap': ['indexability status', 'in xml sitemap', 'in_sitemap', 'sitemap', 'in_xml_sitemap',
                   'sitemap_status', 'xml sitemap'],
    'page_title': ['title 1', 'page_title', 'title', 'meta_title', 'page title', 'seo title'],
    'meta_description': ['meta description 1', 'meta_description', 'description', 'meta description'],
    'h1': ['h1-1', 'h1', 'h1 1', 'heading 1', 'first h1'],
    'word_count': ['word count', 'word_count', 'words', 'content_word_count', 'text_word_count',
                   'text content', 'body word count'],
    'last_modified': ['last modified', 'last_modified', 'lastmod', 'modified', 'date modified',
                      'last-modified', 'modified date'],
}

GA_COLUMN_MAP = {
    'url': ['url', 'page_path', 'page', 'landing_page', 'page_location'],
    'sessions': ['sessions', 'session_count', 'visits'],
    'sessions_prev': ['sessions_prev', 'sessions_previous', 'previous_sessions', 'sessions_yoy',
                      'sessions_last_year', 'prior_sessions'],
    'conversions': ['conversions', 'goal_completions', 'conversion_count', 'key_events', 'keyEvents'],
    'bounce_rate': ['bounce_rate', 'bounceRate', 'bounce rate'],
    'avg_session_duration': ['avg_session_duration', 'averageSessionDuration', 'average session duration',
                             'session_duration', 'avgSessionDuration'],
    'ecom_revenue': ['ecom_revenue', 'revenue', 'transactionRevenue', 'purchaseRevenue', 'totalRevenue',
                     'ecommerce_revenue', 'total_revenue'],
}

GSC_COLUMN_MAP = {
    'url': ['url', 'page', 'top_pages', 'landing_page'],
    'avg_position': ['avg_position', 'position', 'average_position', 'avg_pos'],
    'ctr': ['ctr', 'click_through_rate', 'clickthrough_rate'],
    'clicks': ['clicks', 'click_count'],
    'impressions': ['impressions', 'impression_count'],
    'primary_keyword': ['primary_keyword', 'query', 'top_query', 'keyword'],
}

# Backlink Data Column Mapping - supports Ahrefs, Semrush, Moz, Majestic, etc.
# Supports both per-URL aggregated exports AND per-backlink raw exports (like SEMRush)
BACKLINK_COLUMN_MAP = {
    # Target URL (the page receiving the backlink)
    'url': ['url', 'target url', 'target_url', 'page', 'target', 'page url', 'target page'],
    # Source URL (the page providing the backlink) - used for per-backlink formats
    'source_url': ['source url', 'source_url', 'source', 'referring url', 'referring_url',
                   'from url', 'from_url', 'linking page'],
    # Pre-aggregated referring domains count
    'referring_domains': ['referring domains', 'referring_domains', 'ref domains', 'ref_domains',
                          'domains', 'rd', 'dofollow referring domains', 'root domains'],
    # Pre-aggregated backlinks count
    'backlinks': ['backlinks', 'backlink_count', 'external_links', 'links', 'total backlinks',
                  'dofollow backlinks', 'external backlinks', 'inbound links'],
    # Authority score (source page authority for per-backlink, target page for aggregated)
    'authority': ['ur', 'url rating', 'url_rating', 'authority score', 'authority_score',
                  'page authority', 'page_authority', 'pa', 'trust flow', 'citation flow',
                  'domain rating', 'dr', 'as', 'page ascore', 'ascore'],
    # Anchor text
    'anchor_texts': ['anchor_texts', 'anchors', 'anchor_text', 'anchor', 'top anchor'],
    # Nofollow indicator (for per-backlink formats)
    'nofollow': ['nofollow', 'no_follow', 'is_nofollow', 'rel_nofollow'],
}

# Keyword Tracking Data Column Mapping - supports SEMrush Position Tracking, Ahrefs Rank Tracker, etc.
KEYWORD_COLUMN_MAP = {
    # Target URL that ranks for the keyword
    'url': ['url', 'landing page', 'landing_page', 'page', 'page url', 'ranking url', 'target url'],
    # The keyword/query
    'keyword': ['keyword', 'query', 'search term', 'keyphrase', 'key phrase', 'target keyword'],
    # Search volume for the keyword
    'volume': ['volume', 'search volume', 'search_volume', 'monthly volume', 'avg. volume',
               'average volume', 'monthly searches'],
    # Current ranking position
    'position': ['position', 'rank', 'ranking', 'current position', 'current rank',
                 'google position', 'serp position'],
    # Previous ranking position (for tracking changes)
    'prev_position': ['previous position', 'prev_position', 'previous rank', 'last position',
                      'position change'],
    # Keyword difficulty
    'difficulty': ['difficulty', 'kd', 'keyword difficulty', 'kw difficulty', 'competition'],
    # CPC value
    'cpc': ['cpc', 'cost per click', 'avg cpc', 'avg. cpc'],
    # Intent classification
    'intent': ['intent', 'search intent', 'keyword intent'],
}


# =============================================================================
# API PROVIDER INTERFACES (Future Extension Points)
# =============================================================================
# These abstract base classes define the interface for future API integrations.
# Implement a provider by subclassing and registering with the provider registry.

from abc import ABC, abstractmethod
from typing import Protocol, runtime_checkable

@runtime_checkable
class CrawlProvider(Protocol):
    """Interface for crawl data providers (JetOctopus, custom crawlers, etc.)"""
    provider_name: str

    async def fetch_crawl_data(self, site_url: str, **options) -> pd.DataFrame:
        """Fetch crawl data for a site. Returns DataFrame with CRAWL_COLUMN_MAP fields."""
        ...

    def get_required_credentials(self) -> List[str]:
        """Return list of required credential keys (e.g., ['api_key', 'project_id'])"""
        ...


@runtime_checkable
class BacklinkProvider(Protocol):
    """Interface for backlink data providers (Ahrefs, Semrush, Moz, etc.)"""
    provider_name: str

    async def fetch_backlink_data(self, site_url: str, **options) -> pd.DataFrame:
        """Fetch backlink data for a site. Returns DataFrame with BACKLINK_COLUMN_MAP fields."""
        ...

    def get_required_credentials(self) -> List[str]:
        """Return list of required credential keys (e.g., ['api_key'])"""
        ...


# Provider registries (populated when providers are implemented)
CRAWL_PROVIDERS: Dict[str, type] = {}
BACKLINK_PROVIDERS: Dict[str, type] = {}

def register_crawl_provider(name: str):
    """Decorator to register a crawl provider"""
    def decorator(cls):
        CRAWL_PROVIDERS[name] = cls
        return cls
    return decorator

def register_backlink_provider(name: str):
    """Decorator to register a backlink provider"""
    def decorator(cls):
        BACKLINK_PROVIDERS[name] = cls
        return cls
    return decorator


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def find_column(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    df_columns_lower = {col.lower().strip(): col for col in df.columns}
    for name in possible_names:
        if name.lower() in df_columns_lower:
            return df_columns_lower[name.lower()]
    return None


def map_columns(df: pd.DataFrame, column_map: Dict[str, List[str]], source_name: str) -> pd.DataFrame:
    result = pd.DataFrame()
    for standard_name, possible_names in column_map.items():
        found_col = find_column(df, possible_names)
        if found_col:
            result[standard_name] = df[found_col]
        else:
            result[standard_name] = None
    return result


def normalize_indexable(value) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, bool):
        return value
    str_val = str(value).lower().strip()
    if str_val in ['true', '1', 'yes', 'index', 'indexable']:
        return True
    if str_val in ['false', '0', 'no', 'noindex', 'non-indexable', 'not indexable']:
        return False
    return True


def normalize_boolean(value) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    str_val = str(value).lower().strip()
    return str_val in ['true', '1', 'yes', 'y']


def normalize_url(url: str) -> str:
    if pd.isna(url) or not url:
        return ''
    url = str(url).strip().lower()
    if url.endswith('/') and len(url) > 1:
        parsed = urlparse(url)
        if parsed.path != '/':
            url = url.rstrip('/')
    return url


def url_to_page_path(url: str) -> str:
    """Extract normalized page path from a full URL for GA4 join.

    Examples:
        https://example.com/states/tennessee/ -> /states/tennessee
        https://example.com/ -> /
        /states/tennessee/ -> /states/tennessee
    """
    if pd.isna(url) or not url:
        return '/'
    url = str(url).strip()

    try:
        # If it's already just a path (starts with /), use it directly
        if url.startswith('/'):
            path = url
        else:
            parsed = urlparse(url)
            path = parsed.path or '/'

        # Normalize: lowercase, strip trailing slash (except for root)
        path = path.lower()
        if path != '/' and path.endswith('/'):
            path = path.rstrip('/')

        # Ensure it starts with /
        if not path.startswith('/'):
            path = '/' + path

        return path if path else '/'
    except Exception:
        return '/'


def normalize_page_path(path: str) -> str:
    """Normalize a GA4 pagePath value for joining.

    GA4 returns pagePath like '/states/tennessee/' - we normalize it to match url_to_page_path output.
    """
    if pd.isna(path) or not path:
        return '/'
    path = str(path).strip().lower()

    # Ensure starts with /
    if not path.startswith('/'):
        path = '/' + path

    # Strip trailing slash (except for root)
    if path != '/' and path.endswith('/'):
        path = path.rstrip('/')

    return path if path else '/'


def classify_page_type(url: str, page_title: str = '', content_type: str = '') -> str:
    """
    Classify page into expanded taxonomy:
    - Home Page: Root/homepage
    - Local Lander: Location/city/area pages
    - Landing Page: Service/product/solution pages
    - Lead Generation: Contact, quote, demo request pages
    - Blog Post: Individual blog articles
    - Blog Category: Blog listing/category/tag pages
    - Resource / Guide: Resources, guides, whitepapers, case studies
    - Site Info: About, privacy, terms, careers pages
    - Other: Anything else
    """
    if not url or pd.isna(url):
        return 'Other'
    url_lower = str(url).lower()
    title_lower = str(page_title).lower() if page_title else ''

    try:
        parsed = urlparse(url_lower)
        path = parsed.path.replace('\\', '/')
    except Exception:
        path = url_lower.replace('\\', '/')

    # Home Page - root path only
    if path in ['/', ''] or path.rstrip('/') == '':
        return 'Home Page'

    # Lead Generation - contact, quote, demo pages (check early - high priority)
    lead_gen_patterns = ['/contact', '/get-quote', '/request-quote', '/free-quote',
                         '/demo', '/request-demo', '/schedule', '/book-', '/appointment',
                         '/consultation', '/free-estimate', '/get-started', '/signup', '/sign-up',
                         '/register', '/inquiry', '/enquiry']
    if any(pattern in path for pattern in lead_gen_patterns):
        return 'Lead Generation'

    # Blog Category - blog listing/category/tag pages (check before Blog Post)
    blog_category_patterns = ['/blog/', '/blog', '/news/', '/news', '/articles/', '/posts/']
    # Check if it's a category/tag/author page or the main blog listing
    if any(pattern in path for pattern in blog_category_patterns):
        # If path ends with /blog or /blog/ or has category/tag/author/page patterns
        category_indicators = ['/category/', '/tag/', '/author/', '/page/', '/topics/']
        if (path.rstrip('/').endswith('/blog') or
            path.rstrip('/').endswith('/news') or
            path.rstrip('/').endswith('/articles') or
            any(ind in path for ind in category_indicators)):
            return 'Blog Category'
        # Otherwise it's likely a blog post (has additional path segments)
        return 'Blog Post'

    # Local Lander - location/city/area pages
    local_patterns = ['/location/', '/locations/', '/city/', '/cities/',
                      '/service-area/', '/service-areas/', '/area/', '/areas/',
                      '/near-me', '/local/', '/region/', '/state/', '/county/']
    if any(pattern in path for pattern in local_patterns):
        return 'Local Lander'

    # Resource / Guide - educational content, downloads, case studies
    resource_patterns = ['/resource/', '/resources/', '/guide/', '/guides/',
                        '/whitepaper/', '/whitepapers/', '/ebook/', '/ebooks/',
                        '/case-study/', '/case-studies/', '/casestudy/', '/casestudies/',
                        '/download/', '/downloads/', '/library/', '/learn/',
                        '/how-to/', '/tutorial/', '/tutorials/', '/faq/', '/faqs/',
                        '/knowledge-base/', '/kb/', '/help-center/', '/documentation/']
    if any(pattern in path for pattern in resource_patterns):
        return 'Resource / Guide'

    # Site Info - corporate/legal/info pages
    site_info_patterns = ['/about', '/privacy', '/terms', '/legal/', '/disclaimer',
                          '/careers/', '/jobs/', '/team/', '/leadership/',
                          '/cookie', '/gdpr', '/accessibility', '/sitemap',
                          '/press/', '/media/', '/awards/', '/testimonials/',
                          '/partners/', '/affiliate/', '/referral/']
    if any(pattern in path for pattern in site_info_patterns):
        return 'Site Info'

    # Landing Page - service/product/solution pages
    landing_patterns = ['/service/', '/services/', '/solutions/', '/solution/',
                       '/products/', '/product/', '/what-we-do/', '/offerings/',
                       '/industries/', '/industry/', '/capabilities/', '/expertise/']
    if any(pattern in path for pattern in landing_patterns):
        return 'Landing Page'

    return 'Other'


# =============================================================================
# DATA LOADING
# =============================================================================

def read_file_to_dataframe(file_content: bytes, filename: str = "") -> pd.DataFrame:
    """Read file content into DataFrame, auto-detecting CSV or Excel format"""
    filename_lower = filename.lower() if filename else ""

    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        return pd.read_excel(io.BytesIO(file_content))
    else:
        # Try CSV with different encodings, fall back to Excel if all fail
        encodings_to_try = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        for encoding in encodings_to_try:
            try:
                return pd.read_csv(io.BytesIO(file_content), low_memory=False, encoding=encoding)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                # If it's not an encoding error, try Excel
                break
        # Try Excel as last resort
        try:
            return pd.read_excel(io.BytesIO(file_content))
        except Exception:
            raise ValueError(f"Could not read file as CSV or Excel: {filename}")


def load_crawl_data(file_content: bytes, filename: str = "") -> pd.DataFrame:
    df = read_file_to_dataframe(file_content, filename)
    df_mapped = map_columns(df, CRAWL_COLUMN_MAP, 'Crawl')

    if 'indexable' in df_mapped.columns and df_mapped['indexable'] is not None:
        df_mapped['indexable'] = df_mapped['indexable'].apply(normalize_indexable)
    if 'in_sitemap' in df_mapped.columns and df_mapped['in_sitemap'] is not None:
        df_mapped['in_sitemap'] = df_mapped['in_sitemap'].apply(normalize_boolean)

    numeric_cols = ['status_code', 'inlinks', 'outlinks', 'crawl_depth', 'word_count']
    for col in numeric_cols:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0).astype(int)

    return df_mapped


def load_ga_data(file_content: Optional[bytes], filename: str = "") -> Optional[pd.DataFrame]:
    if file_content is None:
        return None
    df = read_file_to_dataframe(file_content, filename)
    df_mapped = map_columns(df, GA_COLUMN_MAP, 'GA4')
    # Convert numeric columns
    numeric_cols = ['sessions', 'sessions_prev', 'conversions', 'bounce_rate', 'avg_session_duration', 'ecom_revenue']
    for col in numeric_cols:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)
    return df_mapped


def load_gsc_data(file_content: Optional[bytes], filename: str = "") -> Optional[pd.DataFrame]:
    if file_content is None:
        return None
    df = read_file_to_dataframe(file_content, filename)
    df_mapped = map_columns(df, GSC_COLUMN_MAP, 'GSC')
    for col in ['avg_position', 'ctr', 'clicks', 'impressions']:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)
    return df_mapped


def load_keyword_data(file_content: Optional[bytes], filename: str = "") -> Optional[pd.DataFrame]:
    """Load keyword tracking data and aggregate by URL to get best keyword per page.

    Returns DataFrame with columns: url, main_kw, main_kw_volume, main_kw_ranking, best_kw, best_kw_ranking
    - main_kw: Keyword with highest search volume for the URL
    - best_kw: Keyword with best (lowest) ranking position for the URL
    """
    if file_content is None:
        return None

    df = read_file_to_dataframe(file_content, filename)
    df_mapped = map_columns(df, KEYWORD_COLUMN_MAP, 'Keywords')

    # Check required columns
    if 'url' not in df_mapped.columns or 'keyword' not in df_mapped.columns:
        logger.warning("Keyword data missing required columns (url, keyword)")
        return None

    # Normalize URLs
    df_mapped['url'] = df_mapped['url'].apply(normalize_url)
    df_mapped = df_mapped[df_mapped['url'] != '']

    # Convert numeric columns
    for col in ['volume', 'position', 'difficulty', 'cpc']:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)

    # Ensure position is valid (greater than 0)
    if 'position' in df_mapped.columns:
        df_mapped.loc[df_mapped['position'] <= 0, 'position'] = 999

    result_rows = []

    for url, group in df_mapped.groupby('url'):
        row_data = {'url': url}

        # Main KW: keyword with highest volume
        if 'volume' in group.columns:
            main_kw_row = group.loc[group['volume'].idxmax()]
            row_data['main_kw'] = main_kw_row.get('keyword', '')
            row_data['main_kw_volume'] = main_kw_row.get('volume', 0)
            row_data['main_kw_ranking'] = main_kw_row.get('position', 0) if main_kw_row.get('position', 0) < 999 else 0
        else:
            # If no volume, use first keyword
            first_row = group.iloc[0]
            row_data['main_kw'] = first_row.get('keyword', '')
            row_data['main_kw_volume'] = 0
            row_data['main_kw_ranking'] = first_row.get('position', 0) if first_row.get('position', 0) < 999 else 0

        # Best KW: keyword with best (lowest) ranking position
        if 'position' in group.columns:
            valid_positions = group[group['position'] < 999]
            if not valid_positions.empty:
                best_kw_row = valid_positions.loc[valid_positions['position'].idxmin()]
                row_data['best_kw'] = best_kw_row.get('keyword', '')
                row_data['best_kw_ranking'] = best_kw_row.get('position', 0)
            else:
                row_data['best_kw'] = ''
                row_data['best_kw_ranking'] = 0
        else:
            row_data['best_kw'] = ''
            row_data['best_kw_ranking'] = 0

        result_rows.append(row_data)

    result_df = pd.DataFrame(result_rows)
    logger.info(f"Loaded keyword data: {len(df_mapped)} keywords across {len(result_df)} URLs")
    return result_df


def extract_domain(url: str) -> str:
    """Extract domain from URL for referring domain counting"""
    if pd.isna(url) or not url:
        return ''
    try:
        from urllib.parse import urlparse
        parsed = urlparse(str(url))
        domain = parsed.netloc or parsed.path.split('/')[0]
        # Remove www. prefix for consistent domain counting
        if domain.startswith('www.'):
            domain = domain[4:]
        return domain.lower()
    except:
        return ''


def load_backlink_data(file_content: Optional[bytes], filename: str = "") -> Optional[pd.DataFrame]:
    """Load and process backlink data, supporting both aggregated and per-backlink formats.

    Detects format automatically:
    - If 'source_url' column exists: per-backlink format (SEMRush style) - aggregates by target URL
    - Otherwise: pre-aggregated format (Ahrefs style) - uses data as-is
    """
    if file_content is None:
        return None

    df = read_file_to_dataframe(file_content, filename)
    df_mapped = map_columns(df, BACKLINK_COLUMN_MAP, 'Backlinks')

    # Check if this is a per-backlink format (has source_url column)
    # This indicates SEMRush-style export where each row is one backlink
    if 'source_url' in df_mapped.columns and 'url' in df_mapped.columns:
        logger.info("Detected per-backlink format (SEMRush style) - aggregating by target URL")

        # Normalize target URLs
        df_mapped['url'] = df_mapped['url'].apply(normalize_url)
        df_mapped = df_mapped[df_mapped['url'] != '']

        # Extract source domains for referring domain counting
        df_mapped['source_domain'] = df_mapped['source_url'].apply(extract_domain)

        # Determine if link is dofollow (not nofollow)
        if 'nofollow' in df_mapped.columns:
            df_mapped['is_dofollow'] = ~df_mapped['nofollow'].fillna(False).astype(bool)
        else:
            df_mapped['is_dofollow'] = True

        # Aggregate by target URL
        aggregated = df_mapped.groupby('url').agg(
            backlinks=('url', 'count'),  # Count total backlinks
            referring_domains=('source_domain', 'nunique'),  # Count unique referring domains
            dofollow_links=('is_dofollow', 'sum'),  # Count dofollow links
            authority=('authority', 'max') if 'authority' in df_mapped.columns else ('url', lambda x: 0),  # Max authority score
        ).reset_index()

        # Ensure numeric types
        for col in ['backlinks', 'referring_domains', 'dofollow_links', 'authority']:
            if col in aggregated.columns:
                aggregated[col] = pd.to_numeric(aggregated[col], errors='coerce').fillna(0).astype(int)

        logger.info(f"Aggregated {len(df_mapped)} backlinks into {len(aggregated)} URLs")
        return aggregated

    else:
        # Pre-aggregated format - use as-is
        logger.info("Detected pre-aggregated backlink format")
        for col in ['referring_domains', 'backlinks', 'authority']:
            if col in df_mapped.columns:
                df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)
        return df_mapped


# =============================================================================
# DATA QUALITY VALIDATION
# =============================================================================

class DataQualityReport:
    """Container for data quality validation results."""

    def __init__(self):
        self.missing_columns: Dict[str, List[str]] = {}
        self.empty_columns: List[str] = []
        self.sparse_columns: Dict[str, float] = {}
        self.url_match_stats: Dict[str, Dict] = {}
        self.data_source_coverage: Dict[str, int] = {}
        self.issues: List[Dict] = []
        self.warnings: List[str] = []

    def add_issue(self, category: str, severity: str, description: str,
                  affected_count: int = 0, examples: List[str] = None):
        """Add a specific data quality issue."""
        self.issues.append({
            'category': category,
            'severity': severity,
            'description': description,
            'affected_count': affected_count,
            'examples': examples or []
        })

    def to_dataframe(self) -> pd.DataFrame:
        """Convert issues to a DataFrame for Excel export."""
        if not self.issues:
            return pd.DataFrame(columns=['Category', 'Severity', 'Description', 'Affected Count', 'Examples'])

        rows = []
        for issue in self.issues:
            rows.append({
                'Category': issue['category'],
                'Severity': issue['severity'],
                'Description': issue['description'],
                'Affected Count': issue['affected_count'],
                'Examples': '; '.join(str(e) for e in issue['examples'][:5]) if issue['examples'] else ''
            })

        return pd.DataFrame(rows)


def validate_data_quality(
    merged_df: pd.DataFrame,
    crawl_df: pd.DataFrame,
    ga_df: Optional[pd.DataFrame],
    gsc_df: Optional[pd.DataFrame],
    backlink_df: Optional[pd.DataFrame]
) -> DataQualityReport:
    """
    Validate merged data quality and identify missing/problematic data.
    """
    report = DataQualityReport()
    total_urls = len(merged_df)

    logger.info("=" * 60)
    logger.info("DATA QUALITY VALIDATION")
    logger.info("=" * 60)

    crawl_urls = set(merged_df['url'].dropna().unique())

    # GA4 URL matching
    if ga_df is not None:
        if 'page_path' in ga_df.columns:
            # API uses page_path for GA joining
            ga_paths = set(ga_df['page_path'].apply(normalize_page_path).dropna().unique())
            ga_paths = ga_paths - {''}
            crawl_paths = set(merged_df['page_path'].dropna().unique()) if 'page_path' in merged_df.columns else set()
            matched_ga = crawl_paths & ga_paths
            unmatched_ga = ga_paths - crawl_paths
            total_ga = len(ga_paths)
        elif 'url' in ga_df.columns:
            ga_urls = set(ga_df['url'].apply(normalize_url).dropna().unique())
            ga_urls = ga_urls - {''}
            matched_ga = crawl_urls & ga_urls
            unmatched_ga = ga_urls - crawl_urls
            total_ga = len(ga_urls)
        else:
            matched_ga = set()
            unmatched_ga = set()
            total_ga = 0

        report.data_source_coverage['GA4'] = len(matched_ga)
        report.url_match_stats['GA4'] = {
            'total_source_urls': total_ga,
            'matched': len(matched_ga),
            'unmatched': len(unmatched_ga),
            'match_rate': len(matched_ga) / total_ga * 100 if total_ga else 0
        }

        logger.info(f"GA4: {len(matched_ga)}/{total_ga} URLs matched ({report.url_match_stats['GA4']['match_rate']:.1f}%)")

        if report.url_match_stats['GA4']['match_rate'] < 50 and total_ga > 0:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'GA4 URL match rate is low ({report.url_match_stats["GA4"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_ga),
                examples=list(unmatched_ga)[:10]
            )
    else:
        report.data_source_coverage['GA4'] = 0
        report.warnings.append('GA4 data not provided - sessions/conversions will be 0')

    # GSC URL matching
    if gsc_df is not None and 'url' in gsc_df.columns:
        gsc_urls = set(gsc_df['url'].apply(normalize_url).dropna().unique())
        gsc_urls = gsc_urls - {''}
        matched_gsc = crawl_urls & gsc_urls
        unmatched_gsc = gsc_urls - crawl_urls

        report.data_source_coverage['GSC'] = len(matched_gsc)
        report.url_match_stats['GSC'] = {
            'total_source_urls': len(gsc_urls),
            'matched': len(matched_gsc),
            'unmatched': len(unmatched_gsc),
            'match_rate': len(matched_gsc) / len(gsc_urls) * 100 if gsc_urls else 0
        }

        logger.info(f"GSC: {len(matched_gsc)}/{len(gsc_urls)} URLs matched ({report.url_match_stats['GSC']['match_rate']:.1f}%)")

        if report.url_match_stats['GSC']['match_rate'] < 50:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'GSC URL match rate is low ({report.url_match_stats["GSC"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_gsc),
                examples=list(unmatched_gsc)[:10]
            )
    else:
        report.data_source_coverage['GSC'] = 0
        report.warnings.append('GSC data not provided - rankings/CTR will be 0')

    # Backlink URL matching
    if backlink_df is not None and 'url' in backlink_df.columns:
        bl_urls = set(backlink_df['url'].apply(normalize_url).dropna().unique())
        bl_urls = bl_urls - {''}
        matched_bl = crawl_urls & bl_urls
        unmatched_bl = bl_urls - crawl_urls

        report.data_source_coverage['Backlinks'] = len(matched_bl)
        report.url_match_stats['Backlinks'] = {
            'total_source_urls': len(bl_urls),
            'matched': len(matched_bl),
            'unmatched': len(unmatched_bl),
            'match_rate': len(matched_bl) / len(bl_urls) * 100 if bl_urls else 0
        }

        logger.info(f"Backlinks: {len(matched_bl)}/{len(bl_urls)} URLs matched ({report.url_match_stats['Backlinks']['match_rate']:.1f}%)")

        if report.url_match_stats['Backlinks']['match_rate'] < 50:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'Backlink URL match rate is low ({report.url_match_stats["Backlinks"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_bl),
                examples=list(unmatched_bl)[:10]
            )
    else:
        report.data_source_coverage['Backlinks'] = 0
        report.warnings.append('Backlink data not provided - referring_domains/backlinks will be 0')

    # Check for empty/sparse columns
    critical_columns = {
        'sessions': ('GA4', 0),
        'conversions': ('GA4', 0),
        'avg_position': ('GSC', 0.0),
        'ctr': ('GSC', 0.0),
        'clicks': ('GSC', 0),
        'impressions': ('GSC', 0),
        'primary_keyword': ('GSC', ''),
        'referring_domains': ('Backlinks', 0),
        'backlinks': ('Backlinks', 0),
        'word_count': ('Crawl', 0),
        'page_title': ('Crawl', ''),
        'meta_description': ('Crawl', ''),
    }

    for col, (source, default) in critical_columns.items():
        if col not in merged_df.columns:
            report.empty_columns.append(col)
            continue

        if isinstance(default, str):
            empty_mask = (merged_df[col].isna()) | (merged_df[col] == '') | (merged_df[col] == default)
        else:
            empty_mask = (merged_df[col].isna()) | (merged_df[col] == default)

        empty_count = empty_mask.sum()
        empty_pct = empty_count / total_urls * 100 if total_urls > 0 else 0
        report.sparse_columns[col] = empty_pct

        if empty_pct == 100:
            report.empty_columns.append(col)
            report.add_issue(
                category='Empty Column',
                severity='High',
                description=f'Column "{col}" has no data (100% empty/default). Source: {source}',
                affected_count=empty_count
            )
        elif empty_pct >= 90:
            report.add_issue(
                category='Sparse Data',
                severity='High',
                description=f'Column "{col}" is {empty_pct:.1f}% empty/default. Source: {source}',
                affected_count=empty_count
            )

    # Check data inconsistencies
    if 'sessions' in merged_df.columns and 'avg_position' in merged_df.columns:
        traffic_no_rank = merged_df[(merged_df['sessions'] > 0) & (merged_df['avg_position'] == 0)]
        if len(traffic_no_rank) > 0:
            report.add_issue(
                category='Data Inconsistency',
                severity='Medium',
                description=f'URLs with traffic but no GSC ranking data.',
                affected_count=len(traffic_no_rank),
                examples=list(traffic_no_rank['url'].head(10))
            )

    high_issues = sum(1 for i in report.issues if i['severity'] == 'High')
    medium_issues = sum(1 for i in report.issues if i['severity'] == 'Medium')

    logger.info(f"Data Quality: {len(report.issues)} issues ({high_issues} High, {medium_issues} Medium)")
    logger.info("=" * 60)

    return report


def generate_data_quality_sheets(report: DataQualityReport) -> Dict[str, pd.DataFrame]:
    """Generate DataFrames for data quality reporting in Excel."""
    sheets = {}

    sheets['issues'] = report.to_dataframe()

    coverage_data = []
    for source, count in report.data_source_coverage.items():
        stats = report.url_match_stats.get(source, {})
        coverage_data.append({
            'Data Source': source,
            'URLs Matched': count,
            'Total Source URLs': stats.get('total_source_urls', 'N/A'),
            'Match Rate (%)': f"{stats.get('match_rate', 0):.1f}" if stats else 'N/A',
            'Unmatched URLs': stats.get('unmatched', 0)
        })
    sheets['coverage'] = pd.DataFrame(coverage_data)

    sparsity_data = []
    for col, pct in sorted(report.sparse_columns.items(), key=lambda x: -x[1]):
        sparsity_data.append({
            'Column': col,
            'Empty/Default (%)': f"{pct:.1f}",
            'Status': 'Critical' if pct >= 90 else ('Warning' if pct >= 70 else 'OK')
        })
    sheets['column_quality'] = pd.DataFrame(sparsity_data)

    return sheets


# =============================================================================
# MERGING & PROCESSING
# =============================================================================

def merge_datasets(crawl_df, ga_df, gsc_df, backlink_df) -> pd.DataFrame:
    df = crawl_df.copy()
    df['url'] = df['url'].apply(normalize_url)
    df = df[df['url'] != '']

    # Create page_path column for GA4 joining
    df['page_path'] = df['url'].apply(url_to_page_path)

    # GA4 data joins on page_path (not full URL)
    if ga_df is not None and 'page_path' in ga_df.columns:
        ga_df = ga_df.copy()
        # Normalize page_path in GA data (should already be normalized from fetch)
        ga_df['page_path'] = ga_df['page_path'].apply(normalize_page_path)

        # Build aggregation dict dynamically based on available columns
        ga_agg_cols = {}
        if 'sessions' in ga_df.columns:
            ga_agg_cols['sessions'] = 'sum'
        if 'conversions' in ga_df.columns:
            ga_agg_cols['conversions'] = 'sum'
        if 'bounce_rate' in ga_df.columns:
            ga_agg_cols['bounce_rate'] = 'mean'
        if 'avg_session_duration' in ga_df.columns:
            ga_agg_cols['avg_session_duration'] = 'mean'
        if 'ecom_revenue' in ga_df.columns:
            ga_agg_cols['ecom_revenue'] = 'sum'
        if 'sessions_prev' in ga_df.columns:
            ga_agg_cols['sessions_prev'] = 'sum'
        if ga_agg_cols:
            ga_agg = ga_df.groupby('page_path').agg(ga_agg_cols).reset_index()
            df = df.merge(ga_agg, on='page_path', how='left', suffixes=('', '_ga'))
            logger.info(f"Merged GA4 data: {len(ga_agg)} paths, {df['sessions'].notna().sum()} matches")

    # Also support old-style GA data with 'url' column (from CSV uploads)
    elif ga_df is not None and 'url' in ga_df.columns:
        ga_df = ga_df.copy()
        ga_df['url'] = ga_df['url'].apply(normalize_url)
        ga_df = ga_df[ga_df['url'] != '']
        # Build aggregation dict dynamically based on available columns
        ga_agg_cols = {}
        if 'sessions' in ga_df.columns:
            ga_agg_cols['sessions'] = 'sum'
        if 'conversions' in ga_df.columns:
            ga_agg_cols['conversions'] = 'sum'
        if 'bounce_rate' in ga_df.columns:
            ga_agg_cols['bounce_rate'] = 'mean'
        if 'avg_session_duration' in ga_df.columns:
            ga_agg_cols['avg_session_duration'] = 'mean'
        if 'ecom_revenue' in ga_df.columns:
            ga_agg_cols['ecom_revenue'] = 'sum'
        if 'sessions_prev' in ga_df.columns:
            ga_agg_cols['sessions_prev'] = 'sum'
        if ga_agg_cols:
            ga_agg = ga_df.groupby('url').agg(ga_agg_cols).reset_index()
            df = df.merge(ga_agg, on='url', how='left', suffixes=('', '_ga'))

    if gsc_df is not None and 'url' in gsc_df.columns:
        gsc_df = gsc_df.copy()
        gsc_df['url'] = gsc_df['url'].apply(normalize_url)
        gsc_df = gsc_df[gsc_df['url'] != '']
        gsc_agg_cols = {}
        if 'avg_position' in gsc_df.columns:
            gsc_agg_cols['avg_position'] = 'mean'
        if 'ctr' in gsc_df.columns:
            gsc_agg_cols['ctr'] = 'mean'
        if 'clicks' in gsc_df.columns:
            gsc_agg_cols['clicks'] = 'sum'
        if 'impressions' in gsc_df.columns:
            gsc_agg_cols['impressions'] = 'sum'
        if 'primary_keyword' in gsc_df.columns:
            gsc_agg_cols['primary_keyword'] = 'first'
        if gsc_agg_cols:
            gsc_agg = gsc_df.groupby('url').agg(gsc_agg_cols).reset_index()
            df = df.merge(gsc_agg, on='url', how='left', suffixes=('', '_gsc'))

    if backlink_df is not None and 'url' in backlink_df.columns:
        backlink_df = backlink_df.copy()
        backlink_df['url'] = backlink_df['url'].apply(normalize_url)
        backlink_df = backlink_df[backlink_df['url'] != '']
        bl_agg_cols = {}
        if 'referring_domains' in backlink_df.columns:
            bl_agg_cols['referring_domains'] = 'sum'
        if 'backlinks' in backlink_df.columns:
            bl_agg_cols['backlinks'] = 'sum'
        if 'dofollow_links' in backlink_df.columns:
            bl_agg_cols['dofollow_links'] = 'sum'
        if 'authority' in backlink_df.columns:
            bl_agg_cols['authority'] = 'max'  # Take highest authority score if multiple entries
        if bl_agg_cols:
            bl_agg = backlink_df.groupby('url').agg(bl_agg_cols).reset_index()
            df = df.merge(bl_agg, on='url', how='left', suffixes=('', '_bl'))

    expected_columns = {
        # Crawl data columns
        'url': '', 'status_code': 0, 'indexable': True, 'canonical_url': '',
        'meta_robots': '', 'content_type': '', 'h1': '',
        'inlinks': 0, 'outlinks': 0, 'crawl_depth': 0, 'in_sitemap': False,
        'page_title': '', 'meta_description': '', 'word_count': 0,
        'last_modified': '',
        # GA data columns
        'sessions': 0, 'conversions': 0, 'bounce_rate': 0.0, 'avg_session_duration': 0.0,
        'ecom_revenue': 0.0, 'sessions_prev': 0,
        # GSC data columns
        'avg_position': 0.0, 'ctr': 0.0, 'clicks': 0, 'impressions': 0, 'primary_keyword': '',
        # Backlink data columns
        'referring_domains': 0, 'backlinks': 0, 'dofollow_links': 0, 'authority': 0,
    }

    for col, default in expected_columns.items():
        if col not in df.columns:
            df[col] = default
        else:
            if isinstance(default, str):
                df[col] = df[col].fillna(default).astype(str)
            elif isinstance(default, bool):
                df[col] = df[col].fillna(default).astype(bool)
            elif isinstance(default, int):
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default).astype(int)
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default)

    return df


def assign_actions(row, low_traffic_threshold=5, thin_content_threshold=1000,
                   high_rank_max_position=20.0, low_ctr_threshold=0.05) -> Tuple[List[str], List[str]]:
    """
    Assign Technical Actions and Content Actions to a URL.

    Technical Actions: Infrastructure/SEO technical changes (redirects, schema, sitemaps, internal links)
    Content Actions: Content quality changes (rewrite, refresh, update metadata) - NOT redirects
    """
    technical_actions = []
    content_actions = []

    status_code = int(row.get('status_code', 0))
    indexable = bool(row.get('indexable', True))
    canonical_url = str(row.get('canonical_url', '')).strip()
    url = str(row.get('url', '')).strip().lower()
    inlinks = int(row.get('inlinks', 0))
    crawl_depth = int(row.get('crawl_depth', 0))
    in_sitemap = bool(row.get('in_sitemap', False))
    page_title = str(row.get('page_title', '')).strip()
    meta_description = str(row.get('meta_description', '')).strip()
    word_count = int(row.get('word_count', 0))
    sessions = float(row.get('sessions', 0))
    conversions = float(row.get('conversions', 0))
    avg_position = float(row.get('avg_position', 0))
    ctr = float(row.get('ctr', 0))
    impressions = float(row.get('impressions', 0))
    referring_domains = int(row.get('referring_domains', 0))
    backlinks = int(row.get('backlinks', 0))
    page_type = str(row.get('page_type', 'Other'))

    # Track if page should be removed/redirected (affects content action assignment)
    marked_for_removal = False

    # =========================================================================
    # TECHNICAL ACTIONS - Infrastructure/SEO technical changes
    # =========================================================================

    # Status code fixes (redirects are TECHNICAL, not content)
    if status_code == 302:
        technical_actions.append('Convert to 301')

    if status_code == 404:
        if backlinks > 0 or referring_domains > 0:
            technical_actions.append('301 Redirect (has backlinks)')
            marked_for_removal = True
        elif inlinks > 0:
            technical_actions.append('Remove Internal Links')
            marked_for_removal = True
        else:
            marked_for_removal = True

    if status_code in [301, 308]:
        technical_actions.append('Review Redirect Chain')
        marked_for_removal = True

    # Low-value pages that should be redirected (TECHNICAL decision)
    if (status_code == 200 and sessions <= low_traffic_threshold and conversions == 0):
        if backlinks > 0 or referring_domains > 0:
            # Has link equity - redirect to preserve it
            technical_actions.append('301 Redirect (low value, has links)')
            marked_for_removal = True
        elif sessions == 0 and impressions == 0 and word_count < 300:
            # Zero-value thin content - delete
            technical_actions.append('Delete (404)')
            marked_for_removal = True

    # Sitemap management
    if not indexable and in_sitemap:
        technical_actions.append('Remove from Sitemap')
    if indexable and status_code == 200 and sessions > 0 and not in_sitemap:
        technical_actions.append('Add to Sitemap')

    # Canonicalization
    if canonical_url and canonical_url.lower() != url:
        technical_actions.append('Review Canonical')

    # Internal linking
    if status_code == 200 and not marked_for_removal:
        if sessions > 0 and inlinks == 0:
            technical_actions.append('Add Internal Links')
        important_pages = ['Home Page', 'Landing Page', 'Local Lander', 'Lead Generation']
        if crawl_depth >= 4 and page_type in important_pages:
            if 'Add Internal Links' not in technical_actions:
                technical_actions.append('Improve Page Depth')

    # Schema markup recommendations
    if status_code == 200 and not marked_for_removal:
        if page_type in ['Blog Post']:
            technical_actions.append('Add Schema: Article')
        elif page_type == 'Local Lander':
            technical_actions.append('Add Schema: LocalBusiness')
        elif page_type == 'Home Page':
            technical_actions.append('Add Schema: Organization')
        elif page_type == 'Resource / Guide':
            technical_actions.append('Add Schema: HowTo/FAQ')
        elif page_type == 'Lead Generation':
            technical_actions.append('Add Schema: ContactPage')

    # =========================================================================
    # CONTENT ACTIONS - Content quality improvements (only for live pages)
    # =========================================================================

    if status_code == 200 and not marked_for_removal:
        # Thin content needs rewrite
        if (word_count < thin_content_threshold and
            (sessions > 0 or avg_position > 0 or impressions > 0)):
            content_actions.append('Rewrite (Thin Content)')

        # Ranking content that could be improved
        elif word_count >= thin_content_threshold:
            # Content refresh for pages ranking but could do better
            if (sessions > 0 and 2 < avg_position <= high_rank_max_position):
                content_actions.append('Refresh')

            # Link building target for pages with no backlinks
            if (sessions > 0 and 3 <= avg_position <= 20 and
                referring_domains == 0 and backlinks == 0):
                content_actions.append('Target w/ Links')

        # Metadata improvements
        if not meta_description:
            content_actions.append('Add Meta Description')
        elif (avg_position > 0 and avg_position <= high_rank_max_position and
              ctr < low_ctr_threshold and impressions > 100):
            content_actions.append('Improve Meta Description')

        if not page_title:
            content_actions.append('Add Page Title')
        elif (avg_position > 0 and avg_position <= high_rank_max_position and
              ctr < low_ctr_threshold and impressions > 100):
            content_actions.append('Improve Page Title')

    # Default if no content actions needed
    if not content_actions and status_code == 200 and not marked_for_removal:
        content_actions.append('Leave As Is')

    return technical_actions, content_actions


def assign_priority(row) -> str:
    """Assign priority based on page importance and required actions."""
    status_code = int(row.get('status_code', 0))
    page_type = str(row.get('page_type', 'Other'))
    technical_actions = str(row.get('technical_actions', ''))
    content_actions = str(row.get('content_actions', ''))

    # High-priority page types
    important_pages = ['Home Page', 'Landing Page', 'Local Lander', 'Lead Generation']

    # Error status codes are always high priority
    if status_code in [404, 302, 500, 502, 503]:
        return 'High'

    # Important pages needing content work are high priority
    if page_type in important_pages:
        if 'Rewrite' in content_actions or 'Refresh' in content_actions:
            return 'High'
        if '301 Redirect' in technical_actions or 'Delete' in technical_actions:
            return 'High'  # Something's wrong with an important page

    # Redirect/delete decisions for less important pages
    if '301 Redirect' in technical_actions or 'Delete' in technical_actions:
        return 'Medium'

    # Any pages needing content work
    if 'Rewrite' in content_actions or 'Refresh' in content_actions:
        return 'Medium'

    return 'Low'


def determine_final_url(row) -> str:
    """
    Determine Final URL for pages that need redirecting/merging.

    Returns:
    - Empty string for pages that don't need redirecting
    - '[MANUAL ENTRY NEEDED]' for pages marked for 301 redirect (user must specify target)
    - The canonical URL if page is canonicalized to a different URL
    """
    technical_actions = str(row.get('technical_actions', ''))
    canonical = str(row.get('canonical_url', '')).strip()
    url = str(row.get('url', '')).strip().lower()
    status_code = int(row.get('status_code', 200))

    # If page is already a redirect, the target URL should be looked up
    if status_code in [301, 302, 307, 308]:
        return '[MANUAL ENTRY NEEDED]'

    # If page is marked for 301 redirect (any reason), user needs to specify target
    if '301 Redirect' in technical_actions:
        # If it's canonicalized to another URL, suggest that as the target
        if canonical and canonical.lower() != url:
            return canonical
        return '[MANUAL ENTRY NEEDED]'

    # If page is marked for deletion, no final URL needed
    if 'Delete' in technical_actions:
        return ''

    # If page is canonicalized but not marked for redirect, show the canonical
    if canonical and canonical.lower() != url:
        return canonical

    return ''


def generate_summary(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    summaries = {}

    tech_actions_flat = []
    for actions in df['technical_actions']:
        if actions:
            for action in str(actions).split(', '):
                action = action.strip()
                if action:
                    tech_actions_flat.append(action)

    tech_counts = pd.Series(tech_actions_flat).value_counts().reset_index()
    tech_counts.columns = ['Technical Action', 'Count']
    summaries['technical_actions'] = tech_counts

    content_actions_flat = []
    for actions in df['content_actions']:
        if actions:
            for action in str(actions).split(', '):
                action = action.strip()
                if action:
                    content_actions_flat.append(action)

    content_counts = pd.Series(content_actions_flat).value_counts().reset_index()
    content_counts.columns = ['Content Action', 'Count']
    summaries['content_actions'] = content_counts

    priority_counts = df['priority'].value_counts().reset_index()
    priority_counts.columns = ['Priority', 'Count']
    priority_order = ['High', 'Medium', 'Low']
    priority_counts['sort_order'] = priority_counts['Priority'].apply(
        lambda x: priority_order.index(x) if x in priority_order else 99
    )
    priority_counts = priority_counts.sort_values('sort_order').drop('sort_order', axis=1)
    summaries['priority'] = priority_counts

    page_type_counts = df['page_type'].value_counts().reset_index()
    page_type_counts.columns = ['Page Type', 'Count']
    summaries['page_type'] = page_type_counts

    status_counts = df['status_code'].value_counts().reset_index()
    status_counts.columns = ['Status Code', 'Count']
    status_counts = status_counts.sort_values('Status Code')
    summaries['status_code'] = status_counts

    return summaries


def extract_page_path(url: str) -> str:
    """Extract page path from URL (no protocol, no domain, no querystring)"""
    if pd.isna(url) or not url:
        return '/'
    try:
        parsed = urlparse(str(url))
        path = parsed.path if parsed.path else '/'
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        return path
    except:
        return '/'


def format_indexability(row) -> Tuple[str, str]:
    """
    Return (Index/Noindex label, Indexation Status explanation).

    Index/Noindex: Simple binary status
    Indexation Status: WHY the page is not indexable (reason), not just repeating the label
    """
    indexable = row.get('indexable', True)
    meta_robots = str(row.get('meta_robots', '')).lower()
    status_code = int(row.get('status_code', 200))
    canonical = str(row.get('canonical_url', '')).strip()
    url = str(row.get('url', '')).strip().lower()

    # Determine index/noindex label and specific reason
    if status_code >= 500:
        index_label = 'Non-Indexable'
        status_reason = 'Server Error'
    elif status_code == 404:
        index_label = 'Non-Indexable'
        status_reason = 'Not Found (404)'
    elif status_code in [301, 302, 307, 308]:
        index_label = 'Non-Indexable'
        status_reason = 'Redirected'
    elif status_code >= 400:
        index_label = 'Non-Indexable'
        status_reason = 'Client Error'
    elif 'noindex' in meta_robots:
        index_label = 'Non-Indexable'
        status_reason = 'Noindex Tag'
    elif 'nofollow' in meta_robots and 'noindex' not in meta_robots:
        # Nofollow but not noindex - still indexable
        index_label = 'Indexable'
        status_reason = 'Indexable (nofollow)'
    elif canonical and canonical.lower() != url:
        index_label = 'Non-Indexable'
        status_reason = 'Canonicalised'
    elif not indexable:
        index_label = 'Non-Indexable'
        status_reason = 'Blocked'  # Generic blocked reason
    else:
        index_label = 'Indexable'
        status_reason = 'OK'  # Clear signal that page is fine

    return index_label, status_reason


def create_excel_report(
    df: pd.DataFrame,
    summaries: Dict[str, pd.DataFrame],
    quality_report: Optional[DataQualityReport] = None
) -> bytes:
    """Create WQA Excel report with proper 3-header-row structure and data quality sheet"""
    output = io.BytesIO()

    # Define the 38-column structure
    # Row 1: Section headings (sparse)
    section_headers = {
        0: '', 1: 'ANALYSIS', 2: '', 3: '', 4: '', 5: '',
        6: 'KEYWORD PERFORMANCE', 7: '', 8: '', 9: '', 10: '', 11: '',
        12: 'TRAFFIC PERFORMANCE', 13: '', 14: '', 15: '',
        16: 'ENGAGEMENT', 17: '',
        18: 'CONVERSIONS', 19: '', 20: '', 21: '',
        22: 'ON PAGE', 23: '', 24: '', 25: '', 26: '', 27: '', 28: '', 29: '', 30: '',
        31: 'TECHNICAL', 32: '', 33: '', 34: '', 35: '', 36: '', 37: ''
    }

    # Row 2: Source labels
    source_labels = {
        0: 'Formula', 1: 'SF', 2: 'Manual', 3: 'Manual', 4: 'Manual', 5: 'Manual',
        6: 'SEMrush', 7: 'SEMrush', 8: 'SEMrush', 9: 'SEMrush', 10: 'SEMrush', 11: 'SEMrush',
        12: 'GSC', 13: 'GA', 14: 'GA', 15: 'GA',
        16: 'GA', 17: 'GA',
        18: 'GA', 19: 'GA', 20: 'GA', 21: 'GA',
        22: 'SF', 23: 'SF', 24: 'GSC', 25: 'SF', 26: 'SF', 27: 'SF', 28: 'SF', 29: 'SF', 30: 'Ahrefs',
        31: 'SF', 32: 'SF', 33: 'SF', 34: 'SF', 35: 'SF', 36: 'Formula', 37: 'SF'
    }

    # Row 3: Column names
    column_names = [
        'page-path', 'URL', 'Category', 'Technical Action', 'Content Action', 'Final URL',
        'Main KW', 'Volume', 'Ranking', '"Best" KW', 'Volume', 'Ranking',
        'Impressions', 'Sessions', '% Change Sessions', 'Losing Traffic?',
        'Bounce rate (%)', 'Average session duration',
        'Conversions (All Goals)', 'Conversion Rate (%)', 'Ecom Revenue Generated', 'Ecom Conversion Rate',
        'Type', 'Current Title', 'SERP CTR', 'Meta', 'H1', 'Word Count', 'Inlinks', 'Outlinks', 'DOFOLLOW Links',
        'Canonical Link Element', 'Status Code', 'Index / Noindex', 'Indexation Status', 'Page Depth', 'In Sitemap?', 'Last Modified'
    ]

    # Build data rows
    data_rows = []
    for _, row in df.iterrows():
        # Get canonical URL or regular URL
        canonical = str(row.get('canonical_url', '')).strip()
        url = str(row.get('url', '')).strip()
        display_url = canonical if canonical and canonical != '' else url

        # Extract page path
        page_path = extract_page_path(display_url)

        # Get indexability info
        index_label, indexation_status = format_indexability(row)

        # Calculate conversion rate
        sessions = float(row.get('sessions', 0))
        conversions = float(row.get('conversions', 0))
        conv_rate = (conversions / sessions) if sessions > 0 else ''

        # Calculate % change sessions and losing traffic
        sessions_prev = float(row.get('sessions_prev', 0))
        if sessions_prev > 0:
            pct_change = (sessions - sessions_prev) / sessions_prev
            losing_traffic = 'Yes' if pct_change <= -0.1 else 'No'
        else:
            pct_change = ''
            losing_traffic = 'No YoY Data'

        # Format in_sitemap
        in_sitemap = 'Yes' if row.get('in_sitemap', False) else 'No'

        # Get best keyword data - prefer keyword file data, fall back to GSC
        best_kw = row.get('best_kw', '')
        best_kw_volume = row.get('best_kw_volume', '')
        best_kw_ranking = row.get('best_kw_ranking', '')

        # Fall back to GSC data if no keyword file data
        if not best_kw:
            best_kw = row.get('primary_keyword', '')
            # Only show GSC ranking if we have a keyword from GSC
            best_kw_ranking = row.get('avg_position', '') if best_kw else ''
            best_kw_volume = ''  # GSC doesn't provide volume

        # Build row data matching column order
        data_row = [
            page_path,                                          # 0: page-path
            display_url,                                        # 1: URL
            row.get('page_type', 'Other'),                      # 2: Category
            row.get('technical_actions', ''),                   # 3: Technical Action
            row.get('content_actions', ''),                     # 4: Content Action
            row.get('final_url', ''),                           # 5: Final URL
            row.get('main_kw', ''),                             # 6: Main KW
            row.get('main_kw_volume', ''),                      # 7: Volume (Main KW)
            row.get('main_kw_ranking', ''),                     # 8: Ranking (Main KW)
            best_kw,                                            # 9: "Best" KW
            best_kw_volume,                                     # 10: Volume (Best KW)
            best_kw_ranking,                                    # 11: Ranking (Best KW)
            row.get('impressions', 0),                          # 12: Impressions
            sessions if sessions > 0 else '',                   # 13: Sessions
            pct_change,                                         # 14: % Change Sessions
            losing_traffic,                                     # 15: Losing Traffic?
            row.get('bounce_rate', ''),                         # 16: Bounce rate (%)
            row.get('avg_session_duration', ''),                # 17: Average session duration
            conversions if conversions > 0 else '',             # 18: Conversions (All Goals)
            conv_rate,                                          # 19: Conversion Rate (%)
            row.get('ecom_revenue', ''),                        # 20: Ecom Revenue Generated
            row.get('ecom_conv_rate', ''),                      # 21: Ecom Conversion Rate
            row.get('content_type', ''),                        # 22: Type
            row.get('page_title', ''),                          # 23: Current Title
            row.get('ctr', ''),                                 # 24: SERP CTR
            row.get('meta_description', ''),                    # 25: Meta
            row.get('h1', ''),                                  # 26: H1
            row.get('word_count', 0),                           # 27: Word Count
            row.get('inlinks', 0),                              # 28: Inlinks
            row.get('outlinks', 0),                             # 29: Outlinks
            row.get('dofollow_links', 0),                        # 30: DOFOLLOW Links
            row.get('canonical_url', ''),                       # 31: Canonical Link Element
            row.get('status_code', ''),                         # 32: Status Code
            index_label,                                        # 33: Index / Noindex
            indexation_status,                                  # 34: Indexation Status
            row.get('crawl_depth', ''),                         # 35: Page Depth
            in_sitemap,                                         # 36: In Sitemap?
            row.get('last_modified', ''),                       # 37: Last Modified
        ]
        data_rows.append(data_row)

    # Create DataFrame with proper structure
    wqa_df = pd.DataFrame(data_rows, columns=column_names)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write WQA sheet with 3 header rows
        workbook = writer.book
        worksheet = workbook.create_sheet('WQA', 0)

        # Row 1: Section headers with merging and distinct colors
        # Define merge ranges and their labels/colors
        # A1:F1 (cols 1-6), G1:L1 (cols 7-12), M1:P1 (cols 13-16), Q1:R1 (cols 17-18),
        # S1:V1 (cols 19-22), W1:AE1 (cols 23-31), AF1:AL1 (cols 32-38)
        section_merges = [
            ('A1:F1', 'ANALYSIS', '2F5496'),           # Dark Blue
            ('G1:L1', 'KEYWORD PERFORMANCE', '548235'), # Dark Green
            ('M1:P1', 'TRAFFIC PERFORMANCE', 'C65911'), # Dark Orange
            ('Q1:R1', 'ENGAGEMENT', '7030A0'),          # Purple
            ('S1:V1', 'CONVERSIONS', 'BF8F00'),         # Dark Gold
            ('W1:AE1', 'ON PAGE', '385723'),            # Forest Green
            ('AF1:AL1', 'TECHNICAL', '833C0C'),         # Dark Brown
        ]

        from openpyxl.styles import Font, PatternFill, Alignment

        for merge_range, label, color in section_merges:
            # Merge the cells
            worksheet.merge_cells(merge_range)
            # Get the first cell of the merged range to set value and style
            first_cell = merge_range.split(':')[0]
            cell = worksheet[first_cell]
            cell.value = label
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Source labels
        for col_idx, source in source_labels.items():
            worksheet.cell(row=2, column=col_idx + 1, value=source)

        # Row 3: Column names
        for col_idx, col_name in enumerate(column_names):
            worksheet.cell(row=3, column=col_idx + 1, value=col_name)

        # Row 4+: Data rows
        for row_idx, data_row in enumerate(data_rows):
            for col_idx, value in enumerate(data_row):
                worksheet.cell(row=row_idx + 4, column=col_idx + 1, value=value)

        # Apply formatting to Row 2 and Row 3 (Row 1 already styled with merges above)
        from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1

        # Header row styling for Row 3
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx in range(len(column_names)):
            # Source labels (Row 2)
            cell2 = worksheet.cell(row=2, column=col_idx + 1)
            cell2.fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
            cell2.font = Font(italic=True, size=9)

            # Column names (Row 3)
            cell3 = worksheet.cell(row=3, column=col_idx + 1)
            cell3.fill = header_fill
            cell3.font = header_font

        # Freeze panes at row 4 (after headers)
        worksheet.freeze_panes = 'A4'

        # Define column indices for formatting (0-based)
        # Percentage columns: % Change Sessions (14), Bounce rate (16), Conversion Rate (19), Ecom Conv Rate (21), SERP CTR (24)
        percentage_cols = [14, 16, 19, 21, 24]
        # Time duration column: Average session duration (17)
        duration_cols = [17]
        # Date column: Last Modified (37)
        date_cols = [37]
        # Currency column: Ecom Revenue (20)
        currency_cols = [20]
        # Number columns: Impressions (12), Sessions (13), Conversions (18), Word Count (27), Inlinks (28), Outlinks (29), DOFOLLOW (30)
        number_cols = [12, 13, 18, 27, 28, 29, 30]

        # Apply number formatting to data rows
        for row_idx in range(len(data_rows)):
            excel_row = row_idx + 4  # Data starts at row 4

            # Format percentage columns
            for col_idx in percentage_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        val = float(cell.value)
                        # If value is already decimal (0-1 range), use as-is
                        # If value is whole number (like 50 for 50%), convert
                        if val > 1:
                            cell.value = val / 100
                        cell.number_format = FORMAT_PERCENTAGE_00
                    except (ValueError, TypeError):
                        pass

            # Format duration columns (seconds to mm:ss)
            for col_idx in duration_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        seconds = float(cell.value)
                        # Convert seconds to time format (fraction of day)
                        cell.value = seconds / 86400  # Convert to Excel time
                        cell.number_format = '[mm]:ss'
                    except (ValueError, TypeError):
                        pass

            # Format currency columns
            for col_idx in currency_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '$#,##0.00'
                    except (ValueError, TypeError):
                        pass

            # Format number columns with comma separators
            for col_idx in number_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass

        # Auto-adjust column widths (basic)
        for col_idx, col_name in enumerate(column_names):
            col_letter = worksheet.cell(row=1, column=col_idx + 1).column_letter
            # Set reasonable default widths based on column type
            if col_name in ['URL', 'Canonical Link Element', 'Meta', 'Current Title']:
                worksheet.column_dimensions[col_letter].width = 50
            elif col_name in ['page-path', 'Final URL', 'H1']:
                worksheet.column_dimensions[col_letter].width = 35
            elif col_name in ['Technical Action', 'Content Action', 'Category', 'Indexation Status']:
                worksheet.column_dimensions[col_letter].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = 15

        # Also write the raw aggregation data to a separate sheet for reference
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Write summary sheet
        current_row = 0
        summary_order = [
            ('Priority Distribution', 'priority'),
            ('Technical Actions', 'technical_actions'),
            ('Content Actions', 'content_actions'),
            ('Page Types', 'page_type'),
            ('Status Codes', 'status_code'),
        ]
        for title, key in summary_order:
            if key in summaries:
                title_df = pd.DataFrame([[title]], columns=[''])
                title_df.to_excel(writer, sheet_name='Summary', index=False,
                                  header=False, startrow=current_row)
                current_row += 1
                summaries[key].to_excel(writer, sheet_name='Summary', index=False,
                                        startrow=current_row)
                current_row += len(summaries[key]) + 3

        # Sheet 4: Data Quality (if provided)
        if quality_report is not None:
            quality_sheets = generate_data_quality_sheets(quality_report)
            current_row = 0

            # Write header
            header_df = pd.DataFrame([['DATA QUALITY REPORT']], columns=[''])
            header_df.to_excel(writer, sheet_name='Data Quality', index=False,
                              header=False, startrow=current_row)
            current_row += 2

            # Write Data Source Coverage
            title_df = pd.DataFrame([['Data Source Coverage']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['coverage'].empty:
                quality_sheets['coverage'].to_excel(writer, sheet_name='Data Quality',
                                                    index=False, startrow=current_row)
                current_row += len(quality_sheets['coverage']) + 3

            # Write Column Quality
            title_df = pd.DataFrame([['Column Data Quality']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['column_quality'].empty:
                quality_sheets['column_quality'].to_excel(writer, sheet_name='Data Quality',
                                                          index=False, startrow=current_row)
                current_row += len(quality_sheets['column_quality']) + 3

            # Write Issues
            title_df = pd.DataFrame([['Data Quality Issues']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['issues'].empty:
                quality_sheets['issues'].to_excel(writer, sheet_name='Data Quality',
                                                  index=False, startrow=current_row)
                current_row += len(quality_sheets['issues']) + 3
            else:
                no_issues_df = pd.DataFrame([['No data quality issues found!']], columns=[''])
                no_issues_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                     header=False, startrow=current_row)

            # Write warnings if any
            if quality_report.warnings:
                current_row += 2
                title_df = pd.DataFrame([['Warnings']], columns=[''])
                title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                 header=False, startrow=current_row)
                current_row += 1
                warnings_df = pd.DataFrame([[w] for w in quality_report.warnings], columns=['Warning'])
                warnings_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                    startrow=current_row)

            logger.info(f"Wrote Data Quality sheet: {len(quality_report.issues)} issues found")

    output.seek(0)
    return output.getvalue()


# =============================================================================
# GOOGLE API INTEGRATION
# =============================================================================

async def fetch_ga4_properties(credentials: Credentials) -> List[dict]:
    """Fetch list of GA4 properties the user has access to"""
    try:
        service = build('analyticsadmin', 'v1beta', credentials=credentials)
        accounts = service.accountSummaries().list().execute()

        properties = []
        for account in accounts.get('accountSummaries', []):
            for prop in account.get('propertySummaries', []):
                properties.append({
                    'id': prop.get('property', '').replace('properties/', ''),
                    'name': prop.get('displayName', 'Unknown'),
                    'account': account.get('displayName', 'Unknown')
                })
        return properties
    except Exception as e:
        logger.error(f"Error fetching GA4 properties: {e}")
        return []

async def fetch_ga4_report(credentials: Credentials, property_id: str, days: int = 90) -> pd.DataFrame:
    """Fetch GA4 page data with current and previous period for YoY comparison.

    Returns DataFrame with columns:
    - page_path: normalized path for joining with crawl data
    - sessions, conversions, bounce_rate, avg_session_duration, ecom_revenue: current period
    - sessions_prev: previous period sessions for YoY comparison
    """
    try:
        service = build('analyticsdata', 'v1beta', credentials=credentials)

        # Current period: last N days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        # Previous period: same duration, one year ago
        prev_end_date = end_date - timedelta(days=365)
        prev_start_date = prev_end_date - timedelta(days=days)

        # === CURRENT PERIOD QUERY ===
        current_request = {
            'dateRanges': [{
                'startDate': start_date.strftime('%Y-%m-%d'),
                'endDate': end_date.strftime('%Y-%m-%d')
            }],
            'dimensions': [{'name': 'pagePath'}],
            'metrics': [
                {'name': 'sessions'},
                {'name': 'keyEvents'},  # GA4's conversion metric
                {'name': 'bounceRate'},
                {'name': 'averageSessionDuration'},
                {'name': 'purchaseRevenue'}  # E-commerce revenue
            ],
            'limit': 25000
        }

        current_response = service.properties().runReport(
            property=f'properties/{property_id}',
            body=current_request
        ).execute()

        # Build current period data map: page_path -> metrics
        current_data = {}
        for row in current_response.get('rows', []):
            raw_path = row['dimensionValues'][0]['value'] if row.get('dimensionValues') else '/'
            page_path = normalize_page_path(raw_path)
            metrics = row.get('metricValues', [])

            sessions = int(float(metrics[0]['value'])) if len(metrics) > 0 else 0
            conversions = int(float(metrics[1]['value'])) if len(metrics) > 1 else 0
            bounce_rate = float(metrics[2]['value']) if len(metrics) > 2 else 0.0
            avg_duration = float(metrics[3]['value']) if len(metrics) > 3 else 0.0
            revenue = float(metrics[4]['value']) if len(metrics) > 4 else 0.0

            # Aggregate if same path appears multiple times
            if page_path in current_data:
                current_data[page_path]['sessions'] += sessions
                current_data[page_path]['conversions'] += conversions
                current_data[page_path]['ecom_revenue'] += revenue
                # For rates, we'd need weighted average - keep simple for now
            else:
                current_data[page_path] = {
                    'sessions': sessions,
                    'conversions': conversions,
                    'bounce_rate': bounce_rate,
                    'avg_session_duration': avg_duration,
                    'ecom_revenue': revenue
                }

        logger.info(f"GA4 current period: {len(current_data)} unique page paths")

        # === PREVIOUS PERIOD QUERY (for YoY comparison) ===
        prev_request = {
            'dateRanges': [{
                'startDate': prev_start_date.strftime('%Y-%m-%d'),
                'endDate': prev_end_date.strftime('%Y-%m-%d')
            }],
            'dimensions': [{'name': 'pagePath'}],
            'metrics': [{'name': 'sessions'}],
            'limit': 25000
        }

        prev_data = {}
        try:
            prev_response = service.properties().runReport(
                property=f'properties/{property_id}',
                body=prev_request
            ).execute()

            for row in prev_response.get('rows', []):
                raw_path = row['dimensionValues'][0]['value'] if row.get('dimensionValues') else '/'
                page_path = normalize_page_path(raw_path)
                metrics = row.get('metricValues', [])
                sessions_prev = int(float(metrics[0]['value'])) if len(metrics) > 0 else 0

                if page_path in prev_data:
                    prev_data[page_path] += sessions_prev
                else:
                    prev_data[page_path] = sessions_prev

            logger.info(f"GA4 previous period: {len(prev_data)} unique page paths")
        except Exception as prev_e:
            logger.warning(f"Could not fetch previous period GA4 data: {prev_e}")

        # === COMBINE INTO DATAFRAME ===
        rows = []
        all_paths = set(current_data.keys()) | set(prev_data.keys())

        for page_path in all_paths:
            current = current_data.get(page_path, {})
            rows.append({
                'page_path': page_path,
                'sessions': current.get('sessions', 0),
                'conversions': current.get('conversions', 0),
                'bounce_rate': current.get('bounce_rate', 0.0),
                'avg_session_duration': current.get('avg_session_duration', 0.0),
                'ecom_revenue': current.get('ecom_revenue', 0.0),
                'sessions_prev': prev_data.get(page_path, 0)
            })

        result_df = pd.DataFrame(rows) if rows else pd.DataFrame(
            columns=['page_path', 'sessions', 'conversions', 'bounce_rate', 'avg_session_duration', 'ecom_revenue', 'sessions_prev']
        )

        logger.info(f"GA4 combined result: {len(result_df)} rows")
        return result_df

    except Exception as e:
        logger.error(f"Error fetching GA4 report: {e}")
        return pd.DataFrame(
            columns=['page_path', 'sessions', 'conversions', 'bounce_rate', 'avg_session_duration', 'ecom_revenue', 'sessions_prev']
        )

async def fetch_gsc_sites(credentials: Credentials) -> List[dict]:
    """Fetch list of verified GSC sites"""
    try:
        service = build('searchconsole', 'v1', credentials=credentials)
        sites = service.sites().list().execute()

        return [
            {
                'url': site.get('siteUrl', ''),
                'permission': site.get('permissionLevel', 'Unknown')
            }
            for site in sites.get('siteEntry', [])
        ]
    except Exception as e:
        logger.error(f"Error fetching GSC sites: {e}")
        return []

async def fetch_gsc_report(credentials: Credentials, site_url: str, days: int = 90) -> pd.DataFrame:
    """Fetch GSC performance data and return as DataFrame matching CSV format"""
    try:
        service = build('searchconsole', 'v1', credentials=credentials)

        end_date = datetime.now() - timedelta(days=3)  # GSC data has ~3 day lag
        start_date = end_date - timedelta(days=days)

        request_body = {
            'startDate': start_date.strftime('%Y-%m-%d'),
            'endDate': end_date.strftime('%Y-%m-%d'),
            'dimensions': ['page'],
            'rowLimit': 25000
        }

        response = service.searchanalytics().query(
            siteUrl=site_url,
            body=request_body
        ).execute()

        rows = []
        for row in response.get('rows', []):
            rows.append({
                'url': row['keys'][0] if row.get('keys') else '',
                'clicks': row.get('clicks', 0),
                'impressions': row.get('impressions', 0),
                'ctr': row.get('ctr', 0),
                'avg_position': row.get('position', 0)
            })

        # Also fetch top query per page for primary_keyword
        query_request = {
            'startDate': start_date.strftime('%Y-%m-%d'),
            'endDate': end_date.strftime('%Y-%m-%d'),
            'dimensions': ['page', 'query'],
            'rowLimit': 25000
        }

        try:
            query_response = service.searchanalytics().query(
                siteUrl=site_url,
                body=query_request
            ).execute()

            # Group by page and get top query by impressions
            page_keywords = {}
            for row in query_response.get('rows', []):
                page = row['keys'][0] if row.get('keys') else ''
                query = row['keys'][1] if len(row.get('keys', [])) > 1 else ''
                impressions = row.get('impressions', 0)

                if page not in page_keywords or impressions > page_keywords[page]['impressions']:
                    page_keywords[page] = {'query': query, 'impressions': impressions}

            # Add primary_keyword to rows
            for row in rows:
                if row['url'] in page_keywords:
                    row['primary_keyword'] = page_keywords[row['url']]['query']
                else:
                    row['primary_keyword'] = ''
        except Exception as e:
            logger.warning(f"Could not fetch query data: {e}")
            for row in rows:
                row['primary_keyword'] = ''

        return pd.DataFrame(rows) if rows else pd.DataFrame(
            columns=['url', 'clicks', 'impressions', 'ctr', 'avg_position', 'primary_keyword']
        )
    except Exception as e:
        logger.error(f"Error fetching GSC report: {e}")
        return pd.DataFrame(columns=['url', 'clicks', 'impressions', 'ctr', 'avg_position', 'primary_keyword'])


# =============================================================================
# OAUTH ENDPOINTS
# =============================================================================

@app.get("/api/auth/google")
async def google_auth_start(response: Response):
    """Start Google OAuth flow"""
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(
            status_code=503,
            detail="Google OAuth not configured. Set GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET, and GOOGLE_REDIRECT_URI environment variables."
        )

    state = secrets.token_urlsafe(32)
    auth_url = get_google_auth_url(state)

    # Store state in cookie for verification
    redirect = RedirectResponse(url=auth_url, status_code=302)
    redirect.set_cookie(key="oauth_state", value=state, httponly=True, max_age=600, samesite="lax")
    return redirect

@app.get("/api/auth/google/callback")
async def google_auth_callback(
    request: Request,
    code: str = None,
    state: str = None,
    error: str = None,
    oauth_state: Optional[str] = Cookie(None)
):
    """Handle Google OAuth callback"""
    if error:
        return RedirectResponse(url=f"/?error={error}", status_code=302)

    if not code:
        return RedirectResponse(url="/?error=no_code", status_code=302)

    # Verify state
    if state != oauth_state:
        return RedirectResponse(url="/?error=invalid_state", status_code=302)

    try:
        # Exchange code for tokens
        tokens = await exchange_code_for_tokens(code)

        # Encrypt tokens for cookie storage
        encrypted = encrypt_token(tokens)

        # Redirect to app with success
        redirect = RedirectResponse(url="/?connected=google", status_code=302)
        redirect.set_cookie(
            key="google_tokens",
            value=encrypted,
            httponly=True,
            max_age=86400,  # 24 hours
            samesite="lax",
            secure=True
        )
        redirect.delete_cookie(key="oauth_state")
        return redirect
    except Exception as e:
        logger.error(f"OAuth callback error: {e}")
        return RedirectResponse(url=f"/?error=auth_failed", status_code=302)

@app.get("/api/auth/status")
async def auth_status(google_tokens: Optional[str] = Cookie(None)):
    """Check if user is authenticated with Google"""
    if not google_tokens:
        return {"connected": False}

    tokens = decrypt_token(google_tokens)
    if not tokens:
        return {"connected": False}

    return {"connected": True, "has_refresh": "refresh_token" in tokens}

@app.post("/api/auth/logout")
async def logout(response: Response):
    """Clear Google authentication"""
    response = JSONResponse({"success": True})
    response.delete_cookie(key="google_tokens")
    return response

@app.get("/api/ga4/properties")
async def list_ga4_properties(google_tokens: Optional[str] = Cookie(None)):
    """List GA4 properties available to the user"""
    if not google_tokens:
        raise HTTPException(status_code=401, detail="Not authenticated with Google")

    tokens = decrypt_token(google_tokens)
    if not tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired session")

    credentials = get_credentials_from_tokens(tokens)
    properties = await fetch_ga4_properties(credentials)
    return {"properties": properties}

@app.get("/api/gsc/sites")
async def list_gsc_sites(google_tokens: Optional[str] = Cookie(None)):
    """List GSC sites available to the user"""
    if not google_tokens:
        raise HTTPException(status_code=401, detail="Not authenticated with Google")

    tokens = decrypt_token(google_tokens)
    if not tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired session")

    credentials = get_credentials_from_tokens(tokens)
    sites = await fetch_gsc_sites(credentials)
    return {"sites": sites}

@app.post("/api/ga4/report")
async def get_ga4_report(
    property_id: str = Form(...),
    days: int = Form(30),
    google_tokens: Optional[str] = Cookie(None)
):
    """Fetch GA4 report data"""
    if not google_tokens:
        raise HTTPException(status_code=401, detail="Not authenticated with Google")

    tokens = decrypt_token(google_tokens)
    if not tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired session")

    credentials = get_credentials_from_tokens(tokens)
    df = await fetch_ga4_report(credentials, property_id, days)
    return {"rows": df.to_dict('records'), "count": len(df)}

@app.post("/api/gsc/report")
async def get_gsc_report(
    site_url: str = Form(...),
    days: int = Form(90),
    google_tokens: Optional[str] = Cookie(None)
):
    """Fetch GSC report data"""
    if not google_tokens:
        raise HTTPException(status_code=401, detail="Not authenticated with Google")

    tokens = decrypt_token(google_tokens)
    if not tokens:
        raise HTTPException(status_code=401, detail="Invalid or expired session")

    credentials = get_credentials_from_tokens(tokens)
    df = await fetch_gsc_report(credentials, site_url, days)
    return {"rows": df.to_dict('records'), "count": len(df)}


# =============================================================================
# API ENDPOINTS
# =============================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the main HTML page"""
    return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WQA Generator - Website Quality Audit</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container {
            max-width: 900px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 40px;
        }
        h1 { color: #333; margin-bottom: 10px; font-size: 2rem; }
        .subtitle { color: #666; margin-bottom: 30px; font-size: 1.1rem; }
        .form-section { margin-bottom: 25px; }
        .form-section h3 {
            color: #444; margin-bottom: 10px; font-size: 1rem;
            display: flex; align-items: center; gap: 8px;
        }
        .required { color: #e74c3c; }
        .optional { color: #27ae60; font-size: 0.85rem; }
        .file-input-wrapper { position: relative; overflow: hidden; display: inline-block; width: 100%; }
        .file-input-wrapper input[type=file] {
            font-size: 100px; position: absolute; left: 0; top: 0;
            opacity: 0; cursor: pointer; width: 100%; height: 100%;
        }
        .file-input-btn {
            background: #f8f9fa; border: 2px dashed #ddd; border-radius: 8px;
            padding: 20px; text-align: center; transition: all 0.3s; cursor: pointer;
        }
        .file-input-btn:hover { border-color: #667eea; background: #f0f0ff; }
        .file-input-btn.has-file { border-color: #27ae60; background: #e8f8f0; }
        .file-input-btn.disabled { opacity: 0.5; cursor: not-allowed; }
        .file-name { margin-top: 8px; font-size: 0.9rem; color: #27ae60; }
        .threshold-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin-top: 20px; }
        .threshold-item label { display: block; font-size: 0.85rem; color: #666; margin-bottom: 5px; }
        .threshold-item input, .threshold-item select {
            width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 1rem;
        }
        .submit-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; border: none; padding: 15px 40px; font-size: 1.1rem;
            border-radius: 8px; cursor: pointer; width: 100%; margin-top: 20px;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .submit-btn:hover { transform: translateY(-2px); box-shadow: 0 5px 20px rgba(102, 126, 234, 0.4); }
        .submit-btn:disabled { opacity: 0.6; cursor: not-allowed; transform: none; }
        .status { margin-top: 20px; padding: 15px; border-radius: 8px; display: none; }
        .status.loading { display: block; background: #fff3cd; color: #856404; }
        .status.success { display: block; background: #d4edda; color: #155724; }
        .status.error { display: block; background: #f8d7da; color: #721c24; }
        .info-box {
            background: #e8f4fd; border-left: 4px solid #667eea;
            padding: 15px; margin-bottom: 25px; border-radius: 0 8px 8px 0;
        }
        .info-box p { color: #444; font-size: 0.9rem; line-height: 1.5; }
        .helper-text {
            margin-top: 10px; padding: 12px 15px; background: #f8f9fa;
            border-radius: 6px; font-size: 0.85rem; color: #555; line-height: 1.6;
        }
        .helper-text strong { color: #444; }

        /* Google OAuth styles */
        .google-section {
            background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
            border: 2px solid #dee2e6; border-radius: 12px;
            padding: 25px; margin-bottom: 30px;
        }
        .google-section h2 { color: #333; font-size: 1.3rem; margin-bottom: 15px; display: flex; align-items: center; gap: 10px; }
        .google-btn {
            display: inline-flex; align-items: center; gap: 10px;
            background: white; border: 2px solid #4285f4; color: #4285f4;
            padding: 12px 24px; border-radius: 8px; font-size: 1rem;
            cursor: pointer; transition: all 0.3s; font-weight: 500;
        }
        .google-btn:hover { background: #4285f4; color: white; }
        .google-btn.connected { background: #34a853; border-color: #34a853; color: white; }
        .google-btn svg { width: 20px; height: 20px; }
        .google-status { margin-top: 15px; font-size: 0.9rem; }
        .google-status.connected { color: #34a853; }
        .google-status.not-connected { color: #666; }
        .logout-btn {
            background: none; border: none; color: #dc3545;
            cursor: pointer; font-size: 0.85rem; margin-left: 15px;
            text-decoration: underline;
        }

        /* Data source toggle */
        .data-source-toggle {
            display: flex; gap: 10px; margin-bottom: 15px;
            background: #f1f3f4; padding: 5px; border-radius: 8px;
        }
        .toggle-btn {
            flex: 1; padding: 10px 15px; border: none; background: transparent;
            border-radius: 6px; cursor: pointer; font-size: 0.9rem;
            transition: all 0.2s; color: #666;
        }
        .toggle-btn.active { background: white; color: #667eea; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
        .toggle-btn:hover:not(.active) { background: rgba(255,255,255,0.5); }

        /* Property/site selectors */
        .api-selector { margin-top: 15px; }
        .api-selector select {
            width: 100%; padding: 12px; border: 2px solid #ddd;
            border-radius: 8px; font-size: 1rem; background: white;
        }
        .api-selector select:focus { border-color: #667eea; outline: none; }
        .api-selector.has-selection select { border-color: #27ae60; }
        .api-selector label { display: block; font-size: 0.85rem; color: #666; margin-bottom: 8px; }
        .loading-indicator { color: #666; font-style: italic; padding: 10px 0; }

        .hidden { display: none !important; }

        @media (max-width: 600px) {
            .threshold-grid { grid-template-columns: 1fr; }
            .container { padding: 20px; }
            .data-source-toggle { flex-direction: column; }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>WQA Generator</h1>
        <p class="subtitle">Website Quality Audit Report Generator for SEO Agencies</p>

        <div class="info-box">
            <p><strong>How it works:</strong> Upload your crawl data (required) and add GA4/GSC data either by connecting your Google account or uploading CSVs. The tool merges all data sources, applies SEO rules, and generates a comprehensive Excel report with prioritized actions.</p>
        </div>

        <!-- Google OAuth Section -->
        <div class="google-section">
            <h2>
                <svg viewBox="0 0 24 24" width="24" height="24"><path fill="#4285F4" d="M22.56 12.25c0-.78-.07-1.53-.2-2.25H12v4.26h5.92c-.26 1.37-1.04 2.53-2.21 3.31v2.77h3.57c2.08-1.92 3.28-4.74 3.28-8.09z"/><path fill="#34A853" d="M12 23c2.97 0 5.46-.98 7.28-2.66l-3.57-2.77c-.98.66-2.23 1.06-3.71 1.06-2.86 0-5.29-1.93-6.16-4.53H2.18v2.84C3.99 20.53 7.7 23 12 23z"/><path fill="#FBBC05" d="M5.84 14.09c-.22-.66-.35-1.36-.35-2.09s.13-1.43.35-2.09V7.07H2.18C1.43 8.55 1 10.22 1 12s.43 3.45 1.18 4.93l2.85-2.22.81-.62z"/><path fill="#EA4335" d="M12 5.38c1.62 0 3.06.56 4.21 1.64l3.15-3.15C17.45 2.09 14.97 1 12 1 7.7 1 3.99 3.47 2.18 7.07l3.66 2.84c.87-2.6 3.3-4.53 6.16-4.53z"/></svg>
                Connect Google Data
            </h2>
            <p style="color: #666; margin-bottom: 15px; font-size: 0.9rem;">Connect your Google account to automatically pull GA4 and Search Console data instead of uploading CSVs.</p>

            <div id="googleAuthSection">
                <button class="google-btn" id="googleConnectBtn" onclick="window.location.href='/api/auth/google'">
                    <svg viewBox="0 0 24 24"><path fill="currentColor" d="M22.56 12.25c0-.78-.07-1.53-.2-2.25H12v4.26h5.92c-.26 1.37-1.04 2.53-2.21 3.31v2.77h3.57c2.08-1.92 3.28-4.74 3.28-8.09z"/><path fill="currentColor" d="M12 23c2.97 0 5.46-.98 7.28-2.66l-3.57-2.77c-.98.66-2.23 1.06-3.71 1.06-2.86 0-5.29-1.93-6.16-4.53H2.18v2.84C3.99 20.53 7.7 23 12 23z"/></svg>
                    Connect with Google
                </button>
                <div class="google-status not-connected" id="googleStatus">Not connected</div>
            </div>

            <!-- Property/Site selectors (shown when connected) -->
            <div id="googleSelectors" class="hidden">
                <div class="api-selector" id="ga4SelectorWrapper">
                    <label>GA4 Property</label>
                    <select id="ga4PropertySelect">
                        <option value="">-- Select GA4 Property --</option>
                    </select>
                </div>
                <div class="api-selector" id="gscSelectorWrapper" style="margin-top: 15px;">
                    <label>Search Console Site</label>
                    <select id="gscSiteSelect">
                        <option value="">-- Select GSC Site --</option>
                    </select>
                </div>
            </div>
        </div>

        <form id="wqaForm" enctype="multipart/form-data">
            <div class="form-section">
                <h3><span class="required">*</span> Crawl Data (Required)</h3>
                <div class="file-input-wrapper">
                    <div class="file-input-btn" id="crawlBtn">
                        <div>Click or drag to upload crawl file (CSV or Excel)</div>
                        <div class="file-name" id="crawlFileName"></div>
                    </div>
                    <input type="file" name="crawl_file" id="crawlFile" accept=".csv,.xlsx,.xls" required>
                </div>
                <div class="helper-text">
                    Upload a full-site crawl report with one row per URL. Recommended: Screaming Frog "Internal → Export → All" or Sitebulb "All URLs" table export.<br>
                    <strong>Required columns (or equivalents):</strong> URL, Status Code, Indexability/Meta Robots, Canonical URL, Word Count, Title, Meta Description, H1, Crawl Depth, Inlinks, Sitemap status
                </div>
            </div>

            <!-- GA4 Section with toggle -->
            <div class="form-section" id="ga4Section">
                <h3>GA4 Data <span class="optional">(Optional)</span></h3>
                <div class="data-source-toggle" id="ga4Toggle">
                    <button type="button" class="toggle-btn active" data-source="csv" onclick="toggleDataSource('ga4', 'csv')">Upload CSV</button>
                    <button type="button" class="toggle-btn" data-source="api" onclick="toggleDataSource('ga4', 'api')" id="ga4ApiToggle" disabled>Use Google API</button>
                </div>
                <div id="ga4CsvUpload">
                    <div class="file-input-wrapper">
                        <div class="file-input-btn" id="gaBtn">
                            <div>Click or drag to upload GA4 file (CSV or Excel)</div>
                            <div class="file-name" id="gaFileName"></div>
                        </div>
                        <input type="file" name="ga_file" id="gaFile" accept=".csv,.xlsx,.xls">
                    </div>
                </div>
                <div id="ga4ApiInfo" class="hidden" style="padding: 15px; background: #e8f8f0; border-radius: 8px; color: #155724;">
                    <strong>Using Google API</strong> - Data will be pulled from the selected GA4 property above.
                </div>
            </div>

            <!-- GSC Section with toggle -->
            <div class="form-section" id="gscSection">
                <h3>GSC Data <span class="optional">(Optional)</span></h3>
                <div class="data-source-toggle" id="gscToggle">
                    <button type="button" class="toggle-btn active" data-source="csv" onclick="toggleDataSource('gsc', 'csv')">Upload CSV</button>
                    <button type="button" class="toggle-btn" data-source="api" onclick="toggleDataSource('gsc', 'api')" id="gscApiToggle" disabled>Use Google API</button>
                </div>
                <div id="gscCsvUpload">
                    <div class="file-input-wrapper">
                        <div class="file-input-btn" id="gscBtn">
                            <div>Click or drag to upload GSC file (CSV or Excel)</div>
                            <div class="file-name" id="gscFileName"></div>
                        </div>
                        <input type="file" name="gsc_file" id="gscFile" accept=".csv,.xlsx,.xls">
                    </div>
                </div>
                <div id="gscApiInfo" class="hidden" style="padding: 15px; background: #e8f8f0; border-radius: 8px; color: #155724;">
                    <strong>Using Google API</strong> - Data will be pulled from the selected GSC site above.
                </div>
            </div>

            <div class="form-section">
                <h3>Backlink Data <span class="optional">(Optional)</span></h3>
                <div class="file-input-wrapper">
                    <div class="file-input-btn" id="backlinkBtn">
                        <div>Click or drag to upload backlinks file (CSV or Excel)</div>
                        <div class="file-name" id="backlinkFileName"></div>
                    </div>
                    <input type="file" name="backlink_file" id="backlinkFile" accept=".csv,.xlsx,.xls">
                </div>
                <div class="helper-text">
                    Upload per-URL backlink metrics from your link tool (Ahrefs, Semrush, Moz, Majestic).<br>
                    Recommended: "Best by links" / "Indexed pages" export.<br>
                    <strong>Required columns (or equivalents):</strong> URL, Referring Domains, Backlinks, Page Authority (UR/PA/AS)
                </div>
            </div>

            <div class="form-section">
                <h3>Keyword Tracking Data <span class="optional">(Optional)</span></h3>
                <div class="file-input-wrapper">
                    <div class="file-input-btn" id="keywordBtn">
                        <div>Click or drag to upload keyword tracking file (CSV or Excel)</div>
                        <div class="file-name" id="keywordFileName"></div>
                    </div>
                    <input type="file" name="keyword_file" id="keywordFile" accept=".csv,.xlsx,.xls">
                </div>
                <div class="helper-text">
                    Upload keyword rankings data from your rank tracker (Semrush, Ahrefs, AccuRanker, etc.).<br>
                    This populates Main KW, Best KW, and Volume columns.<br>
                    <strong>Required columns:</strong> URL/Landing Page, Keyword<br>
                    <strong>Optional columns:</strong> Volume, Position/Rank, Difficulty, CPC
                </div>
            </div>

            <div class="form-section">
                <h3>Thresholds</h3>
                <div class="threshold-grid">
                    <div class="threshold-item">
                        <label>Low Traffic (sessions)</label>
                        <input type="number" name="low_traffic_threshold" value="5" min="0">
                    </div>
                    <div class="threshold-item">
                        <label>Thin Content (words)</label>
                        <input type="number" name="thin_content_threshold" value="1000" min="0">
                    </div>
                    <div class="threshold-item">
                        <label>High Rank Max Position</label>
                        <input type="number" name="high_rank_max_position" value="20" min="1" step="0.1">
                    </div>
                    <div class="threshold-item">
                        <label>Low CTR Threshold</label>
                        <input type="number" name="low_ctr_threshold" value="0.05" min="0" max="1" step="0.01">
                    </div>
                </div>
            </div>

            <!-- Hidden fields for API data source flags -->
            <input type="hidden" name="use_ga4_api" id="useGa4Api" value="false">
            <input type="hidden" name="ga4_property_id" id="ga4PropertyId" value="">
            <input type="hidden" name="use_gsc_api" id="useGscApi" value="false">
            <input type="hidden" name="gsc_site_url" id="gscSiteUrl" value="">

            <button type="submit" class="submit-btn" id="submitBtn">Generate WQA Report</button>
        </form>

        <div id="status" class="status"></div>
    </div>

    <script>
        let isGoogleConnected = false;
        let ga4DataSource = 'csv';
        let gscDataSource = 'csv';

        // Check auth status on page load
        async function checkAuthStatus() {
            try {
                const response = await fetch('/api/auth/status');
                const data = await response.json();

                isGoogleConnected = data.connected;
                updateGoogleUI();

                if (isGoogleConnected) {
                    await loadGoogleData();
                }
            } catch (error) {
                console.error('Error checking auth status:', error);
            }
        }

        function updateGoogleUI() {
            const connectBtn = document.getElementById('googleConnectBtn');
            const status = document.getElementById('googleStatus');
            const selectors = document.getElementById('googleSelectors');
            const ga4ApiToggle = document.getElementById('ga4ApiToggle');
            const gscApiToggle = document.getElementById('gscApiToggle');

            if (isGoogleConnected) {
                connectBtn.classList.add('connected');
                connectBtn.innerHTML = '<svg viewBox="0 0 24 24" width="20" height="20"><path fill="currentColor" d="M9 16.17L4.83 12l-1.42 1.41L9 19 21 7l-1.41-1.41z"/></svg> Connected';
                status.className = 'google-status connected';
                status.innerHTML = 'Connected to Google <button class="logout-btn" onclick="logout()">Disconnect</button>';
                selectors.classList.remove('hidden');
                ga4ApiToggle.disabled = false;
                gscApiToggle.disabled = false;
            } else {
                connectBtn.classList.remove('connected');
                connectBtn.innerHTML = '<svg viewBox="0 0 24 24" width="20" height="20"><path fill="currentColor" d="M22.56 12.25c0-.78-.07-1.53-.2-2.25H12v4.26h5.92c-.26 1.37-1.04 2.53-2.21 3.31v2.77h3.57c2.08-1.92 3.28-4.74 3.28-8.09z"/></svg> Connect with Google';
                status.className = 'google-status not-connected';
                status.textContent = 'Not connected';
                selectors.classList.add('hidden');
                ga4ApiToggle.disabled = true;
                gscApiToggle.disabled = true;
            }
        }

        async function loadGoogleData() {
            // Load GA4 properties
            try {
                const ga4Response = await fetch('/api/ga4/properties');
                if (ga4Response.ok) {
                    const ga4Data = await ga4Response.json();
                    const select = document.getElementById('ga4PropertySelect');
                    select.innerHTML = '<option value="">-- Select GA4 Property --</option>';
                    ga4Data.properties.forEach(prop => {
                        select.innerHTML += `<option value="${prop.id}">${prop.name} (${prop.account})</option>`;
                    });
                }
            } catch (error) {
                console.error('Error loading GA4 properties:', error);
            }

            // Load GSC sites
            try {
                const gscResponse = await fetch('/api/gsc/sites');
                if (gscResponse.ok) {
                    const gscData = await gscResponse.json();
                    const select = document.getElementById('gscSiteSelect');
                    select.innerHTML = '<option value="">-- Select GSC Site --</option>';
                    gscData.sites.forEach(site => {
                        select.innerHTML += `<option value="${site.url}">${site.url}</option>`;
                    });
                }
            } catch (error) {
                console.error('Error loading GSC sites:', error);
            }
        }

        async function logout() {
            try {
                await fetch('/api/auth/logout', { method: 'POST' });
                isGoogleConnected = false;
                toggleDataSource('ga4', 'csv');
                toggleDataSource('gsc', 'csv');
                updateGoogleUI();
            } catch (error) {
                console.error('Error logging out:', error);
            }
        }

        function toggleDataSource(type, source) {
            const toggleBtns = document.querySelectorAll(`#${type}Toggle .toggle-btn`);
            toggleBtns.forEach(btn => {
                btn.classList.toggle('active', btn.dataset.source === source);
            });

            const csvUpload = document.getElementById(`${type}CsvUpload`);
            const apiInfo = document.getElementById(`${type}ApiInfo`);
            const useApiField = document.getElementById(`use${type.charAt(0).toUpperCase() + type.slice(1)}Api`);

            if (source === 'csv') {
                csvUpload.classList.remove('hidden');
                apiInfo.classList.add('hidden');
                useApiField.value = 'false';
            } else {
                csvUpload.classList.add('hidden');
                apiInfo.classList.remove('hidden');
                useApiField.value = 'true';
            }

            if (type === 'ga4') ga4DataSource = source;
            if (type === 'gsc') gscDataSource = source;
        }

        // File input handlers
        function setupFileInput(inputId, btnId, nameId) {
            const input = document.getElementById(inputId);
            const btn = document.getElementById(btnId);
            const nameEl = document.getElementById(nameId);

            input.addEventListener('change', function() {
                if (this.files.length > 0) {
                    nameEl.textContent = this.files[0].name;
                    btn.classList.add('has-file');
                } else {
                    nameEl.textContent = '';
                    btn.classList.remove('has-file');
                }
            });
        }

        setupFileInput('crawlFile', 'crawlBtn', 'crawlFileName');
        setupFileInput('gaFile', 'gaBtn', 'gaFileName');
        setupFileInput('gscFile', 'gscBtn', 'gscFileName');
        setupFileInput('backlinkFile', 'backlinkBtn', 'backlinkFileName');
        setupFileInput('keywordFile', 'keywordBtn', 'keywordFileName');

        // Update hidden fields when selectors change
        document.getElementById('ga4PropertySelect').addEventListener('change', function() {
            document.getElementById('ga4PropertyId').value = this.value;
        });
        document.getElementById('gscSiteSelect').addEventListener('change', function() {
            document.getElementById('gscSiteUrl').value = this.value;
        });

        // Form submission
        document.getElementById('wqaForm').addEventListener('submit', async function(e) {
            e.preventDefault();

            const submitBtn = document.getElementById('submitBtn');
            const status = document.getElementById('status');

            submitBtn.disabled = true;
            submitBtn.textContent = 'Generating Report...';
            status.className = 'status loading';
            status.textContent = 'Processing your data. This may take a moment...';

            const formData = new FormData(this);

            try {
                const response = await fetch('/api/generate', {
                    method: 'POST',
                    body: formData
                });

                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = 'wqa_report.xlsx';
                    document.body.appendChild(a);
                    a.click();
                    window.URL.revokeObjectURL(url);
                    a.remove();

                    status.className = 'status success';
                    status.textContent = 'Report generated successfully! Download should start automatically.';
                } else {
                    const error = await response.json();
                    throw new Error(error.detail || 'Failed to generate report');
                }
            } catch (error) {
                status.className = 'status error';
                status.textContent = 'Error: ' + error.message;
            } finally {
                submitBtn.disabled = false;
                submitBtn.textContent = 'Generate WQA Report';
            }
        });

        // Check for connection status on URL params
        const urlParams = new URLSearchParams(window.location.search);
        if (urlParams.get('connected') === 'google') {
            // Clear the URL param
            window.history.replaceState({}, document.title, window.location.pathname);
        }
        if (urlParams.get('error')) {
            const status = document.getElementById('status');
            status.className = 'status error';
            status.textContent = 'Google authentication failed: ' + urlParams.get('error');
            window.history.replaceState({}, document.title, window.location.pathname);
        }

        // Initialize
        checkAuthStatus();
    </script>
</body>
</html>
"""


@app.post("/api/generate")
async def generate_report(
    request: Request,
    crawl_file: UploadFile = File(...),
    ga_file: Optional[UploadFile] = File(None),
    gsc_file: Optional[UploadFile] = File(None),
    backlink_file: Optional[UploadFile] = File(None),
    keyword_file: Optional[UploadFile] = File(None),
    low_traffic_threshold: int = Form(5),
    thin_content_threshold: int = Form(1000),
    high_rank_max_position: float = Form(20.0),
    low_ctr_threshold: float = Form(0.05),
    use_ga4_api: str = Form("false"),
    ga4_property_id: str = Form(""),
    use_gsc_api: str = Form("false"),
    gsc_site_url: str = Form(""),
    google_tokens: Optional[str] = Cookie(None)
):
    """Generate WQA report from uploaded CSV/Excel files or Google API data"""
    try:
        # Read crawl file (required)
        crawl_content = await crawl_file.read()
        crawl_df = load_crawl_data(crawl_content, crawl_file.filename or "")
        logger.info(f"Loaded crawl data: {len(crawl_df)} rows")

        # Get GA4 data - either from API or file
        ga_df = None
        if use_ga4_api == "true" and ga4_property_id and google_tokens:
            tokens = decrypt_token(google_tokens)
            if tokens:
                credentials = get_credentials_from_tokens(tokens)
                ga_df = await fetch_ga4_report(credentials, ga4_property_id)
                logger.info(f"Loaded GA4 data from API: {len(ga_df)} rows")
        elif ga_file and ga_file.filename:
            ga_content = await ga_file.read()
            if ga_content:
                ga_df = load_ga_data(ga_content, ga_file.filename or "")
                logger.info(f"Loaded GA4 data from file: {len(ga_df)} rows")

        # Get GSC data - either from API or file
        gsc_df = None
        if use_gsc_api == "true" and gsc_site_url and google_tokens:
            tokens = decrypt_token(google_tokens)
            if tokens:
                credentials = get_credentials_from_tokens(tokens)
                gsc_df = await fetch_gsc_report(credentials, gsc_site_url)
                logger.info(f"Loaded GSC data from API: {len(gsc_df)} rows")
        elif gsc_file and gsc_file.filename:
            gsc_content = await gsc_file.read()
            if gsc_content:
                gsc_df = load_gsc_data(gsc_content, gsc_file.filename or "")
                logger.info(f"Loaded GSC data from file: {len(gsc_df)} rows")

        backlink_df = None
        if backlink_file and backlink_file.filename:
            backlink_content = await backlink_file.read()
            if backlink_content:
                backlink_df = load_backlink_data(backlink_content, backlink_file.filename or "")
                logger.info(f"Loaded backlink data: {len(backlink_df)} rows")

        # Load keyword tracking data
        keyword_df = None
        if keyword_file and keyword_file.filename:
            keyword_content = await keyword_file.read()
            if keyword_content:
                keyword_df = load_keyword_data(keyword_content, keyword_file.filename or "")
                if keyword_df is not None:
                    logger.info(f"Loaded keyword data: {len(keyword_df)} rows")

        # Merge datasets
        df = merge_datasets(crawl_df, ga_df, gsc_df, backlink_df)

        # Merge keyword data if available
        if keyword_df is not None and not keyword_df.empty:
            df = df.merge(keyword_df, on='url', how='left')
            logger.info(f"Merged keyword data into main DataFrame")
        logger.info(f"Merged data: {len(df)} rows")

        # Classify page types
        df['page_type'] = df['url'].apply(classify_page_type)

        # Assign actions
        actions_result = df.apply(
            lambda row: assign_actions(
                row,
                low_traffic_threshold=low_traffic_threshold,
                thin_content_threshold=thin_content_threshold,
                high_rank_max_position=high_rank_max_position,
                low_ctr_threshold=low_ctr_threshold
            ),
            axis=1,
            result_type='expand'
        )
        df['technical_actions'] = actions_result[0].apply(lambda x: ', '.join(x) if x else '')
        df['content_actions'] = actions_result[1].apply(lambda x: ', '.join(x) if x else '')

        # Assign priority
        df['priority'] = df.apply(assign_priority, axis=1)

        # Determine Final URL for redirects/merges
        df['final_url'] = df.apply(determine_final_url, axis=1)

        # Generate summary
        summaries = generate_summary(df)

        # Validate data quality
        quality_report = validate_data_quality(df, crawl_df, ga_df, gsc_df, backlink_df)

        # Log data quality warnings
        high_issues = sum(1 for i in quality_report.issues if i['severity'] == 'High')
        if high_issues > 0:
            logger.warning(f"ATTENTION: {high_issues} high-severity data quality issues found!")

        # Create Excel report with data quality sheet
        excel_bytes = create_excel_report(df, summaries, quality_report)

        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=wqa_report.xlsx"}
        )

    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "WQA Generator"}
