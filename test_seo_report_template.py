#!/usr/bin/env python3
"""
Verification script for SEO Report Template sheet in WQA Generator output.

This script verifies that the SEO Report Template sheet is correctly created
with all expected labels and structure.

Usage:
    python test_seo_report_template.py <path_to_wqa_output.xlsx>

Example:
    python test_seo_report_template.py wqa_report.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def verify_seo_report_template(excel_path: str) -> bool:
    """
    Verify that the SEO Report Template sheet exists and has correct structure.

    Args:
        excel_path: Path to the WQA output Excel file

    Returns:
        True if all verifications pass, False otherwise
    """
    errors = []
    warnings = []

    # Check file exists
    if not Path(excel_path).exists():
        print(f"ERROR: File not found: {excel_path}")
        return False

    # Load workbook (not read_only to access freeze_panes attribute)
    try:
        wb = load_workbook(excel_path, read_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}")
        return False

    # Verify sheet exists
    sheet_name = 'SEO Report Template'
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found in workbook")
        print(f"  Available sheets: {wb.sheetnames}")
        return False

    ws = wb[sheet_name]
    print(f"[OK] Sheet '{sheet_name}' exists")

    # Define expected values
    expected_values = {
        # Header fields
        ('A1', 'Company:'): "Header field 'Company:'",
        ('A2', 'Created by:'): "Header field 'Created by:'",
        ('A3', 'Month/Year:'): "Header field 'Month/Year:'",
        ('A4', 'Key Figures:'): "Header field 'Key Figures:'",

        # Section 1: Pages Report
        ('A6', 'Section 1: Pages Report'): "Section 1 title",
        ('A7', 'Page'): "Pages table header 'Page'",
        ('B7', 'URL'): "Pages table header 'URL'",
        ('C7', 'Traffic This Month'): "Pages table header 'Traffic This Month'",
        ('D7', 'Change From Last Month'): "Pages table header 'Change From Last Month'",
        ('E7', 'Changes Made to the Page'): "Pages table header 'Changes Made to the Page'",

        # Section 2: Top Keywords
        ('A35', 'Section 2: Top Keywords'): "Section 2 title",
        ('A36', 'Top Keywords'): "Keywords table header 'Top Keywords'",
        ('B36', 'Rank'): "Keywords table header 'Rank'",
        ('C36', 'Change in Rank From Last Month'): "Keywords table header 'Change in Rank'",
        ('D36', 'Volume'): "Keywords table header 'Volume'",
        ('E36', 'Difficulty (Ahrefs)'): "Keywords table header 'Difficulty (Ahrefs)'",
    }

    # Verify each expected value
    for (cell_ref, expected_value), description in expected_values.items():
        actual_value = ws[cell_ref].value
        if actual_value != expected_value:
            errors.append(f"Cell {cell_ref}: Expected '{expected_value}', got '{actual_value}' ({description})")
        else:
            print(f"[OK] {cell_ref} = '{expected_value}'")

    # Verify freeze panes
    if ws.freeze_panes != 'A8':
        warnings.append(f"Freeze panes: Expected 'A8', got '{ws.freeze_panes}'")
    else:
        print(f"[OK] Freeze panes at A8")

    # Close workbook
    wb.close()

    # Report results
    print()
    if errors:
        print("=" * 50)
        print("ERRORS:")
        for error in errors:
            print(f"  [FAIL] {error}")
        print("=" * 50)

    if warnings:
        print("WARNINGS:")
        for warning in warnings:
            print(f"  [WARN] {warning}")

    if not errors:
        print("[OK] All verifications PASSED")
        return True
    else:
        print(f"[FAIL] {len(errors)} verification(s) FAILED")
        return False


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python test_seo_report_template.py <path_to_wqa_output.xlsx>")
        print()
        print("Example:")
        print("  python test_seo_report_template.py wqa_report.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    success = verify_seo_report_template(excel_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
