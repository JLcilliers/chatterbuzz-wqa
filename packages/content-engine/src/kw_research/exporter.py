"""
Keyword Research Excel Exporter — produces a multi-sheet workbook with
cluster data, content gaps, and query patterns.
"""

import io
import logging

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_SHEET_CONFIGS = {
    "Clusters": {
        "columns": [
            "keyword", "cluster_id", "cluster_label", "intent",
            "page_type", "estimated_volume_tier", "parent_topic",
            "relevance_score",
        ],
        "widths": {
            "keyword": 35,
            "cluster_id": 12,
            "cluster_label": 28,
            "intent": 16,
            "page_type": 18,
            "estimated_volume_tier": 18,
            "parent_topic": 24,
            "relevance_score": 16,
        },
    },
    "Content Gaps": {
        "columns": [
            "query", "page", "impressions", "clicks",
            "position", "ctr", "gap_reason",
        ],
        "widths": {
            "query": 40,
            "page": 50,
            "impressions": 14,
            "clicks": 12,
            "position": 12,
            "ctr": 10,
            "gap_reason": 24,
        },
    },
    "Query Patterns": {
        "columns": [
            "pattern", "frequency", "avg_position", "avg_ctr",
            "total_impressions", "total_clicks",
        ],
        "widths": {
            "pattern": 30,
            "frequency": 12,
            "avg_position": 14,
            "avg_ctr": 12,
            "total_impressions": 18,
            "total_clicks": 14,
        },
    },
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _style_sheet(ws, config: dict) -> None:
    """Apply header formatting and column widths to a worksheet."""
    columns = config["columns"]
    widths = config["widths"]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 16)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Light borders on data rows
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)
    ):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _safe_df(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Return a copy of *df* with only the requested columns that exist."""
    present = [c for c in columns if c in df.columns]
    out = df[present].copy() if present else pd.DataFrame()
    # Add any missing columns as empty
    for c in columns:
        if c not in out.columns:
            out[c] = ""
    return out[columns]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_keyword_report(
    clusters_df: pd.DataFrame,
    gaps_df: pd.DataFrame,
    patterns_df: pd.DataFrame,
    output: io.BytesIO,
) -> bytes:
    """Create a multi-sheet Excel workbook for the keyword research report.

    Parameters
    ----------
    clusters_df : pd.DataFrame
        Keyword clusters (output of :func:`clusterer.cluster_keywords` or
        :func:`clusterer.score_business_relevance`).
    gaps_df : pd.DataFrame
        Content gap analysis (output of
        :func:`gsc_miner.find_content_gaps`).
    patterns_df : pd.DataFrame
        Query pattern data (output of
        :func:`gsc_miner.mine_keyword_patterns`).
    output : io.BytesIO
        Buffer to write the workbook into.

    Returns
    -------
    bytes
        The raw bytes of the generated ``.xlsx`` file.
    """
    sheets = {
        "Clusters": _safe_df(clusters_df, _SHEET_CONFIGS["Clusters"]["columns"]),
        "Content Gaps": _safe_df(gaps_df, _SHEET_CONFIGS["Content Gaps"]["columns"]),
        "Query Patterns": _safe_df(
            patterns_df, _SHEET_CONFIGS["Query Patterns"]["columns"]
        ),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_sheet(ws, _SHEET_CONFIGS[sheet_name])

    output.seek(0)
    result_bytes = output.read()

    logger.info(
        "Exported keyword report: %d clusters, %d gaps, %d patterns (%.1f KB).",
        len(clusters_df),
        len(gaps_df),
        len(patterns_df),
        len(result_bytes) / 1024,
    )
    return result_bytes
