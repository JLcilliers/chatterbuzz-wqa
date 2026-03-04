"""
Excel report generation for the WQA Engine.

Ported from api/index.py — contains create_excel_report and all helper functions
related to Excel report generation (analytical sheets, data quality sheets, etc.).
"""

import io
import logging
import math
import re
from collections import defaultdict
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# =============================================================================
# DATA QUALITY REPORT
# =============================================================================

class DataQualityReport:
    """Container for data quality validation results."""

    def __init__(self):
        self.missing_columns: Dict[str, List[str]] = {}
        self.empty_columns: List[str] = []
        self.sparse_columns: Dict[str, float] = {}
        self.url_match_stats: Dict[str, Dict] = {}
        self.data_source_coverage: Dict[str, int] = {}
        self.issues: List[Dict] = []
        self.warnings: List[str] = []

    def add_issue(self, category: str, severity: str, description: str,
                  affected_count: int = 0, examples: List[str] = None):
        """Add a specific data quality issue."""
        self.issues.append({
            'category': category,
            'severity': severity,
            'description': description,
            'affected_count': affected_count,
            'examples': examples or []
        })

    def to_dataframe(self) -> pd.DataFrame:
        """Convert issues to a DataFrame for Excel export."""
        if not self.issues:
            return pd.DataFrame(columns=['Category', 'Severity', 'Description', 'Affected Count', 'Examples'])

        rows = []
        for issue in self.issues:
            rows.append({
                'Category': issue['category'],
                'Severity': issue['severity'],
                'Description': issue['description'],
                'Affected Count': issue['affected_count'],
                'Examples': '; '.join(str(e) for e in issue['examples'][:5]) if issue['examples'] else ''
            })

        return pd.DataFrame(rows)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def extract_page_path(url: str) -> str:
    """Extract page path from URL (no protocol, no domain, no querystring)"""
    if pd.isna(url) or not url:
        return '/'
    try:
        parsed = urlparse(str(url))
        path = parsed.path if parsed.path else '/'
        # Ensure path starts with /
        if not path.startswith('/'):
            path = '/' + path
        return path
    except:
        return '/'


def format_indexability(row) -> Tuple[str, str]:
    """
    Return (Index/Noindex label, Indexation Status explanation).

    Index/Noindex: Simple binary status
    Indexation Status: WHY the page is not indexable (reason), not just repeating the label
    """
    indexable = row.get('indexable', True)
    meta_robots = str(row.get('meta_robots', '')).lower()
    status_code = int(row.get('status_code', 200))
    canonical = str(row.get('canonical_url', '')).strip()
    url = str(row.get('url', '')).strip().lower()

    # Determine index/noindex label and specific reason
    if status_code >= 500:
        index_label = 'Non-Indexable'
        status_reason = 'Server Error'
    elif status_code == 404:
        index_label = 'Non-Indexable'
        status_reason = 'Not Found (404)'
    elif status_code in [301, 302, 307, 308]:
        index_label = 'Non-Indexable'
        status_reason = 'Redirected'
    elif status_code >= 400:
        index_label = 'Non-Indexable'
        status_reason = 'Client Error'
    elif 'noindex' in meta_robots:
        index_label = 'Non-Indexable'
        status_reason = 'Noindex Tag'
    elif 'nofollow' in meta_robots and 'noindex' not in meta_robots:
        # Nofollow but not noindex - still indexable
        index_label = 'Indexable'
        status_reason = 'Indexable (nofollow)'
    elif canonical and canonical.lower() != url:
        index_label = 'Non-Indexable'
        status_reason = 'Canonicalised'
    elif not indexable:
        index_label = 'Non-Indexable'
        status_reason = 'Blocked'  # Generic blocked reason
    else:
        index_label = 'Indexable'
        status_reason = 'OK'  # Clear signal that page is fine

    return index_label, status_reason


def generate_data_quality_sheets(report: DataQualityReport) -> Dict[str, pd.DataFrame]:
    """Generate DataFrames for data quality reporting in Excel."""
    sheets = {}

    sheets['issues'] = report.to_dataframe()

    coverage_data = []
    for source, count in report.data_source_coverage.items():
        stats = report.url_match_stats.get(source, {})
        coverage_data.append({
            'Data Source': source,
            'URLs Matched': count,
            'Total Source URLs': stats.get('total_source_urls', 'N/A'),
            'Match Rate (%)': f"{stats.get('match_rate', 0):.1f}" if stats else 'N/A',
            'Unmatched URLs': stats.get('unmatched', 0)
        })
    sheets['coverage'] = pd.DataFrame(coverage_data)

    sparsity_data = []
    for col, pct in sorted(report.sparse_columns.items(), key=lambda x: -x[1]):
        sparsity_data.append({
            'Column': col,
            'Empty/Default (%)': f"{pct:.1f}",
            'Status': 'Critical' if pct >= 90 else ('Warning' if pct >= 70 else 'OK')
        })
    sheets['column_quality'] = pd.DataFrame(sparsity_data)

    return sheets


# =============================================================================
# ANALYTICAL SHEETS BUILDER FUNCTIONS
# =============================================================================

def build_content_to_optimize_api(df: pd.DataFrame) -> pd.DataFrame:
    """Build Content to Optimize sheet - API version."""
    mask = (df['status_code'] == 200) & (df['indexable'] == True)
    candidates = df[mask].copy()

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Avg Position', 'Impressions', 'Clicks', 'CTR (%)',
            'Word Count', 'Has Meta Description', 'Sessions', 'Optimization Signals', 'Priority Score'
        ])

    optimization_signals = []
    priority_scores = []

    for _, row in candidates.iterrows():
        signals = []
        score = 0

        if row['impressions'] >= 100 and row['ctr'] < 0.05:
            signals.append('Low CTR despite impressions')
            score += 3
        elif row['impressions'] >= 50 and row['ctr'] < 0.03:
            signals.append('Very low CTR')
            score += 4

        if 5 <= row['avg_position'] <= 10:
            signals.append('Position 5-10 (near page 1)')
            score += 5
        elif 11 <= row['avg_position'] <= 20:
            signals.append('Position 11-20 (striking distance)')
            score += 3

        meta_desc = str(row.get('meta_description', '')).strip()
        if not meta_desc:
            signals.append('Missing meta description')
            score += 2
        elif len(meta_desc) < 70:
            signals.append('Short meta description')
            score += 1

        if row['word_count'] < 500:
            signals.append('Very thin content (<500 words)')
            score += 3
        elif row['word_count'] < 1000:
            signals.append('Thin content (<1000 words)')
            score += 2

        if row['sessions'] > 0 and row['avg_position'] > 5:
            signals.append('Has traffic, room to improve')
            score += 1

        if row['impressions'] >= 500:
            signals.append('High search demand')
            score += 2
        elif row['impressions'] >= 100:
            signals.append('Moderate search demand')
            score += 1

        optimization_signals.append('; '.join(signals) if signals else 'No specific signals')
        priority_scores.append(score)

    candidates['Optimization Signals'] = optimization_signals
    candidates['Priority Score'] = priority_scores
    candidates = candidates[candidates['Priority Score'] > 0]

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Avg Position', 'Impressions', 'Clicks', 'CTR (%)',
            'Word Count', 'Has Meta Description', 'Sessions', 'Optimization Signals', 'Priority Score'
        ])

    # Sort by priority score descending and cap at top 20
    candidates = candidates.sort_values('Priority Score', ascending=False).head(20)

    return pd.DataFrame({
        'URL': candidates['url'],
        'Page Type': candidates['page_type'],
        'Avg Position': candidates['avg_position'].round(1),
        'Impressions': candidates['impressions'].astype(int),
        'Clicks': candidates['clicks'].astype(int),
        'CTR (%)': (candidates['ctr'] * 100).round(2),
        'Word Count': candidates['word_count'].astype(int),
        'Has Meta Description': candidates['meta_description'].apply(lambda x: 'Yes' if str(x).strip() else 'No'),
        'Sessions': candidates['sessions'].astype(int),
        'Optimization Signals': candidates['Optimization Signals'],
        'Priority Score': candidates['Priority Score']
    })


def build_thin_content_api(df: pd.DataFrame, thin_threshold: int = 1000) -> pd.DataFrame:
    """Build Thin Content Opportunities sheet - API version."""
    mask = (df['status_code'] == 200) & (df['indexable'] == True) & (df['word_count'] < thin_threshold)
    candidates = df[mask].copy()

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Word Count', 'Content Gap', 'Sessions',
            'Impressions', 'Avg Position', 'Referring Domains', 'Opportunity Type', 'Value Score'
        ])

    opportunity_types = []
    value_scores = []

    for _, row in candidates.iterrows():
        opp_type = []
        score = 0

        if row['sessions'] >= 10:
            opp_type.append('Has significant traffic')
            score += 5
        elif row['sessions'] > 0:
            opp_type.append('Has some traffic')
            score += 2

        if row['impressions'] >= 100:
            opp_type.append('High search visibility')
            score += 4
        elif row['impressions'] > 0:
            opp_type.append('Some search visibility')
            score += 1

        if 0 < row['avg_position'] <= 20:
            opp_type.append('Already ranking (pos <= 20)')
            score += 3
        elif 20 < row['avg_position'] <= 50:
            opp_type.append('Has rankings (pos 20-50)')
            score += 1

        if row['referring_domains'] >= 5:
            opp_type.append('Strong backlink profile')
            score += 4
        elif row['referring_domains'] > 0:
            opp_type.append('Has backlinks')
            score += 2

        if row['page_type'] in ['Service', 'Local Lander', 'Home']:
            opp_type.append(f'Strategic page ({row["page_type"]})')
            score += 3

        if row['word_count'] < 300:
            opp_type.append('Severely thin (<300 words)')
        elif row['word_count'] < 500:
            opp_type.append('Very thin (<500 words)')

        opportunity_types.append('; '.join(opp_type) if opp_type else 'Low priority')
        value_scores.append(score)

    candidates['Opportunity Type'] = opportunity_types
    candidates['Value Score'] = value_scores
    candidates = candidates[candidates['Value Score'] > 0]

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Word Count', 'Content Gap', 'Sessions',
            'Impressions', 'Avg Position', 'Referring Domains', 'Opportunity Type', 'Value Score'
        ])

    candidates = candidates.sort_values('Value Score', ascending=False)
    content_gap = thin_threshold - candidates['word_count']

    return pd.DataFrame({
        'URL': candidates['url'],
        'Page Type': candidates['page_type'],
        'Word Count': candidates['word_count'].astype(int),
        'Content Gap': content_gap.astype(int),
        'Sessions': candidates['sessions'].astype(int),
        'Impressions': candidates['impressions'].astype(int),
        'Avg Position': candidates['avg_position'].round(1),
        'Referring Domains': candidates['referring_domains'].astype(int),
        'Opportunity Type': candidates['Opportunity Type'],
        'Value Score': candidates['Value Score']
    })


# =============================================================================
# BUSINESS RELEVANCE & FEASIBILITY FILTERING (API VERSION)
# =============================================================================
# These functions filter and score topics to ensure only realistic, sensible,
# on-brand opportunities are surfaced. A topic must pass all three gates:
# 1. Search Demand Gate (handled by clustering)
# 2. SEO Feasibility Gate (handled by action recommendation)
# 3. Business Relevance Gate (NEW - this module)

# Exclusion terms - topics containing these are automatically excluded
NAVIGATIONAL_TERMS_API = {
    'login', 'signin', 'sign in', 'signup', 'sign up', 'logout', 'log out',
    'register', 'forgot password', 'reset password', 'my account', 'dashboard',
    'admin', 'portal', 'intranet', 'employee', 'staff only'
}

POLICY_LEGAL_TERMS_API = {
    'privacy policy', 'terms of service', 'terms and conditions', 'cookie policy',
    'gdpr', 'disclaimer', 'legal notice', 'refund policy', 'return policy',
    'shipping policy', 'cancellation policy', 'accessibility statement'
}

JUNK_ARTIFACT_TERMS_API = {
    'utm_', 'utm=', '?p=', '?s=', 'index.php', 'page=', 'amp', '&amp',
    'session', 'sessionid', 'jsessionid', 'phpsessid', 'cfid', 'cftoken',
    'www.', 'http://', 'https://', '.html', '.php', '.aspx', '.jsp'
}

AUTHORITY_REQUIRED_TERMS_API = {
    # Medical - requires licensed practitioner
    'diagnosis', 'treatment plan', 'medical advice', 'prescription', 'dosage',
    'symptoms of', 'cure for', 'medication for', 'drug interaction',
    # Legal - requires licensed attorney
    'legal advice', 'attorney', 'lawsuit', 'sue for', 'liability for',
    'contract law', 'legal rights', 'file a claim',
    # Financial - requires licensed advisor
    'investment advice', 'tax advice', 'financial planning', 'retirement planning',
    'stock picks', 'crypto investment', 'guaranteed returns',
    # Government - requires official authority
    'official form', 'government application', 'visa application', 'passport renewal'
}

# Intent classification modifiers
TRANSACTIONAL_MODIFIERS_API = {
    'pricing', 'price', 'cost', 'quote', 'buy', 'purchase', 'order', 'hire',
    'service', 'services', 'provider', 'providers', 'company', 'companies',
    'contractor', 'contractors', 'professional', 'professionals', 'near me',
    'in my area', 'local', 'get', 'schedule', 'book', 'appointment'
}

COMMERCIAL_MODIFIERS_API = {
    'best', 'top', 'review', 'reviews', 'compare', 'comparison', 'vs', 'versus',
    'alternative', 'alternatives', 'software', 'tool', 'tools', 'solution',
    'solutions', 'platform', 'app', 'rated', 'recommended', 'affordable'
}

INFORMATIONAL_MODIFIERS_API = {
    'how', 'what', 'why', 'when', 'where', 'guide', 'tutorial', 'tips',
    'steps', 'process', 'explained', 'meaning', 'definition', 'example',
    'examples', 'learn', 'understand', 'basics', 'introduction', 'overview'
}


def calculate_business_relevance_score_api(
    topic_tokens: set,
    primary_query: str,
    site_context: dict = None
) -> tuple:
    """
    Calculate Business Relevance Score (0-100) for a topic cluster - API version.
    """
    if site_context is None:
        site_context = {}

    query_lower = primary_query.lower()
    notes = []
    exclusion_reason = ""

    # HARD EXCLUSION CHECKS
    for term in NAVIGATIONAL_TERMS_API:
        if term in query_lower:
            return (0, "", True, f"Navigational intent: '{term}'")

    for term in POLICY_LEGAL_TERMS_API:
        if term in query_lower:
            return (0, "", True, f"Policy/legal page: '{term}'")

    for term in JUNK_ARTIFACT_TERMS_API:
        if term in query_lower:
            return (0, "", True, f"Junk/artifact term: '{term}'")

    for term in AUTHORITY_REQUIRED_TERMS_API:
        if term in query_lower:
            return (0, "", True, f"Requires professional authority: '{term}'")

    clean_query = ''.join(c for c in query_lower if c.isalnum() or c.isspace())
    if len(clean_query.strip()) < 5:
        return (0, "", True, "Query too short/malformed")

    if clean_query.replace(' ', '').isdigit():
        return (0, "", True, "Query is only numbers")

    # COMPONENT 1: Site Theme Alignment (40%)
    theme_score = 0
    max_theme_score = 40

    existing_urls = set(str(u).lower() for u in site_context.get('existing_urls', []))
    existing_titles = set(str(t).lower() for t in site_context.get('existing_titles', []))
    top_queries = set(str(q).lower() for q in site_context.get('top_queries', []))
    site_categories = set(str(c).lower() for c in site_context.get('site_categories', []))

    all_site_text = ' '.join(existing_urls) + ' ' + ' '.join(existing_titles) + ' ' + ' '.join(top_queries)
    site_tokens = set(all_site_text.split())

    if topic_tokens and site_tokens:
        overlap = len(topic_tokens & site_tokens)
        overlap_ratio = overlap / len(topic_tokens) if topic_tokens else 0

        if overlap_ratio >= 0.6:
            theme_score = max_theme_score
        elif overlap_ratio >= 0.4:
            theme_score = int(max_theme_score * 0.75)
        elif overlap_ratio >= 0.2:
            theme_score = int(max_theme_score * 0.5)
        elif overlap_ratio > 0:
            theme_score = int(max_theme_score * 0.25)
        else:
            theme_score = 0
            notes.append("No site theme overlap")

    for category in site_categories:
        if category in query_lower:
            theme_score = min(max_theme_score, theme_score + 10)
            break

    # COMPONENT 2: Commercial/Strategic Intent (30%)
    intent_score = 0
    max_intent_score = 30

    transactional_matches = sum(1 for term in TRANSACTIONAL_MODIFIERS_API if term in query_lower)
    commercial_matches = sum(1 for term in COMMERCIAL_MODIFIERS_API if term in query_lower)
    informational_matches = sum(1 for term in INFORMATIONAL_MODIFIERS_API if term in query_lower)

    if transactional_matches > 0:
        intent_score = max_intent_score
    elif commercial_matches > 0:
        intent_score = int(max_intent_score * 0.7)
    elif informational_matches > 0:
        if theme_score >= max_theme_score * 0.5:
            intent_score = int(max_intent_score * 0.4)
        else:
            intent_score = int(max_intent_score * 0.15)
            notes.append("Informational intent with weak commercial tie-in")
    else:
        intent_score = int(max_intent_score * 0.3)

    # COMPONENT 3: Content Feasibility (20%)
    feasibility_score = 20
    questionable_expertise = {
        'diy', 'homemade', 'home remedy', 'self-diagnose', 'without professional',
        'instead of doctor', 'instead of lawyer', 'free legal', 'free medical'
    }

    for term in questionable_expertise:
        if term in query_lower:
            feasibility_score = int(feasibility_score * 0.5)
            notes.append(f"Questionable expertise required: '{term}'")
            break

    # COMPONENT 4: Brand Safety / Nonsense Filter (10%)
    safety_score = 10

    if len(query_lower.split()) > 8:
        safety_score = int(safety_score * 0.7)
        notes.append("Query unusually long")

    word_list = query_lower.split()
    if len(word_list) > len(set(word_list)) + 2:
        safety_score = int(safety_score * 0.5)
        notes.append("Repeated words in query")

    # CALCULATE FINAL SCORE
    total_score = theme_score + intent_score + feasibility_score + safety_score
    total_score = max(0, min(100, total_score))

    if total_score < 50:
        exclusion_reason = f"Low business relevance score ({total_score}/100)"
        return (total_score, '; '.join(notes) if notes else '', True, exclusion_reason)

    return (total_score, '; '.join(notes) if notes else '', False, '')


def build_site_context_from_df_api(df: pd.DataFrame, gsc_df: pd.DataFrame = None) -> dict:
    """
    Build site context dictionary from crawl and GSC data - API version.
    """
    context = {
        'existing_urls': [],
        'existing_titles': [],
        'top_queries': [],
        'site_categories': [],
        'brand_terms': []
    }

    if df is not None and len(df) > 0:
        if 'url' in df.columns:
            context['existing_urls'] = df['url'].dropna().tolist()

        if 'page_title' in df.columns:
            context['existing_titles'] = df['page_title'].dropna().tolist()

        if 'page_type' in df.columns:
            categories = df['page_type'].dropna().unique().tolist()
            for cat in categories:
                context['site_categories'].extend(str(cat).lower().split())
            context['site_categories'] = list(set(context['site_categories']))

    if gsc_df is not None and len(gsc_df) > 0:
        query_col = 'query' if 'query' in gsc_df.columns else 'primary_keyword'
        if query_col in gsc_df.columns and 'impressions' in gsc_df.columns:
            top_queries_df = gsc_df.nlargest(100, 'impressions')
            context['top_queries'] = top_queries_df[query_col].dropna().tolist()
        elif query_col in gsc_df.columns:
            context['top_queries'] = gsc_df[query_col].dropna().head(100).tolist()

    if context['existing_urls']:
        try:
            sample_url = context['existing_urls'][0]
            parsed = urlparse(sample_url)
            domain_parts = parsed.netloc.replace('www.', '').split('.')
            if domain_parts:
                context['brand_terms'] = [domain_parts[0]]
        except:
            pass

    return context


def filter_topics_by_business_relevance_api(
    topics_df: pd.DataFrame,
    site_context: dict = None,
    min_score: int = 50,
    max_topics: int = 15
) -> tuple:
    """
    Filter and score topics by business relevance - API version.
    """
    if topics_df is None or len(topics_df) == 0:
        return topics_df, []

    if site_context is None:
        site_context = {}

    relevance_scores = []
    feasibility_notes = []
    excluded_topics = []

    for _, row in topics_df.iterrows():
        primary_query = str(row.get('Primary Keyword', '')).strip()
        secondary_keywords = str(row.get('Secondary Keywords', '')).strip()

        all_keywords = primary_query + ' ' + secondary_keywords
        tokens = set(all_keywords.lower().split())

        score, notes, excluded, reason = calculate_business_relevance_score_api(
            topic_tokens=tokens,
            primary_query=primary_query,
            site_context=site_context
        )

        if excluded:
            excluded_topics.append({
                'Topic': row.get('Suggested Topic', primary_query),
                'Primary Keyword': primary_query,
                'Score': score,
                'Reason': reason
            })
            relevance_scores.append(0)
            feasibility_notes.append(reason)
        else:
            relevance_scores.append(score)
            feasibility_notes.append(notes if notes else '')

    topics_df = topics_df.copy()
    topics_df['Business Relevance Score'] = relevance_scores
    topics_df['Feasibility Notes'] = feasibility_notes

    mask = topics_df['Business Relevance Score'] >= min_score
    filtered_df = topics_df[mask].copy()

    if len(filtered_df) > 0:
        filtered_df = filtered_df.sort_values(
            ['Priority Score', 'Business Relevance Score'],
            ascending=[False, False]
        )
        filtered_df = filtered_df.head(max_topics)

    logger.info(f"[BUSINESS RELEVANCE] Topics before filtering: {len(topics_df)}")
    logger.info(f"[BUSINESS RELEVANCE] Topics after filtering: {len(filtered_df)}")
    logger.info(f"[BUSINESS RELEVANCE] Topics excluded: {len(excluded_topics)}")

    if excluded_topics:
        exclusion_reasons = {}
        for ex in excluded_topics:
            reason = ex['Reason'].split(':')[0] if ':' in ex['Reason'] else ex['Reason']
            exclusion_reasons[reason] = exclusion_reasons.get(reason, 0) + 1
        logger.info(f"[BUSINESS RELEVANCE] Exclusion breakdown: {exclusion_reasons}")

    return filtered_df, excluded_topics


def cluster_queries_into_topics_api(queries_df: pd.DataFrame, max_topics: int = 20, thin_urls: set = None, url_word_counts: dict = None) -> pd.DataFrame:
    """
    Cluster similar queries into topic groups for new content opportunities.
    API version - uses word-overlap algorithm with Content Action Recommendations.

    Implements Create/Expand/Consolidate decision logic:
    - Create new page: No dominant page, content gap exists
    - Expand existing page: One page ranks but needs depth/optimization
    - Consolidate pages: Multiple pages cannibalize each other
    """
    if queries_df is None or len(queries_df) == 0:
        return pd.DataFrame()

    if thin_urls is None:
        thin_urls = set()

    if url_word_counts is None:
        url_word_counts = {}

    def tokenize(query):
        query = str(query).lower().strip()
        stopwords = {'a', 'an', 'the', 'in', 'on', 'at', 'to', 'for', 'of', 'and', 'or', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'shall', 'can', 'need', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'with', 'as', 'by', 'from', 'how', 'when', 'where', 'why', 'if', 'then', 'so', 'than', 'too', 'very', 'just', 'only', 'also', 'even', 'more', 'most', 'other', 'into', 'over', 'after', 'before', 'between', 'through', 'during', 'under', 'around', 'about', 'near'}
        words = re.findall(r'\b[a-z]{2,}\b', query)
        return [w for w in words if w not in stopwords]

    query_data = []
    for _, row in queries_df.iterrows():
        query = str(row.get('query', row.get('primary_keyword', ''))).strip()
        if not query or query == 'nan':
            continue
        tokens = tokenize(query)
        if tokens:
            query_data.append({
                'query': query, 'tokens': set(tokens),
                'impressions': int(row.get('impressions', 0)),
                'clicks': int(row.get('clicks', 0)),
                'avg_position': float(row.get('avg_position', 0)),
                'url': str(row.get('url', ''))
            })

    if not query_data:
        return pd.DataFrame()

    query_data.sort(key=lambda x: x['impressions'], reverse=True)

    clusters = []
    used_queries = set()

    for seed in query_data:
        if seed['query'] in used_queries:
            continue

        cluster = {
            'primary_query': seed['query'], 'queries': [seed],
            'all_tokens': seed['tokens'].copy(),
            'total_impressions': seed['impressions'], 'total_clicks': seed['clicks'],
            'positions': [seed['avg_position']] if seed['avg_position'] > 0 else [],
            'urls': defaultdict(lambda: {'impressions': 0, 'positions': [], 'clicks': 0})
        }
        # Track URL-level metrics for cannibalization detection
        if seed['url']:
            cluster['urls'][seed['url']]['impressions'] += seed['impressions']
            cluster['urls'][seed['url']]['clicks'] += seed['clicks']
            if seed['avg_position'] > 0:
                cluster['urls'][seed['url']]['positions'].append(seed['avg_position'])
        used_queries.add(seed['query'])

        for candidate in query_data:
            if candidate['query'] in used_queries:
                continue
            overlap = len(seed['tokens'] & candidate['tokens'])
            min_len = min(len(seed['tokens']), len(candidate['tokens']))
            if min_len > 0 and overlap / min_len >= 0.5:
                cluster['queries'].append(candidate)
                cluster['all_tokens'].update(candidate['tokens'])
                cluster['total_impressions'] += candidate['impressions']
                cluster['total_clicks'] += candidate['clicks']
                if candidate['avg_position'] > 0:
                    cluster['positions'].append(candidate['avg_position'])
                # Track URL-level metrics
                if candidate['url']:
                    cluster['urls'][candidate['url']]['impressions'] += candidate['impressions']
                    cluster['urls'][candidate['url']]['clicks'] += candidate['clicks']
                    if candidate['avg_position'] > 0:
                        cluster['urls'][candidate['url']]['positions'].append(candidate['avg_position'])
                used_queries.add(candidate['query'])

        clusters.append(cluster)

    topic_opportunities = []
    for cluster in clusters:
        if len(cluster['queries']) == 1 and cluster['total_impressions'] < 100:
            continue

        if cluster['positions']:
            weighted_pos = sum(q['avg_position'] * q['impressions'] for q in cluster['queries'] if q['avg_position'] > 0)
            total_weight = sum(q['impressions'] for q in cluster['queries'] if q['avg_position'] > 0)
            avg_position = weighted_pos / total_weight if total_weight > 0 else 0
        else:
            avg_position = 0

        # =====================================================================
        # URL-LEVEL METRICS ANALYSIS
        # =====================================================================
        url_metrics = cluster['urls']
        num_urls = len(url_metrics)

        # Calculate per-URL metrics with CTR
        url_analysis = []
        for url, metrics in url_metrics.items():
            url_impressions = metrics['impressions']
            url_clicks = metrics['clicks']
            url_positions = metrics['positions']
            url_avg_pos = sum(url_positions) / len(url_positions) if url_positions else 0
            url_ctr = (url_clicks / url_impressions * 100) if url_impressions > 0 else 0
            url_share = url_impressions / cluster['total_impressions'] if cluster['total_impressions'] > 0 else 0
            url_word_count = url_word_counts.get(url, 0)
            is_thin = url in thin_urls or url_word_count < 1000

            url_analysis.append({
                'url': url, 'impressions': url_impressions, 'clicks': url_clicks,
                'avg_position': url_avg_pos, 'ctr': url_ctr, 'share': url_share,
                'word_count': url_word_count, 'is_thin': is_thin
            })

        # Sort by impressions (strongest first)
        url_analysis.sort(key=lambda x: x['impressions'], reverse=True)

        # Identify dominant page (if any)
        dominant_url = None
        dominant_data = None
        if url_analysis:
            top_url = url_analysis[0]
            if top_url['share'] >= 0.60 and top_url['avg_position'] <= 15:
                if not top_url['is_thin']:
                    dominant_url = top_url['url']
                    dominant_data = top_url

        # =====================================================================
        # CONTENT ACTION RECOMMENDATION LOGIC
        # =====================================================================
        recommended_action = ""
        primary_url = ""
        secondary_urls = ""
        reasoning = ""

        # Track clusters with dominant pages for debugging (but DO NOT exclude them)
        if dominant_url and dominant_data:
            logger.debug(f"[CLUSTER DEBUG] Topic '{cluster['primary_query']}' has dominant page {dominant_url} "
                        f"({dominant_data['share']:.0%} impressions at position {dominant_data['avg_position']:.1f}) - evaluating for action")
            # NOTE: Previously this would `continue` and exclude the cluster.
            # Now we flow through to assign an appropriate action (usually Expand existing page)

        # DECISION LOGIC
        if num_urls == 0:
            recommended_action = "Create new page"
            primary_url = ""
            secondary_urls = ""
            reasoning = f"Search demand exists ({cluster['total_impressions']:,} impressions) but no landing page currently serves these queries. This is a true content gap."

        elif num_urls == 1:
            single_url = url_analysis[0]

            if single_url['avg_position'] > 20:
                recommended_action = "Create new page"
                primary_url = ""
                secondary_urls = ""
                reasoning = f"Existing page ({single_url['url'][:60]}...) ranks poorly at position {single_url['avg_position']:.0f}. A dedicated, well-optimized page would better serve this topic."

            elif single_url['is_thin']:
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page exists but has thin content ({single_url['word_count']} words). Expanding with comprehensive content at position {single_url['avg_position']:.0f} can capture more traffic."

            elif single_url['ctr'] < 2.0 and single_url['impressions'] >= 100:
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} with high impressions ({single_url['impressions']:,}) but very low CTR ({single_url['ctr']:.1f}%). Improve content depth and meta data to boost clicks."

            elif single_url['avg_position'] > 10:
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} but not in top 10. Expanding content depth can improve rankings and capture more traffic."

            else:
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} but doesn't fully capture the topic. Expanding with related subtopics can increase topical authority."

        elif num_urls == 2:
            url1, url2 = url_analysis[0], url_analysis[1]

            if url1['share'] < 0.65 and url2['share'] >= 0.10:
                recommended_action = "Consolidate pages"
                primary_url = url1['url']
                secondary_urls = url2['url']
                reasoning = f"Two pages are competing for the same keywords, diluting rankings. {url1['url'][:50]}... has {url1['share']:.0%} impressions at position {url1['avg_position']:.0f}; {url2['url'][:50]}... has {url2['share']:.0%}. Consolidate into one authoritative page."
            else:
                recommended_action = "Expand existing page"
                primary_url = url1['url']
                secondary_urls = ""
                reasoning = f"One page ({url1['url'][:50]}...) receives most impressions ({url1['share']:.0%}) at position {url1['avg_position']:.0f}. Expand this page to fully own the topic."

        else:
            # 3+ URLs - Consolidation needed
            top_urls = url_analysis[:3]
            meaningful_urls = [u for u in top_urls if u['share'] >= 0.08]

            if len(meaningful_urls) >= 2:
                recommended_action = "Consolidate pages"
                primary_url = meaningful_urls[0]['url']
                secondary_urls = ', '.join([u['url'] for u in meaningful_urls[1:]])
                url_summary = '; '.join([f"{u['url'][:40]}... ({u['share']:.0%})" for u in meaningful_urls])
                reasoning = f"Traffic is fragmented across {num_urls} pages causing keyword cannibalization. Top competing pages: {url_summary}. Consolidate into one authoritative hub."
            elif url_analysis[0]['avg_position'] > 20:
                recommended_action = "Create new page"
                primary_url = ""
                secondary_urls = ""
                reasoning = f"Traffic is split across {num_urls} weak pages (best position: {url_analysis[0]['avg_position']:.0f}). Creating a dedicated, comprehensive page would better serve this topic than consolidating poor performers."
            else:
                recommended_action = "Expand existing page"
                primary_url = url_analysis[0]['url']
                secondary_urls = ""
                reasoning = f"Multiple pages exist but {url_analysis[0]['url'][:50]}... leads with {url_analysis[0]['share']:.0%} impressions. Expand this page to consolidate authority."

        # Generate suggested topic and page type
        primary_query = cluster['primary_query']
        common_tokens = sorted(cluster['all_tokens'], key=lambda t: sum(1 for q in cluster['queries'] if t in q['tokens']), reverse=True)
        suggested_topic = ' '.join(common_tokens[:4]).title() if common_tokens else primary_query.title()

        query_lower = primary_query.lower()
        intent_value = 0.5
        if any(word in query_lower for word in ['how to', 'guide', 'tutorial', 'tips', 'steps']):
            page_type = 'Guide / How-To'
            intent_value = 0.7
        elif any(word in query_lower for word in ['best', 'top', 'review', 'compare', 'vs']):
            page_type = 'Comparison / Review'
            intent_value = 0.9
        elif any(word in query_lower for word in ['what is', 'meaning', 'definition', 'explain']):
            page_type = 'Educational / Explainer'
            intent_value = 0.6
        elif any(word in query_lower for word in ['near me', 'in ', 'local']):
            page_type = 'Local Landing Page'
            intent_value = 1.0
        elif any(word in query_lower for word in ['cost', 'price', 'pricing', 'quote', 'estimate']):
            page_type = 'Service / Pricing Page'
            intent_value = 1.0
        elif any(word in query_lower for word in ['buy', 'order', 'purchase', 'shop']):
            page_type = 'Product / Service Page'
            intent_value = 1.0
        else:
            page_type = 'Blog Post / Article'
            intent_value = 0.5

        secondary = [q['query'] for q in cluster['queries'] if q['query'] != primary_query][:8]

        # =====================================================================
        # PRIORITY SCORE CALCULATION
        # 40% Impressions, 25% Action urgency, 20% Position weakness, 15% Intent value
        # =====================================================================
        impressions_score = min(10, math.log10(max(1, cluster['total_impressions'])) * 2.5)

        # Action urgency score (25% weight)
        if recommended_action == "Create new page":
            action_urgency_score = 10 if num_urls == 0 else 7
        elif recommended_action == "Consolidate pages":
            action_urgency_score = 9  # High urgency - cannibalization hurts rankings
        else:  # Expand
            action_urgency_score = 5  # Lower urgency - existing page works somewhat

        # Position weakness score (20% weight)
        if avg_position == 0 or num_urls == 0:
            position_score = 10
        elif avg_position > 30:
            position_score = 9
        elif avg_position > 20:
            position_score = 7
        elif avg_position > 15:
            position_score = 5
        elif avg_position > 10:
            position_score = 3
        else:
            position_score = 1

        intent_score = intent_value * 10

        priority_score = round(
            (impressions_score * 0.40) +
            (action_urgency_score * 0.25) +
            (position_score * 0.20) +
            (intent_score * 0.15),
            1
        )

        topic_opportunities.append({
            'Suggested Topic': suggested_topic,
            'Primary Keyword': primary_query,
            'Secondary Keywords': ', '.join(secondary) if secondary else '',
            'Total Impressions': cluster['total_impressions'],
            'Avg Position': round(avg_position, 1) if avg_position > 0 else 'N/A',
            'Recommended Action': recommended_action,
            'Primary URL': primary_url,
            'Secondary URLs': secondary_urls,
            'Reasoning': reasoning,
            'Suggested Page Type': page_type,
            'Priority Score': priority_score
        })

    if not topic_opportunities:
        return pd.DataFrame()

    result = pd.DataFrame(topic_opportunities)
    return result.sort_values('Priority Score', ascending=False).head(max_topics)


def build_new_content_opportunities_api(df: pd.DataFrame, gsc_df: Optional[pd.DataFrame] = None, thin_threshold: int = 1000) -> pd.DataFrame:
    """
    Build New Content Opportunities sheet - API version.

    Returns TOPICS (not URLs) derived from query-level GSC data.
    This sheet answers: "What should we write that doesn't exist on the site yet?"

    Includes cannibalization detection to exclude topics where a strong page already exists.
    """
    # Define output columns (includes Business Relevance columns)
    empty_columns = [
        'Suggested Topic', 'Primary Keyword', 'Secondary Keywords', 'Total Impressions',
        'Avg Position', 'Recommended Action', 'Primary URL', 'Secondary URLs', 'Reasoning',
        'Suggested Page Type', 'Priority Score', 'Business Relevance Score', 'Feasibility Notes'
    ]

    if gsc_df is None or len(gsc_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Build set of thin content URLs and url_word_counts dict for Content Action logic
    thin_urls = set()
    url_word_counts = {}
    if df is not None and len(df) > 0:
        word_count_col = None
        for col in ['word_count', 'Word Count', 'wordcount']:
            if col in df.columns:
                word_count_col = col
                break

        url_col = None
        for col in ['url', 'URL', 'Address']:
            if col in df.columns:
                url_col = col
                break

        if word_count_col and url_col:
            df_copy = df.copy()
            df_copy[word_count_col] = pd.to_numeric(df_copy[word_count_col], errors='coerce').fillna(0)
            thin_urls = set(df_copy[df_copy[word_count_col] < thin_threshold][url_col].tolist())
            url_word_counts = dict(zip(df_copy[url_col], df_copy[word_count_col]))
            logger.debug(f"Identified {len(thin_urls)} thin content URLs and {len(url_word_counts)} URL word counts")

    gsc_copy = gsc_df.copy()

    # Map column names if needed
    query_col = None
    for col in ['query', 'primary_keyword', 'Query', 'Keyword', 'Top queries']:
        if col in gsc_copy.columns:
            query_col = col
            break

    if query_col is None:
        if 'primary_keyword' in gsc_copy.columns:
            gsc_copy['query'] = gsc_copy['primary_keyword']
        else:
            logger.warning("GSC data missing query column - cannot build topic-based opportunities")
            return pd.DataFrame(columns=empty_columns)
    elif query_col != 'query':
        gsc_copy['query'] = gsc_copy[query_col]

    # Ensure URL column exists for cannibalization detection
    url_col = None
    for col in ['url', 'URL', 'page', 'Page', 'landing_page']:
        if col in gsc_copy.columns:
            url_col = col
            break

    if url_col and url_col != 'url':
        gsc_copy['url'] = gsc_copy[url_col]
    elif url_col is None:
        gsc_copy['url'] = ''  # No URL data available

    for col in ['impressions', 'clicks', 'avg_position']:
        if col not in gsc_copy.columns:
            gsc_copy[col] = 0
        gsc_copy[col] = pd.to_numeric(gsc_copy[col], errors='coerce').fillna(0)

    qualifying_queries = gsc_copy[gsc_copy['impressions'] >= 10].copy()

    if len(qualifying_queries) == 0:
        return pd.DataFrame(columns=empty_columns)

    result = cluster_queries_into_topics_api(qualifying_queries, max_topics=20, thin_urls=thin_urls, url_word_counts=url_word_counts)

    if len(result) == 0:
        return pd.DataFrame(columns=empty_columns)

    # =========================================================================
    # BUSINESS RELEVANCE FILTERING (Gate 3)
    # =========================================================================
    # Build site context from crawl and GSC data
    site_context = build_site_context_from_df_api(df, gsc_df)

    # Apply business relevance filtering
    filtered_result, excluded_topics = filter_topics_by_business_relevance_api(
        topics_df=result,
        site_context=site_context,
        min_score=50,  # Minimum score threshold
        max_topics=15  # Cap to 15 topics max
    )

    # Log excluded topics for debugging
    if excluded_topics:
        logger.debug(f"[NEW CONTENT] Excluded {len(excluded_topics)} topics by business relevance filter")
        for ex in excluded_topics[:5]:  # Log first 5
            logger.debug(f"  - Excluded: {ex.get('Primary Keyword', 'N/A')} | Reason: {ex.get('Reason', 'N/A')}")

    # Ensure all expected columns exist
    for col in empty_columns:
        if col not in filtered_result.columns:
            filtered_result[col] = ''

    # Reorder columns to match expected output
    return filtered_result[empty_columns]


def build_redirect_merge_plan_api(new_content_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the 'Redirect & Merge Plan' sheet from consolidation recommendations - API version.

    Generates a 301 redirect mapping for every topic where the Recommended Action
    is "Consolidate pages". Each Secondary URL gets its own row mapping to the
    Primary URL.

    Args:
        new_content_df: DataFrame from build_new_content_opportunities_api
            containing Recommended Action, Primary URL, Secondary URLs columns

    Returns:
        DataFrame with columns: Topic, Recommended Action, Primary URL,
        Secondary URL, Redirect Type, Reason
    """
    empty_columns = [
        'Topic', 'Recommended Action', 'Primary URL', 'Secondary URL',
        'Redirect Type', 'Reason'
    ]

    if new_content_df is None or len(new_content_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Filter for consolidation actions only
    consolidations = new_content_df[
        new_content_df['Recommended Action'] == 'Consolidate pages'
    ].copy()

    if len(consolidations) == 0:
        return pd.DataFrame(columns=empty_columns)

    redirect_rows = []

    for _, row in consolidations.iterrows():
        topic = row.get('Suggested Topic', '')
        primary_url = row.get('Primary URL', '')
        secondary_urls_str = row.get('Secondary URLs', '')

        # Skip if no primary URL or secondary URLs
        if not primary_url or not secondary_urls_str:
            continue

        # Parse secondary URLs (comma-separated)
        secondary_urls = [
            url.strip() for url in secondary_urls_str.split(',')
            if url.strip() and url.strip() != primary_url
        ]

        # Skip if no valid secondary URLs after filtering
        if not secondary_urls:
            continue

        # Generate redirect entries for each secondary URL
        for secondary_url in secondary_urls:
            # Generate redirect-specific reasoning
            reason = (
                f"This page competes for the same keywords as the primary page "
                f"({primary_url[:50]}...) but has weaker rankings and engagement. "
                f"Redirecting consolidates authority and prevents keyword cannibalization."
            )

            redirect_rows.append({
                'Topic': topic,
                'Recommended Action': 'Consolidate pages',
                'Primary URL': primary_url,
                'Secondary URL': secondary_url,
                'Redirect Type': '301',
                'Reason': reason
            })

    if not redirect_rows:
        return pd.DataFrame(columns=empty_columns)

    result_df = pd.DataFrame(redirect_rows)

    # Remove any duplicate redirects (same Secondary URL)
    result_df = result_df.drop_duplicates(subset=['Secondary URL'], keep='first')

    # Ensure no self-redirects (Primary URL == Secondary URL)
    result_df = result_df[result_df['Primary URL'] != result_df['Secondary URL']]

    logger.info(f"Generated {len(result_df)} redirect entries from {len(consolidations)} consolidation topics")

    return result_df


def build_merge_playbooks_api(new_content_df: pd.DataFrame, gsc_df: Optional[pd.DataFrame] = None, url_word_counts: Optional[dict] = None) -> pd.DataFrame:
    """
    Build the 'Merge Playbooks' sheet with content-level merge instructions - API version.

    For each "Consolidate pages" action, generates detailed content recommendations
    on what to keep from the Primary URL and what to move from Secondary URLs.

    Args:
        new_content_df: DataFrame from build_new_content_opportunities_api
            containing Recommended Action, Primary URL, Secondary URLs, etc.
        gsc_df: Raw GSC query-level data for analyzing per-URL query coverage
        url_word_counts: Dict mapping URL -> word count (optional)

    Returns:
        DataFrame with columns: Topic, Primary URL, Secondary URL, Keep This Content,
        Move These Sections, Retire This Page, Reasoning
    """
    empty_columns = [
        'Topic', 'Primary URL', 'Secondary URL', 'Keep This Content',
        'Move These Sections', 'Retire This Page', 'Reasoning'
    ]

    if new_content_df is None or len(new_content_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Filter for consolidation actions only
    consolidations = new_content_df[
        new_content_df['Recommended Action'] == 'Consolidate pages'
    ].copy()

    if len(consolidations) == 0:
        return pd.DataFrame(columns=empty_columns)

    if url_word_counts is None:
        url_word_counts = {}

    # Build URL -> queries mapping from GSC data if available
    url_queries = {}  # url -> list of {query, impressions, clicks, position}
    if gsc_df is not None and len(gsc_df) > 0:
        gsc_copy = gsc_df.copy()

        # Find query column
        query_col = None
        for col in ['query', 'primary_keyword', 'Query', 'Keyword', 'Top queries']:
            if col in gsc_copy.columns:
                query_col = col
                break

        # Find URL column
        url_col = None
        for col in ['url', 'URL', 'page', 'Page', 'landing_page']:
            if col in gsc_copy.columns:
                url_col = col
                break

        if query_col and url_col:
            for _, row in gsc_copy.iterrows():
                url = str(row.get(url_col, '')).strip()
                query = str(row.get(query_col, '')).strip()
                if url and query and query != 'nan':
                    if url not in url_queries:
                        url_queries[url] = []
                    url_queries[url].append({
                        'query': query,
                        'impressions': int(row.get('impressions', 0)),
                        'clicks': int(row.get('clicks', 0)),
                        'position': float(row.get('avg_position', row.get('position', 0)))
                    })

    playbook_rows = []

    for _, row in consolidations.iterrows():
        topic = row.get('Suggested Topic', '')
        primary_url = row.get('Primary URL', '')
        secondary_urls_str = row.get('Secondary URLs', '')

        # Skip if no primary URL or secondary URLs
        if not primary_url or not secondary_urls_str:
            continue

        # Parse secondary URLs (comma-separated)
        secondary_urls = [
            url.strip() for url in secondary_urls_str.split(',')
            if url.strip() and url.strip() != primary_url
        ]

        if not secondary_urls:
            continue

        # Get Primary URL's queries
        primary_queries = url_queries.get(primary_url, [])
        primary_query_set = {q['query'].lower() for q in primary_queries}
        primary_word_count = url_word_counts.get(primary_url, 0)

        # Calculate total impressions for Primary URL
        primary_total_impressions = sum(q['impressions'] for q in primary_queries)

        # Build "Keep This Content" for Primary URL
        keep_content_bullets = []

        if primary_queries:
            # Group by position ranges
            top_position_queries = [q for q in primary_queries if q['position'] > 0 and q['position'] <= 5]
            strong_position_queries = [q for q in primary_queries if q['position'] > 5 and q['position'] <= 10]

            # Check for queries ranking 1-5
            if top_position_queries:
                top_query_examples = sorted(top_position_queries, key=lambda x: x['impressions'], reverse=True)[:3]
                query_names = ', '.join([f"'{q['query']}'" for q in top_query_examples])
                keep_content_bullets.append(f"Top-ranking content (positions 1-5): {query_names}")

            # Check for high-impression query clusters
            high_impression_queries = [q for q in primary_queries if q['impressions'] >= 50]
            if high_impression_queries:
                total_high_imp = sum(q['impressions'] for q in high_impression_queries)
                if primary_total_impressions > 0:
                    share = total_high_imp / primary_total_impressions
                    if share >= 0.20:
                        keep_content_bullets.append(f"High-traffic sections ({share:.0%} of impressions)")

            # Check for strong position queries (6-10)
            if strong_position_queries:
                keep_content_bullets.append(f"Content ranking positions 6-10 ({len(strong_position_queries)} queries)")

        # Add word count context
        if primary_word_count >= 1000:
            keep_content_bullets.append(f"Comprehensive content ({primary_word_count:,} words)")
        elif primary_word_count > 0:
            keep_content_bullets.append(f"Existing content ({primary_word_count:,} words)")

        if not keep_content_bullets:
            keep_content_bullets.append("Primary page structure and core content")

        keep_content = '\n'.join([f"• {b}" for b in keep_content_bullets])

        # Generate one row per secondary URL
        for secondary_url in secondary_urls:
            # Get Secondary URL's queries
            secondary_queries = url_queries.get(secondary_url, [])
            secondary_word_count = url_word_counts.get(secondary_url, 0)

            # Build "Move These Sections" for Secondary URL
            move_sections_bullets = []

            if secondary_queries:
                # Find unique queries not covered by Primary
                unique_queries = [
                    q for q in secondary_queries
                    if q['query'].lower() not in primary_query_set
                ]

                # Find queries where Secondary ranks better (>3 positions better)
                better_ranking_queries = []
                for sq in secondary_queries:
                    sq_query_lower = sq['query'].lower()
                    for pq in primary_queries:
                        if pq['query'].lower() == sq_query_lower:
                            if pq['position'] > 0 and sq['position'] > 0:
                                if pq['position'] - sq['position'] > 3:
                                    better_ranking_queries.append(sq)
                            break

                # Add unique query coverage
                if unique_queries:
                    unique_examples = sorted(unique_queries, key=lambda x: x['impressions'], reverse=True)[:3]
                    unique_names = ', '.join([f"'{q['query']}'" for q in unique_examples])
                    move_sections_bullets.append(f"Unique content not on Primary: {unique_names}")

                # Add better-ranking content
                if better_ranking_queries:
                    better_examples = sorted(better_ranking_queries, key=lambda x: x['impressions'], reverse=True)[:2]
                    better_names = ', '.join([f"'{q['query']}'" for q in better_examples])
                    move_sections_bullets.append(f"Content ranking better than Primary: {better_names}")

                # Check for high-value queries on Secondary
                high_value_secondary = [q for q in secondary_queries if q['impressions'] >= 30]
                if high_value_secondary and not move_sections_bullets:
                    hv_examples = sorted(high_value_secondary, key=lambda x: x['impressions'], reverse=True)[:2]
                    hv_names = ', '.join([f"'{q['query']}'" for q in hv_examples])
                    move_sections_bullets.append(f"High-impression queries: {hv_names}")

            # Check if Secondary has more content depth
            if secondary_word_count > primary_word_count and secondary_word_count > 500:
                diff = secondary_word_count - primary_word_count
                move_sections_bullets.append(f"Additional content depth ({diff:,} more words) - review for valuable sections")

            if not move_sections_bullets:
                move_sections_bullets.append("Review page for any unique angles or examples to preserve")

            move_sections = '\n'.join([f"• {b}" for b in move_sections_bullets])

            # Generate reasoning for this specific merge
            reasoning = (
                f"Merging '{secondary_url}' into '{primary_url}' eliminates keyword cannibalization "
                f"for the '{topic}' topic. Primary URL has stronger overall authority; "
                f"Secondary URL's unique content should be integrated before redirect."
            )

            playbook_rows.append({
                'Topic': topic,
                'Primary URL': primary_url,
                'Secondary URL': secondary_url,
                'Keep This Content': keep_content,
                'Move These Sections': move_sections,
                'Retire This Page': 'Yes',
                'Reasoning': reasoning
            })

    if not playbook_rows:
        return pd.DataFrame(columns=empty_columns)

    result_df = pd.DataFrame(playbook_rows)

    logger.info(f"Generated {len(result_df)} merge playbook entries from {len(consolidations)} consolidation topics")

    return result_df


# =============================================================================
# ANALYTICAL SHEETS WRITER
# =============================================================================

def write_analytical_sheets_api(workbook, df: pd.DataFrame, thin_content_threshold: int = 1000, gsc_df: Optional[pd.DataFrame] = None) -> None:
    """Write analytical insight sheets to workbook - API version."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Debug instrumentation: Log GSC query data details before building analytical sheets
    if gsc_df is not None and len(gsc_df) > 0:
        logger.info(f"[ANALYTICAL DEBUG] GSC data received: {len(gsc_df)} rows")
        logger.info(f"[ANALYTICAL DEBUG] Columns: {list(gsc_df.columns)}")
        query_col = 'query' if 'query' in gsc_df.columns else ('primary_keyword' if 'primary_keyword' in gsc_df.columns else None)
        url_col = 'url' if 'url' in gsc_df.columns else ('page' if 'page' in gsc_df.columns else None)
        if query_col:
            unique_queries = gsc_df[query_col].nunique() if query_col in gsc_df.columns else 0
            logger.info(f"[ANALYTICAL DEBUG] Unique queries ({query_col}): {unique_queries}")
        if url_col:
            unique_pages = gsc_df[url_col].nunique() if url_col in gsc_df.columns else 0
            logger.info(f"[ANALYTICAL DEBUG] Unique pages ({url_col}): {unique_pages}")
        if 'impressions' in gsc_df.columns:
            total_impressions = gsc_df['impressions'].sum()
            qualifying = len(gsc_df[gsc_df['impressions'] >= 10])
            logger.info(f"[ANALYTICAL DEBUG] Total impressions: {total_impressions:,.0f}, Qualifying rows (>=10 impr): {qualifying}")
            if qualifying > 0:
                top_queries = gsc_df.nlargest(5, 'impressions')
                for _, row in top_queries.iterrows():
                    q = row.get(query_col, 'N/A') if query_col else 'N/A'
                    u = row.get(url_col, 'N/A') if url_col else 'N/A'
                    logger.info(f"[ANALYTICAL DEBUG]   - '{q}' ({row.get('impressions', 0):,.0f} impr) -> {u[:60]}...")
    else:
        logger.warning(f"[ANALYTICAL DEBUG] GSC data is None or empty - topic clustering will not run")

    # Sheet 1: Content to Optimize
    content_df = build_content_to_optimize_api(df)
    ws1 = workbook.create_sheet(title='Content to Optimize')

    if len(content_df) > 0:
        for col_idx, col_name in enumerate(content_df.columns, start=1):
            cell = ws1.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(content_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        ws1.column_dimensions['A'].width = 60
        ws1.column_dimensions['J'].width = 50
    else:
        ws1.cell(row=1, column=1, value='No content optimization opportunities found')
    ws1.freeze_panes = 'A2'
    logger.info(f"Wrote Content to Optimize sheet: {len(content_df)} rows")

    # Sheet 2: Thin Content Opportunities
    thin_df = build_thin_content_api(df, thin_content_threshold)
    ws2 = workbook.create_sheet(title='Thin Content Opportunities')

    if len(thin_df) > 0:
        for col_idx, col_name in enumerate(thin_df.columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(thin_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        ws2.column_dimensions['A'].width = 60
        ws2.column_dimensions['I'].width = 50
    else:
        ws2.cell(row=1, column=1, value='No thin content opportunities found')
    ws2.freeze_panes = 'A2'
    logger.info(f"Wrote Thin Content Opportunities sheet: {len(thin_df)} rows")

    # Sheet 3: New Content Opportunities (Topic-based, not URL-based)
    # Pass thin_content_threshold for cannibalization detection
    new_content_df = build_new_content_opportunities_api(df, gsc_df, thin_content_threshold)
    ws3 = workbook.create_sheet(title='New Content Opportunities')

    if len(new_content_df) > 0:
        for col_idx, col_name in enumerate(new_content_df.columns, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(new_content_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
        # Column widths for topic-based structure with cannibalization detection
        ws3.column_dimensions['A'].width = 35  # Suggested Topic
        ws3.column_dimensions['B'].width = 35  # Primary Keyword
        ws3.column_dimensions['C'].width = 60  # Secondary Keywords
        ws3.column_dimensions['D'].width = 16  # Total Impressions
        ws3.column_dimensions['E'].width = 12  # Avg Position
        ws3.column_dimensions['F'].width = 22  # Recommended Action
        ws3.column_dimensions['G'].width = 60  # Primary URL
        ws3.column_dimensions['H'].width = 80  # Secondary URLs
        ws3.column_dimensions['I'].width = 100  # Reasoning
        ws3.column_dimensions['J'].width = 22  # Suggested Page Type
        ws3.column_dimensions['K'].width = 14  # Priority Score
    else:
        # Provide specific reason for empty New Content Opportunities
        if gsc_df is None:
            empty_reason = "GSC data not connected. Connect Google Search Console or upload a GSC CSV file with query-level data."
        elif len(gsc_df) == 0:
            empty_reason = "GSC returned 0 rows. The property may have no data yet, or the date range returned empty results."
        elif 'query' not in gsc_df.columns and 'primary_keyword' not in gsc_df.columns:
            empty_reason = "GSC data missing 'query' column. Ensure your GSC export includes query-level data, not just page-level aggregates."
        elif 'impressions' in gsc_df.columns and len(gsc_df[gsc_df['impressions'] >= 10]) == 0:
            empty_reason = "No queries with >=10 impressions found. Try a longer date range or check if the property has search traffic."
        else:
            empty_reason = "Topic clustering found no qualifying clusters. Check logs for details."
        ws3.cell(row=1, column=1, value=empty_reason)
        logger.warning(f"[NEW CONTENT] Empty result: {empty_reason}")
    ws3.freeze_panes = 'A2'
    logger.info(f"Wrote New Content Opportunities sheet: {len(new_content_df)} topics")

    # Sheet 4: Redirect & Merge Plan (for consolidation actions)
    redirect_df = build_redirect_merge_plan_api(new_content_df)
    ws4 = workbook.create_sheet(title='Redirect & Merge Plan')

    if len(redirect_df) > 0:
        for col_idx, col_name in enumerate(redirect_df.columns, start=1):
            cell = ws4.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(redirect_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws4.cell(row=row_idx, column=col_idx, value=value)

        # Column widths for redirect plan
        ws4.column_dimensions['A'].width = 35  # Topic
        ws4.column_dimensions['B'].width = 20  # Recommended Action
        ws4.column_dimensions['C'].width = 70  # Primary URL (destination)
        ws4.column_dimensions['D'].width = 70  # Secondary URL (source)
        ws4.column_dimensions['E'].width = 14  # Redirect Type
        ws4.column_dimensions['F'].width = 100  # Reason
    else:
        # Provide specific reason for empty Redirect & Merge Plan
        if len(new_content_df) == 0:
            redirect_empty_reason = "No topic clusters generated - see 'New Content Opportunities' sheet for details."
        elif 'Recommended Action' not in new_content_df.columns:
            redirect_empty_reason = "Topic data missing 'Recommended Action' column - clustering may have failed."
        elif len(new_content_df[new_content_df['Recommended Action'] == 'Consolidate pages']) == 0:
            redirect_empty_reason = "No 'Consolidate pages' actions found. All topics have single-page or expansion recommendations."
        else:
            redirect_empty_reason = "No consolidation redirects generated - check logs for details."
        ws4.cell(row=1, column=1, value=redirect_empty_reason)

    ws4.freeze_panes = 'A2'
    logger.info(f"Wrote Redirect & Merge Plan sheet: {len(redirect_df)} redirects")

    # Sheet 5: Merge Playbooks (content-level merge instructions)
    # Build url_word_counts for merge playbooks
    url_word_counts = {}
    if df is not None and len(df) > 0:
        word_count_col = None
        for col in ['word_count', 'Word Count', 'wordcount']:
            if col in df.columns:
                word_count_col = col
                break

        url_col = None
        for col in ['url', 'URL', 'Address']:
            if col in df.columns:
                url_col = col
                break

        if word_count_col and url_col:
            df_copy = df.copy()
            df_copy[word_count_col] = pd.to_numeric(df_copy[word_count_col], errors='coerce').fillna(0)
            url_word_counts = dict(zip(df_copy[url_col], df_copy[word_count_col]))

    playbook_df = build_merge_playbooks_api(new_content_df, gsc_df, url_word_counts)
    ws5 = workbook.create_sheet(title='Merge Playbooks')

    if len(playbook_df) > 0:
        for col_idx, col_name in enumerate(playbook_df.columns, start=1):
            cell = ws5.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(playbook_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws5.cell(row=row_idx, column=col_idx, value=value)

        # Column widths for merge playbooks
        ws5.column_dimensions['A'].width = 35  # Topic
        ws5.column_dimensions['B'].width = 70  # Primary URL
        ws5.column_dimensions['C'].width = 70  # Secondary URL
        ws5.column_dimensions['D'].width = 60  # Keep This Content
        ws5.column_dimensions['E'].width = 60  # Move These Sections
        ws5.column_dimensions['F'].width = 18  # Retire This Page
        ws5.column_dimensions['G'].width = 100  # Reasoning
    else:
        # Provide specific reason for empty Merge Playbooks
        if len(new_content_df) == 0:
            playbook_empty_reason = "No topic clusters generated - see 'New Content Opportunities' sheet for details."
        elif len(redirect_df) == 0:
            playbook_empty_reason = "No consolidation actions found - see 'Redirect & Merge Plan' sheet for details."
        else:
            playbook_empty_reason = "Merge playbooks could not be generated from the consolidation data - check logs for details."
        ws5.cell(row=1, column=1, value=playbook_empty_reason)

    ws5.freeze_panes = 'A2'
    logger.info(f"Wrote Merge Playbooks sheet: {len(playbook_df)} entries")


# =============================================================================
# MAIN EXCEL REPORT FUNCTION
# =============================================================================

def create_excel_report(
    df: pd.DataFrame,
    summaries: Dict[str, pd.DataFrame],
    quality_report: Optional[DataQualityReport] = None,
    gsc_df: Optional[pd.DataFrame] = None
) -> bytes:
    """Create WQA Excel report with proper 3-header-row structure and data quality sheet"""
    output = io.BytesIO()

    # Define the 38-column structure
    # Row 1: Section headings (sparse)
    section_headers = {
        0: '', 1: 'ANALYSIS', 2: '', 3: '', 4: '', 5: '',
        6: 'KEYWORD PERFORMANCE', 7: '', 8: '', 9: '', 10: '', 11: '',
        12: 'TRAFFIC PERFORMANCE', 13: '', 14: '', 15: '',
        16: 'ENGAGEMENT', 17: '',
        18: 'CONVERSIONS', 19: '', 20: '', 21: '',
        22: 'ON PAGE', 23: '', 24: '', 25: '', 26: '', 27: '', 28: '', 29: '', 30: '',
        31: 'TECHNICAL', 32: '', 33: '', 34: '', 35: '', 36: '', 37: '',
        38: 'AI'
    }

    # Row 2: Source labels
    source_labels = {
        0: 'Formula', 1: 'SF', 2: 'Manual', 3: 'Manual', 4: 'Manual', 5: 'Manual',
        6: 'SEMrush', 7: 'SEMrush', 8: 'SEMrush', 9: 'SEMrush', 10: 'SEMrush', 11: 'SEMrush',
        12: 'GSC', 13: 'GA', 14: 'GA', 15: 'GA',
        16: 'GA', 17: 'GA',
        18: 'GA', 19: 'GA', 20: 'GA', 21: 'GA',
        22: 'SF', 23: 'SF', 24: 'GSC', 25: 'SF', 26: 'SF', 27: 'SF', 28: 'SF', 29: 'SF', 30: 'Ahrefs',
        31: 'SF', 32: 'SF', 33: 'SF', 34: 'SF', 35: 'SF', 36: 'Formula', 37: 'SF',
        38: 'Claude'
    }

    # Row 3: Column names
    column_names = [
        'page-path', 'URL', 'Category', 'Technical Action', 'Content Action', 'Final URL',
        'Main KW', 'Volume', 'Ranking', '"Best" KW', 'Volume', 'Ranking',
        'Impressions', 'Sessions', '% Change Sessions', 'Losing Traffic?',
        'Bounce rate (%)', 'Average session duration',
        'Conversions (All Goals)', 'Conversion Rate (%)', 'Ecom Revenue Generated', 'Ecom Conversion Rate',
        'Type', 'Current Title', 'SERP CTR', 'Meta', 'H1', 'Word Count', 'Inlinks', 'Outlinks', 'DOFOLLOW Links',
        'Canonical Link Element', 'Status Code', 'Index / Noindex', 'Indexation Status', 'Page Depth', 'In Sitemap?', 'Last Modified',
        'AI Recommendation'
    ]

    # Build data rows
    data_rows = []
    for _, row in df.iterrows():
        # Get canonical URL or regular URL
        canonical = str(row.get('canonical_url', '')).strip()
        url = str(row.get('url', '')).strip()
        display_url = canonical if canonical and canonical != '' else url

        # Extract page path
        page_path = extract_page_path(display_url)

        # Get indexability info
        index_label, indexation_status = format_indexability(row)

        # Calculate conversion rate
        sessions = float(row.get('sessions', 0))
        conversions = float(row.get('conversions', 0))
        conv_rate = (conversions / sessions) if sessions > 0 else ''

        # Calculate % change sessions and losing traffic
        sessions_prev = float(row.get('sessions_prev', 0))
        if sessions_prev > 0:
            pct_change = (sessions - sessions_prev) / sessions_prev
            losing_traffic = 'Yes' if pct_change <= -0.1 else 'No'
        else:
            pct_change = ''
            losing_traffic = 'No YoY Data'

        # Format in_sitemap
        in_sitemap = 'Yes' if row.get('in_sitemap', False) else 'No'

        # Get best keyword data - prefer keyword file data, fall back to GSC
        best_kw = row.get('best_kw', '')
        best_kw_volume = row.get('best_kw_volume', '')
        best_kw_ranking = row.get('best_kw_ranking', '')

        # Fall back to GSC data if no keyword file data
        if not best_kw:
            best_kw = row.get('primary_keyword', '')
            # Only show GSC ranking if we have a keyword from GSC
            best_kw_ranking = row.get('avg_position', '') if best_kw else ''
            best_kw_volume = ''  # GSC doesn't provide volume

        # Build row data matching column order
        data_row = [
            page_path,                                          # 0: page-path
            display_url,                                        # 1: URL
            row.get('page_type', 'Other'),                      # 2: Category
            row.get('technical_actions', ''),                   # 3: Technical Action
            row.get('content_actions', ''),                     # 4: Content Action
            row.get('final_url', ''),                           # 5: Final URL
            row.get('main_kw', ''),                             # 6: Main KW
            row.get('main_kw_volume', ''),                      # 7: Volume (Main KW)
            row.get('main_kw_ranking', ''),                     # 8: Ranking (Main KW)
            best_kw,                                            # 9: "Best" KW
            best_kw_volume,                                     # 10: Volume (Best KW)
            best_kw_ranking,                                    # 11: Ranking (Best KW)
            row.get('impressions', 0),                          # 12: Impressions
            sessions if sessions > 0 else '',                   # 13: Sessions
            pct_change,                                         # 14: % Change Sessions
            losing_traffic,                                     # 15: Losing Traffic?
            row.get('bounce_rate', ''),                         # 16: Bounce rate (%)
            row.get('avg_session_duration', ''),                # 17: Average session duration
            conversions if conversions > 0 else '',             # 18: Conversions (All Goals)
            conv_rate,                                          # 19: Conversion Rate (%)
            row.get('ecom_revenue', ''),                        # 20: Ecom Revenue Generated
            row.get('ecom_conv_rate', ''),                      # 21: Ecom Conversion Rate
            row.get('content_type', ''),                        # 22: Type
            row.get('page_title', ''),                          # 23: Current Title
            row.get('ctr', ''),                                 # 24: SERP CTR
            row.get('meta_description', ''),                    # 25: Meta
            row.get('h1', ''),                                  # 26: H1
            row.get('word_count', 0),                           # 27: Word Count
            row.get('inlinks', 0),                              # 28: Inlinks
            row.get('outlinks', 0),                             # 29: Outlinks
            row.get('dofollow_links', 0),                        # 30: DOFOLLOW Links
            row.get('canonical_url', ''),                       # 31: Canonical Link Element
            row.get('status_code', ''),                         # 32: Status Code
            index_label,                                        # 33: Index / Noindex
            indexation_status,                                  # 34: Indexation Status
            row.get('crawl_depth', ''),                         # 35: Page Depth
            in_sitemap,                                         # 36: In Sitemap?
            row.get('last_modified', ''),                       # 37: Last Modified
            row.get('ai_recommendation', ''),                   # 38: AI Recommendation
        ]
        data_rows.append(data_row)

    # Create DataFrame with proper structure
    wqa_df = pd.DataFrame(data_rows, columns=column_names)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write WQA sheet with 3 header rows
        workbook = writer.book
        worksheet = workbook.create_sheet('WQA', 0)

        # Row 1: Section headers with merging and distinct colors
        # Define merge ranges and their labels/colors
        # A1:F1 (cols 1-6), G1:L1 (cols 7-12), M1:P1 (cols 13-16), Q1:R1 (cols 17-18),
        # S1:V1 (cols 19-22), W1:AE1 (cols 23-31), AF1:AL1 (cols 32-38)
        section_merges = [
            ('A1:F1', 'ANALYSIS', '2F5496'),           # Dark Blue
            ('G1:L1', 'KEYWORD PERFORMANCE', '548235'), # Dark Green
            ('M1:P1', 'TRAFFIC PERFORMANCE', 'C65911'), # Dark Orange
            ('Q1:R1', 'ENGAGEMENT', '7030A0'),          # Purple
            ('S1:V1', 'CONVERSIONS', 'BF8F00'),         # Dark Gold
            ('W1:AE1', 'ON PAGE', '385723'),            # Forest Green
            ('AF1:AL1', 'TECHNICAL', '833C0C'),         # Dark Brown
            ('AM1:AM1', 'AI', '1A73E8'),               # Google Blue
        ]

        for merge_range, label, color in section_merges:
            # Merge the cells
            worksheet.merge_cells(merge_range)
            # Get the first cell of the merged range to set value and style
            first_cell = merge_range.split(':')[0]
            cell = worksheet[first_cell]
            cell.value = label
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: Source labels
        for col_idx, source in source_labels.items():
            worksheet.cell(row=2, column=col_idx + 1, value=source)

        # Row 3: Column names
        for col_idx, col_name in enumerate(column_names):
            worksheet.cell(row=3, column=col_idx + 1, value=col_name)

        # Row 4+: Data rows
        for row_idx, data_row in enumerate(data_rows):
            for col_idx, value in enumerate(data_row):
                worksheet.cell(row=row_idx + 4, column=col_idx + 1, value=value)

        # Apply formatting to Row 2 and Row 3 (Row 1 already styled with merges above)

        # Header row styling for Row 3
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx in range(len(column_names)):
            # Source labels (Row 2)
            cell2 = worksheet.cell(row=2, column=col_idx + 1)
            cell2.fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
            cell2.font = Font(italic=True, size=9)

            # Column names (Row 3)
            cell3 = worksheet.cell(row=3, column=col_idx + 1)
            cell3.fill = header_fill
            cell3.font = header_font

        # Freeze panes at row 4 (after headers)
        worksheet.freeze_panes = 'A4'

        # Define column indices for formatting (0-based)
        # Percentage columns: % Change Sessions (14), Bounce rate (16), Conversion Rate (19), Ecom Conv Rate (21), SERP CTR (24)
        percentage_cols = [14, 16, 19, 21, 24]
        # Time duration column: Average session duration (17)
        duration_cols = [17]
        # Date column: Last Modified (37)
        date_cols = [37]
        # Currency column: Ecom Revenue (20)
        currency_cols = [20]
        # Number columns: Impressions (12), Sessions (13), Conversions (18), Word Count (27), Inlinks (28), Outlinks (29), DOFOLLOW (30)
        number_cols = [12, 13, 18, 27, 28, 29, 30]

        # Apply number formatting to data rows
        for row_idx in range(len(data_rows)):
            excel_row = row_idx + 4  # Data starts at row 4

            # Format percentage columns
            for col_idx in percentage_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        val = float(cell.value)
                        # If value is already decimal (0-1 range), use as-is
                        # If value is whole number (like 50 for 50%), convert
                        if val > 1:
                            cell.value = val / 100
                        cell.number_format = FORMAT_PERCENTAGE_00
                    except (ValueError, TypeError):
                        pass

            # Format duration columns (seconds to mm:ss)
            for col_idx in duration_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        seconds = float(cell.value)
                        # Convert seconds to time format (fraction of day)
                        cell.value = seconds / 86400  # Convert to Excel time
                        cell.number_format = '[mm]:ss'
                    except (ValueError, TypeError):
                        pass

            # Format currency columns
            for col_idx in currency_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '$#,##0.00'
                    except (ValueError, TypeError):
                        pass

            # Format number columns with comma separators
            for col_idx in number_cols:
                cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                if cell.value is not None and cell.value != '':
                    try:
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass

        # Auto-adjust column widths (basic)
        for col_idx, col_name in enumerate(column_names):
            col_letter = get_column_letter(col_idx + 1)
            # Set reasonable default widths based on column type
            if col_name in ['URL', 'Canonical Link Element', 'Meta', 'Current Title']:
                worksheet.column_dimensions[col_letter].width = 50
            elif col_name in ['page-path', 'Final URL', 'H1', 'AI Recommendation']:
                worksheet.column_dimensions[col_letter].width = 45
            elif col_name in ['Technical Action', 'Content Action', 'Category', 'Indexation Status']:
                worksheet.column_dimensions[col_letter].width = 20
            else:
                worksheet.column_dimensions[col_letter].width = 15

        # Also write the raw aggregation data to a separate sheet for reference
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Write summary sheet
        current_row = 0
        summary_order = [
            ('Priority Distribution', 'priority'),
            ('Technical Actions', 'technical_actions'),
            ('Content Actions', 'content_actions'),
            ('Page Types', 'page_type'),
            ('Status Codes', 'status_code'),
            ('AI Recommendations', 'ai_recommendations'),
        ]
        for title, key in summary_order:
            if key in summaries:
                title_df = pd.DataFrame([[title]], columns=[''])
                title_df.to_excel(writer, sheet_name='Summary', index=False,
                                  header=False, startrow=current_row)
                current_row += 1
                summaries[key].to_excel(writer, sheet_name='Summary', index=False,
                                        startrow=current_row)
                current_row += len(summaries[key]) + 3

        # Sheet 4: Data Quality (if provided)
        if quality_report is not None:
            quality_sheets = generate_data_quality_sheets(quality_report)
            current_row = 0

            # Write header
            header_df = pd.DataFrame([['DATA QUALITY REPORT']], columns=[''])
            header_df.to_excel(writer, sheet_name='Data Quality', index=False,
                              header=False, startrow=current_row)
            current_row += 2

            # Write Data Source Coverage
            title_df = pd.DataFrame([['Data Source Coverage']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['coverage'].empty:
                quality_sheets['coverage'].to_excel(writer, sheet_name='Data Quality',
                                                    index=False, startrow=current_row)
                current_row += len(quality_sheets['coverage']) + 3

            # Write Column Quality
            title_df = pd.DataFrame([['Column Data Quality']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['column_quality'].empty:
                quality_sheets['column_quality'].to_excel(writer, sheet_name='Data Quality',
                                                          index=False, startrow=current_row)
                current_row += len(quality_sheets['column_quality']) + 3

            # Write Issues
            title_df = pd.DataFrame([['Data Quality Issues']], columns=[''])
            title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                             header=False, startrow=current_row)
            current_row += 1

            if not quality_sheets['issues'].empty:
                quality_sheets['issues'].to_excel(writer, sheet_name='Data Quality',
                                                  index=False, startrow=current_row)
                current_row += len(quality_sheets['issues']) + 3
            else:
                no_issues_df = pd.DataFrame([['No data quality issues found!']], columns=[''])
                no_issues_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                     header=False, startrow=current_row)

            # Write warnings if any
            if quality_report.warnings:
                current_row += 2
                title_df = pd.DataFrame([['Warnings']], columns=[''])
                title_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                 header=False, startrow=current_row)
                current_row += 1
                warnings_df = pd.DataFrame([[w] for w in quality_report.warnings], columns=['Warning'])
                warnings_df.to_excel(writer, sheet_name='Data Quality', index=False,
                                    startrow=current_row)

            logger.info(f"Wrote Data Quality sheet: {len(quality_report.issues)} issues found")

        # Sheets 5-7: Analytical Insight Sheets
        write_analytical_sheets_api(workbook, df, thin_content_threshold=1000, gsc_df=gsc_df)

    output.seek(0)
    return output.getvalue()
