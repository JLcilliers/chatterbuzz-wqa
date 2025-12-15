#!/usr/bin/env python3
"""
Website Quality Audit (WQA) Generator

A Python tool that automatically generates Website Quality Audit reports for SEO agencies.
Ingests multiple CSV exports (crawl, GA4, GSC, backlinks), joins them into a single
URL-level dataset, applies rule-based SEO logic, and exports comprehensive Excel reports.

Usage:
    python wqa_generator.py --crawl crawl.csv --output wqa_report.xlsx
    python wqa_generator.py --crawl crawl.csv --ga ga.csv --gsc gsc.csv --backlinks backlinks.csv --output wqa_report.xlsx

Author: Generated for SEO Agency WQA automation
Python Version: 3.11+
"""

import argparse
import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# COLUMN MAPPING CONFIGURATION
# =============================================================================
# Adapt these mappings if your CSV exports use different column names

CRAWL_COLUMN_MAP = {
    'url': ['url', 'address', 'page_url', 'page'],
    'status_code': ['status_code', 'status', 'http_status', 'response_code'],
    'indexable': ['indexable', 'indexability', 'is_indexable', 'index_status'],
    'canonical_url': ['canonical_url', 'canonical', 'canonical_link', 'canonical_tag'],
    'inlinks': ['inlinks', 'internal_inlinks', 'inlinks_count', 'unique_inlinks'],
    'outlinks': ['outlinks', 'internal_outlinks', 'outlinks_count', 'unique_outlinks'],
    'crawl_depth': ['crawl_depth', 'depth', 'click_depth', 'level'],
    'in_sitemap': ['in_sitemap', 'sitemap', 'in_xml_sitemap', 'sitemap_status'],
    'page_title': ['page_title', 'title', 'title_1', 'meta_title'],
    'meta_description': ['meta_description', 'description', 'meta_description_1'],
    'word_count': ['word_count', 'words', 'content_word_count', 'text_word_count'],
}

GA_COLUMN_MAP = {
    'url': ['url', 'page_path', 'page', 'landing_page', 'page_location'],
    'sessions': ['sessions', 'session_count', 'visits'],
    'conversions': ['conversions', 'goal_completions', 'conversion_count', 'key_events'],
}

GSC_COLUMN_MAP = {
    'url': ['url', 'page', 'top_pages', 'landing_page'],
    'avg_position': ['avg_position', 'position', 'average_position', 'avg_pos'],
    'ctr': ['ctr', 'click_through_rate', 'clickthrough_rate'],
    'clicks': ['clicks', 'click_count'],
    'impressions': ['impressions', 'impression_count'],
    'primary_keyword': ['primary_keyword', 'query', 'top_query', 'keyword'],
}

BACKLINK_COLUMN_MAP = {
    'url': ['url', 'target_url', 'page', 'target'],
    'referring_domains': ['referring_domains', 'ref_domains', 'domains', 'rd'],
    'backlinks': ['backlinks', 'backlink_count', 'external_links', 'links'],
    'anchor_texts': ['anchor_texts', 'anchors', 'anchor_text'],
}


# =============================================================================
# CLI ARGUMENT PARSING
# =============================================================================

def parse_args() -> argparse.Namespace:
    """
    Parse command-line arguments for the WQA generator.

    Returns:
        argparse.Namespace: Parsed arguments with all configuration options.
    """
    parser = argparse.ArgumentParser(
        description='Website Quality Audit (WQA) Generator - Generate comprehensive SEO audit reports',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python wqa_generator.py --crawl crawl.csv --output report.xlsx
  python wqa_generator.py --crawl crawl.csv --ga ga.csv --gsc gsc.csv --backlinks backlinks.csv --output report.xlsx
  python wqa_generator.py --crawl crawl.csv --output report.xlsx --low-traffic-threshold 10 --thin-content-threshold 500
        '''
    )

    # Required arguments
    parser.add_argument(
        '--crawl',
        type=str,
        required=True,
        help='Path to crawl CSV file (required)'
    )
    parser.add_argument(
        '--output',
        type=str,
        required=True,
        help='Path to output Excel file (required)'
    )

    # Optional data source arguments
    parser.add_argument(
        '--ga',
        type=str,
        default=None,
        help='Path to GA4 CSV file (optional)'
    )
    parser.add_argument(
        '--gsc',
        type=str,
        default=None,
        help='Path to GSC CSV file (optional)'
    )
    parser.add_argument(
        '--backlinks',
        type=str,
        default=None,
        help='Path to backlinks CSV file (optional)'
    )

    # Threshold arguments
    parser.add_argument(
        '--low-traffic-threshold',
        type=int,
        default=5,
        help='Sessions threshold for low traffic classification (default: 5)'
    )
    parser.add_argument(
        '--thin-content-threshold',
        type=int,
        default=1000,
        help='Word count threshold for thin content (default: 1000)'
    )
    parser.add_argument(
        '--high-rank-max-position',
        type=float,
        default=20.0,
        help='Maximum position to consider for ranking optimization (default: 20)'
    )
    parser.add_argument(
        '--low-ctr-threshold',
        type=float,
        default=0.05,
        help='CTR threshold for metadata optimization (default: 0.05 = 5%%)'
    )

    return parser.parse_args()


# =============================================================================
# DATA LOADING FUNCTIONS
# =============================================================================

def find_column(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """
    Find the first matching column name from a list of possibilities.

    Args:
        df: DataFrame to search in
        possible_names: List of possible column names to look for

    Returns:
        The first matching column name found, or None if no match
    """
    df_columns_lower = {col.lower().strip(): col for col in df.columns}
    for name in possible_names:
        if name.lower() in df_columns_lower:
            return df_columns_lower[name.lower()]
    return None


def map_columns(df: pd.DataFrame, column_map: Dict[str, List[str]], source_name: str) -> pd.DataFrame:
    """
    Map DataFrame columns to standardized names using a mapping dictionary.

    Args:
        df: DataFrame with original column names
        column_map: Dictionary mapping standard names to possible source names
        source_name: Name of the data source for logging

    Returns:
        DataFrame with standardized column names
    """
    result = pd.DataFrame()

    for standard_name, possible_names in column_map.items():
        found_col = find_column(df, possible_names)
        if found_col:
            result[standard_name] = df[found_col]
        else:
            logger.warning(f"[{source_name}] Column '{standard_name}' not found. Tried: {possible_names}")
            result[standard_name] = None

    return result


def load_crawl_data(path: str) -> pd.DataFrame:
    """
    Load and normalize crawl data from CSV.

    Args:
        path: Path to the crawl CSV file

    Returns:
        DataFrame with standardized crawl data columns

    Raises:
        FileNotFoundError: If the crawl file doesn't exist
    """
    logger.info(f"Loading crawl data from: {path}")

    if not Path(path).exists():
        raise FileNotFoundError(f"Crawl file not found: {path}")

    df = pd.read_csv(path, low_memory=False)
    logger.info(f"Loaded {len(df)} rows from crawl data")

    # Map columns to standard names
    df_mapped = map_columns(df, CRAWL_COLUMN_MAP, 'Crawl')

    # Normalize indexable column to boolean
    if 'indexable' in df_mapped.columns and df_mapped['indexable'] is not None:
        df_mapped['indexable'] = df_mapped['indexable'].apply(normalize_indexable)

    # Normalize in_sitemap column to boolean
    if 'in_sitemap' in df_mapped.columns and df_mapped['in_sitemap'] is not None:
        df_mapped['in_sitemap'] = df_mapped['in_sitemap'].apply(normalize_boolean)

    # Ensure numeric columns
    numeric_cols = ['status_code', 'inlinks', 'outlinks', 'crawl_depth', 'word_count']
    for col in numeric_cols:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0).astype(int)

    return df_mapped


def load_ga_data(path: Optional[str]) -> Optional[pd.DataFrame]:
    """
    Load and normalize GA4 data from CSV.

    Args:
        path: Path to the GA4 CSV file, or None if not provided

    Returns:
        DataFrame with standardized GA data columns, or None if path is None
    """
    if path is None:
        logger.info("GA4 data not provided - skipping")
        return None

    if not Path(path).exists():
        logger.warning(f"GA4 file not found: {path} - skipping")
        return None

    logger.info(f"Loading GA4 data from: {path}")
    df = pd.read_csv(path, low_memory=False)
    logger.info(f"Loaded {len(df)} rows from GA4 data")

    df_mapped = map_columns(df, GA_COLUMN_MAP, 'GA4')

    # Ensure numeric columns
    for col in ['sessions', 'conversions']:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)

    return df_mapped


def load_gsc_data(path: Optional[str]) -> Optional[pd.DataFrame]:
    """
    Load and normalize GSC data from CSV.

    Args:
        path: Path to the GSC CSV file, or None if not provided

    Returns:
        DataFrame with standardized GSC data columns, or None if path is None
    """
    if path is None:
        logger.info("GSC data not provided - skipping")
        return None

    if not Path(path).exists():
        logger.warning(f"GSC file not found: {path} - skipping")
        return None

    logger.info(f"Loading GSC data from: {path}")
    df = pd.read_csv(path, low_memory=False)
    logger.info(f"Loaded {len(df)} rows from GSC data")

    df_mapped = map_columns(df, GSC_COLUMN_MAP, 'GSC')

    # Ensure numeric columns
    for col in ['avg_position', 'ctr', 'clicks', 'impressions']:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)

    return df_mapped


def load_backlink_data(path: Optional[str]) -> Optional[pd.DataFrame]:
    """
    Load and normalize backlink data from CSV.

    Args:
        path: Path to the backlinks CSV file, or None if not provided

    Returns:
        DataFrame with standardized backlink data columns, or None if path is None
    """
    if path is None:
        logger.info("Backlink data not provided - skipping")
        return None

    if not Path(path).exists():
        logger.warning(f"Backlink file not found: {path} - skipping")
        return None

    logger.info(f"Loading backlink data from: {path}")
    df = pd.read_csv(path, low_memory=False)
    logger.info(f"Loaded {len(df)} rows from backlink data")

    df_mapped = map_columns(df, BACKLINK_COLUMN_MAP, 'Backlinks')

    # Ensure numeric columns
    for col in ['referring_domains', 'backlinks']:
        if col in df_mapped.columns:
            df_mapped[col] = pd.to_numeric(df_mapped[col], errors='coerce').fillna(0)

    return df_mapped


# =============================================================================
# DATA NORMALIZATION HELPERS
# =============================================================================

def normalize_indexable(value) -> bool:
    """
    Normalize indexable field to boolean.

    Args:
        value: Raw value from CSV (could be bool, string, or other)

    Returns:
        Boolean indicating if the page is indexable
    """
    if pd.isna(value):
        return True  # Assume indexable if not specified

    if isinstance(value, bool):
        return value

    str_val = str(value).lower().strip()

    # Check for indexable indicators
    if str_val in ['true', '1', 'yes', 'index', 'indexable']:
        return True

    # Check for non-indexable indicators
    if str_val in ['false', '0', 'no', 'noindex', 'non-indexable', 'not indexable']:
        return False

    return True  # Default to indexable


def normalize_boolean(value) -> bool:
    """
    Normalize boolean-like field to boolean.

    Args:
        value: Raw value from CSV

    Returns:
        Boolean value
    """
    if pd.isna(value):
        return False

    if isinstance(value, bool):
        return value

    str_val = str(value).lower().strip()
    return str_val in ['true', '1', 'yes', 'y']


def normalize_url(url: str) -> str:
    """
    Normalize URL for consistent matching.

    Args:
        url: Raw URL string

    Returns:
        Normalized URL string
    """
    if pd.isna(url) or not url:
        return ''

    url = str(url).strip().lower()

    # Remove trailing slash for consistency (except for root)
    if url.endswith('/') and len(url) > 1:
        parsed = urlparse(url)
        if parsed.path != '/':
            url = url.rstrip('/')

    return url


# =============================================================================
# DATASET MERGING
# =============================================================================

def merge_datasets(
    crawl_df: pd.DataFrame,
    ga_df: Optional[pd.DataFrame],
    gsc_df: Optional[pd.DataFrame],
    backlink_df: Optional[pd.DataFrame]
) -> pd.DataFrame:
    """
    Merge all datasets into a single URL-level DataFrame.

    Uses the crawl data as the base universe and left joins other data sources.

    Args:
        crawl_df: Crawl data (required, forms the base)
        ga_df: GA4 data (optional)
        gsc_df: GSC data (optional)
        backlink_df: Backlink data (optional)

    Returns:
        Merged DataFrame with all columns
    """
    logger.info("Merging datasets...")

    # Start with crawl data as base
    df = crawl_df.copy()

    # Normalize URLs in base dataset
    df['url'] = df['url'].apply(normalize_url)
    df = df[df['url'] != '']  # Remove empty URLs

    # Merge GA4 data
    if ga_df is not None and 'url' in ga_df.columns:
        ga_df = ga_df.copy()
        ga_df['url'] = ga_df['url'].apply(normalize_url)
        ga_df = ga_df[ga_df['url'] != '']

        # Aggregate GA data by URL (in case of duplicates)
        ga_agg = ga_df.groupby('url').agg({
            'sessions': 'sum',
            'conversions': 'sum'
        }).reset_index()

        df = df.merge(ga_agg, on='url', how='left', suffixes=('', '_ga'))
        logger.info(f"Merged GA4 data - {len(ga_agg)} unique URLs")

    # Merge GSC data
    if gsc_df is not None and 'url' in gsc_df.columns:
        gsc_df = gsc_df.copy()
        gsc_df['url'] = gsc_df['url'].apply(normalize_url)
        gsc_df = gsc_df[gsc_df['url'] != '']

        # Aggregate GSC data by URL
        gsc_agg_cols = {}
        if 'avg_position' in gsc_df.columns:
            gsc_agg_cols['avg_position'] = 'mean'
        if 'ctr' in gsc_df.columns:
            gsc_agg_cols['ctr'] = 'mean'
        if 'clicks' in gsc_df.columns:
            gsc_agg_cols['clicks'] = 'sum'
        if 'impressions' in gsc_df.columns:
            gsc_agg_cols['impressions'] = 'sum'
        if 'primary_keyword' in gsc_df.columns:
            gsc_agg_cols['primary_keyword'] = 'first'

        if gsc_agg_cols:
            gsc_agg = gsc_df.groupby('url').agg(gsc_agg_cols).reset_index()
            df = df.merge(gsc_agg, on='url', how='left', suffixes=('', '_gsc'))
            logger.info(f"Merged GSC data - {len(gsc_agg)} unique URLs")

    # Merge backlink data
    if backlink_df is not None and 'url' in backlink_df.columns:
        backlink_df = backlink_df.copy()
        backlink_df['url'] = backlink_df['url'].apply(normalize_url)
        backlink_df = backlink_df[backlink_df['url'] != '']

        # Aggregate backlink data by URL
        bl_agg_cols = {}
        if 'referring_domains' in backlink_df.columns:
            bl_agg_cols['referring_domains'] = 'sum'
        if 'backlinks' in backlink_df.columns:
            bl_agg_cols['backlinks'] = 'sum'

        if bl_agg_cols:
            bl_agg = backlink_df.groupby('url').agg(bl_agg_cols).reset_index()
            df = df.merge(bl_agg, on='url', how='left', suffixes=('', '_bl'))
            logger.info(f"Merged backlink data - {len(bl_agg)} unique URLs")

    # Ensure all expected columns exist with defaults
    expected_columns = {
        'url': '',
        'status_code': 0,
        'indexable': True,
        'canonical_url': '',
        'inlinks': 0,
        'outlinks': 0,
        'crawl_depth': 0,
        'in_sitemap': False,
        'page_title': '',
        'meta_description': '',
        'word_count': 0,
        'sessions': 0,
        'conversions': 0,
        'avg_position': 0.0,
        'ctr': 0.0,
        'clicks': 0,
        'impressions': 0,
        'referring_domains': 0,
        'backlinks': 0,
        'primary_keyword': '',
    }

    for col, default in expected_columns.items():
        if col not in df.columns:
            df[col] = default
        else:
            # Fill NaN values with defaults
            if isinstance(default, str):
                df[col] = df[col].fillna(default).astype(str)
            elif isinstance(default, bool):
                df[col] = df[col].fillna(default).astype(bool)
            elif isinstance(default, int):
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default).astype(int)
            else:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default)

    logger.info(f"Final merged dataset: {len(df)} URLs with {len(df.columns)} columns")
    return df


# =============================================================================
# PAGE TYPE CLASSIFICATION
# =============================================================================

def classify_page_type(url: str) -> str:
    """
    Classify the page type based on URL patterns.

    Args:
        url: The URL to classify

    Returns:
        Page type string: 'Home', 'Blog', 'Local Lander', 'Service', or 'Other'
    """
    if not url or pd.isna(url):
        return 'Other'

    url_lower = str(url).lower()

    # Parse URL to get path
    try:
        parsed = urlparse(url_lower)
        path = parsed.path
    except Exception:
        path = url_lower

    # Home page detection
    if path in ['/', ''] or path.rstrip('/') == '':
        return 'Home'

    # Blog/News detection
    if any(pattern in path for pattern in ['/blog/', '/blog', '/news/', '/news', '/articles/', '/posts/']):
        return 'Blog'

    # Local lander detection
    if any(pattern in path for pattern in ['/location/', '/locations/', '/city/', '/cities/',
                                            '/service-area/', '/service-areas/', '/area/', '/areas/',
                                            '/near-me/', '/local/']):
        return 'Local Lander'

    # Service page detection
    if any(pattern in path for pattern in ['/service/', '/services/', '/solutions/', '/products/',
                                            '/what-we-do/', '/offerings/']):
        return 'Service'

    return 'Other'


# =============================================================================
# RULES ENGINE - ACTION ASSIGNMENT
# =============================================================================

def assign_actions(
    row: pd.Series,
    low_traffic_threshold: int = 5,
    thin_content_threshold: int = 1000,
    high_rank_max_position: float = 20.0,
    low_ctr_threshold: float = 0.05
) -> Tuple[List[str], List[str]]:
    """
    Apply rule-based SEO logic to assign technical and content actions.

    Args:
        row: A single row from the URL-level DataFrame
        low_traffic_threshold: Sessions threshold for low traffic
        thin_content_threshold: Word count threshold for thin content
        high_rank_max_position: Maximum position for ranking optimization
        low_ctr_threshold: CTR threshold for metadata optimization

    Returns:
        Tuple of (technical_actions, content_actions) lists
    """
    technical_actions: List[str] = []
    content_actions: List[str] = []

    # Extract row values with safe defaults
    status_code = int(row.get('status_code', 0))
    indexable = bool(row.get('indexable', True))
    canonical_url = str(row.get('canonical_url', '')).strip()
    url = str(row.get('url', '')).strip().lower()
    inlinks = int(row.get('inlinks', 0))
    outlinks = int(row.get('outlinks', 0))
    crawl_depth = int(row.get('crawl_depth', 0))
    in_sitemap = bool(row.get('in_sitemap', False))
    page_title = str(row.get('page_title', '')).strip()
    meta_description = str(row.get('meta_description', '')).strip()
    word_count = int(row.get('word_count', 0))
    sessions = float(row.get('sessions', 0))
    conversions = float(row.get('conversions', 0))
    avg_position = float(row.get('avg_position', 0))
    ctr = float(row.get('ctr', 0))
    impressions = float(row.get('impressions', 0))
    referring_domains = int(row.get('referring_domains', 0))
    backlinks = int(row.get('backlinks', 0))
    page_type = str(row.get('page_type', 'Other'))

    # =========================================================================
    # TECHNICAL ACTIONS
    # =========================================================================

    # 6.1 Status code & redirect logic
    if status_code == 302:
        technical_actions.append('301 Redirect')

    if status_code == 404:
        if backlinks > 0 or referring_domains > 0:
            technical_actions.append('301 Redirect')
        elif backlinks == 0 and inlinks > 0:
            technical_actions.append('Remove Internal Links')

    if status_code in [301, 308]:
        technical_actions.append('Review Redirect')

    # 6.2 Indexability, sitemap, and robots
    if not indexable and in_sitemap:
        technical_actions.append('Remove from Sitemap')

    if indexable and status_code == 200 and sessions > 0 and not in_sitemap:
        technical_actions.append('Add to Sitemap')

    # 6.3 Canonical tags
    if canonical_url and canonical_url.lower() != url:
        technical_actions.append('Canonicalize')

    # 6.4 Internal links & crawl depth
    if sessions > 0 and inlinks == 0:
        technical_actions.append('Add Internal Links')

    if crawl_depth >= 4 and page_type in ['Home', 'Service', 'Local Lander']:
        if 'Add Internal Links' not in technical_actions:
            technical_actions.append('Add Internal Links')

    # 6.5 Schema hints
    if page_type == 'Blog':
        technical_actions.append('Add Schema: Article')
    elif page_type == 'Local Lander':
        technical_actions.append('Add Schema: LocalBusiness')
    elif page_type == 'Home':
        technical_actions.append('Add Schema: Organization')

    # =========================================================================
    # CONTENT ACTIONS
    # =========================================================================

    # 7.1 Deletion / redirect candidates
    # Hard delete (404) candidate
    if (sessions <= low_traffic_threshold and
        conversions == 0 and
        backlinks == 0 and
        referring_domains == 0 and
        status_code == 200):  # Only for live pages
        content_actions.append('Delete (404)')

    # Redirect candidate (content-level consolidation)
    elif (sessions <= low_traffic_threshold and
          conversions == 0 and
          (backlinks > 0 or referring_domains > 0) and
          status_code == 200):
        content_actions.append('301 Redirect')

    # 7.2 Thin vs substantial content
    # Rewrite
    if (word_count < thin_content_threshold and
        (sessions > 0 or avg_position > 0) and
        status_code == 200 and
        'Delete (404)' not in content_actions and
        '301 Redirect' not in content_actions):
        content_actions.append('Rewrite')

    # Refresh
    if (word_count >= thin_content_threshold and
        sessions > 0 and
        2 < avg_position <= high_rank_max_position and
        status_code == 200 and
        'Delete (404)' not in content_actions and
        '301 Redirect' not in content_actions):
        content_actions.append('Refresh')

    # Target w/ Links
    if (word_count >= thin_content_threshold and
        sessions > 0 and
        3 <= avg_position <= 20 and
        referring_domains == 0 and
        status_code == 200 and
        'Delete (404)' not in content_actions and
        '301 Redirect' not in content_actions):
        content_actions.append('Target w/ Links')

    # 7.3 Metadata optimizations
    # Update Meta Description
    if status_code == 200:
        if (not meta_description or
            (avg_position <= high_rank_max_position and
             ctr < low_ctr_threshold and
             impressions > 100)):
            content_actions.append('Update Meta Description')

    # Update Page Title
    if status_code == 200:
        if (not page_title or
            (avg_position <= high_rank_max_position and
             ctr < low_ctr_threshold and
             impressions > 100)):
            content_actions.append('Update Page Title')

    # 7.4 Default / No-action
    if not content_actions:
        content_actions.append('Leave As Is')

    return technical_actions, content_actions


# =============================================================================
# PRIORITY ASSIGNMENT
# =============================================================================

def assign_priority(row: pd.Series) -> str:
    """
    Assign priority level to a URL based on status, page type, and actions.

    Args:
        row: A single row from the URL-level DataFrame

    Returns:
        Priority string: 'High', 'Medium', or 'Low'
    """
    status_code = int(row.get('status_code', 0))
    page_type = str(row.get('page_type', 'Other'))
    content_actions = str(row.get('content_actions', ''))

    # High priority conditions
    if status_code in [404, 302]:
        return 'High'

    if page_type in ['Home', 'Service', 'Local Lander']:
        if 'Rewrite' in content_actions or 'Refresh' in content_actions:
            return 'High'

    # Medium priority conditions
    if 'Delete (404)' in content_actions or '301 Redirect' in content_actions:
        if page_type not in ['Home', 'Service', 'Local Lander']:
            return 'Medium'

    return 'Low'


# =============================================================================
# DATA QUALITY VALIDATION
# =============================================================================

class DataQualityReport:
    """Container for data quality validation results."""

    def __init__(self):
        self.missing_columns: Dict[str, List[str]] = {}  # source -> missing columns
        self.empty_columns: List[str] = []  # columns with all empty/default values
        self.sparse_columns: Dict[str, float] = {}  # column -> % missing
        self.url_match_stats: Dict[str, Dict] = {}  # source -> match stats
        self.data_source_coverage: Dict[str, int] = {}  # source -> rows matched
        self.issues: List[Dict] = []  # list of specific issues found
        self.warnings: List[str] = []  # general warnings

    def add_issue(self, category: str, severity: str, description: str,
                  affected_count: int = 0, examples: List[str] = None):
        """Add a specific data quality issue."""
        self.issues.append({
            'category': category,
            'severity': severity,
            'description': description,
            'affected_count': affected_count,
            'examples': examples or []
        })

    def to_dataframe(self) -> pd.DataFrame:
        """Convert issues to a DataFrame for Excel export."""
        if not self.issues:
            return pd.DataFrame(columns=['Category', 'Severity', 'Description', 'Affected Count', 'Examples'])

        rows = []
        for issue in self.issues:
            rows.append({
                'Category': issue['category'],
                'Severity': issue['severity'],
                'Description': issue['description'],
                'Affected Count': issue['affected_count'],
                'Examples': '; '.join(issue['examples'][:5]) if issue['examples'] else ''
            })

        return pd.DataFrame(rows)


def validate_data_quality(
    merged_df: pd.DataFrame,
    crawl_df: pd.DataFrame,
    ga_df: Optional[pd.DataFrame],
    gsc_df: Optional[pd.DataFrame],
    backlink_df: Optional[pd.DataFrame]
) -> DataQualityReport:
    """
    Validate merged data quality and identify missing/problematic data.

    Args:
        merged_df: The merged URL-level DataFrame
        crawl_df: Original crawl data
        ga_df: Original GA4 data (or None)
        gsc_df: Original GSC data (or None)
        backlink_df: Original backlink data (or None)

    Returns:
        DataQualityReport with all identified issues
    """
    report = DataQualityReport()
    total_urls = len(merged_df)

    logger.info("=" * 60)
    logger.info("DATA QUALITY VALIDATION")
    logger.info("=" * 60)

    # =========================================================================
    # 1. Check URL matching coverage from each source
    # =========================================================================

    crawl_urls = set(merged_df['url'].dropna().unique())

    # GA4 URL matching
    if ga_df is not None and 'url' in ga_df.columns:
        ga_urls = set(ga_df['url'].apply(normalize_url).dropna().unique())
        ga_urls = ga_urls - {''}
        matched_ga = crawl_urls & ga_urls
        unmatched_ga = ga_urls - crawl_urls

        report.data_source_coverage['GA4'] = len(matched_ga)
        report.url_match_stats['GA4'] = {
            'total_source_urls': len(ga_urls),
            'matched': len(matched_ga),
            'unmatched': len(unmatched_ga),
            'match_rate': len(matched_ga) / len(ga_urls) * 100 if ga_urls else 0
        }

        # Check how many crawl URLs have GA data
        urls_with_ga = merged_df[merged_df['sessions'] > 0]['url'].nunique()
        ga_coverage = urls_with_ga / total_urls * 100 if total_urls > 0 else 0

        logger.info(f"GA4: {len(matched_ga)}/{len(ga_urls)} URLs matched ({report.url_match_stats['GA4']['match_rate']:.1f}%)")

        if report.url_match_stats['GA4']['match_rate'] < 50:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'GA4 URL match rate is low ({report.url_match_stats["GA4"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_ga),
                examples=list(unmatched_ga)[:10]
            )
        elif report.url_match_stats['GA4']['match_rate'] < 80:
            report.add_issue(
                category='URL Matching',
                severity='Medium',
                description=f'GA4 URL match rate is moderate ({report.url_match_stats["GA4"]["match_rate"]:.1f}%).',
                affected_count=len(unmatched_ga),
                examples=list(unmatched_ga)[:10]
            )
    else:
        report.data_source_coverage['GA4'] = 0
        report.warnings.append('GA4 data not provided - sessions/conversions will be 0')

    # GSC URL matching
    if gsc_df is not None and 'url' in gsc_df.columns:
        gsc_urls = set(gsc_df['url'].apply(normalize_url).dropna().unique())
        gsc_urls = gsc_urls - {''}
        matched_gsc = crawl_urls & gsc_urls
        unmatched_gsc = gsc_urls - crawl_urls

        report.data_source_coverage['GSC'] = len(matched_gsc)
        report.url_match_stats['GSC'] = {
            'total_source_urls': len(gsc_urls),
            'matched': len(matched_gsc),
            'unmatched': len(unmatched_gsc),
            'match_rate': len(matched_gsc) / len(gsc_urls) * 100 if gsc_urls else 0
        }

        logger.info(f"GSC: {len(matched_gsc)}/{len(gsc_urls)} URLs matched ({report.url_match_stats['GSC']['match_rate']:.1f}%)")

        if report.url_match_stats['GSC']['match_rate'] < 50:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'GSC URL match rate is low ({report.url_match_stats["GSC"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_gsc),
                examples=list(unmatched_gsc)[:10]
            )
        elif report.url_match_stats['GSC']['match_rate'] < 80:
            report.add_issue(
                category='URL Matching',
                severity='Medium',
                description=f'GSC URL match rate is moderate ({report.url_match_stats["GSC"]["match_rate"]:.1f}%).',
                affected_count=len(unmatched_gsc),
                examples=list(unmatched_gsc)[:10]
            )
    else:
        report.data_source_coverage['GSC'] = 0
        report.warnings.append('GSC data not provided - rankings/CTR will be 0')

    # Backlink URL matching
    if backlink_df is not None and 'url' in backlink_df.columns:
        bl_urls = set(backlink_df['url'].apply(normalize_url).dropna().unique())
        bl_urls = bl_urls - {''}
        matched_bl = crawl_urls & bl_urls
        unmatched_bl = bl_urls - crawl_urls

        report.data_source_coverage['Backlinks'] = len(matched_bl)
        report.url_match_stats['Backlinks'] = {
            'total_source_urls': len(bl_urls),
            'matched': len(matched_bl),
            'unmatched': len(unmatched_bl),
            'match_rate': len(matched_bl) / len(bl_urls) * 100 if bl_urls else 0
        }

        logger.info(f"Backlinks: {len(matched_bl)}/{len(bl_urls)} URLs matched ({report.url_match_stats['Backlinks']['match_rate']:.1f}%)")

        if report.url_match_stats['Backlinks']['match_rate'] < 50:
            report.add_issue(
                category='URL Matching',
                severity='High',
                description=f'Backlink URL match rate is low ({report.url_match_stats["Backlinks"]["match_rate"]:.1f}%). Check URL format differences.',
                affected_count=len(unmatched_bl),
                examples=list(unmatched_bl)[:10]
            )
    else:
        report.data_source_coverage['Backlinks'] = 0
        report.warnings.append('Backlink data not provided - referring_domains/backlinks will be 0')

    # =========================================================================
    # 2. Check for empty/sparse columns in merged data
    # =========================================================================

    critical_columns = {
        'sessions': ('GA4', 0),
        'conversions': ('GA4', 0),
        'avg_position': ('GSC', 0.0),
        'ctr': ('GSC', 0.0),
        'clicks': ('GSC', 0),
        'impressions': ('GSC', 0),
        'primary_keyword': ('GSC', ''),
        'referring_domains': ('Backlinks', 0),
        'backlinks': ('Backlinks', 0),
        'word_count': ('Crawl', 0),
        'page_title': ('Crawl', ''),
        'meta_description': ('Crawl', ''),
    }

    for col, (source, default) in critical_columns.items():
        if col not in merged_df.columns:
            report.empty_columns.append(col)
            report.add_issue(
                category='Missing Column',
                severity='High',
                description=f'Column "{col}" is completely missing from {source} data.',
                affected_count=total_urls
            )
            continue

        # Check what percentage of values are default/empty
        if isinstance(default, str):
            empty_mask = (merged_df[col].isna()) | (merged_df[col] == '') | (merged_df[col] == default)
        else:
            empty_mask = (merged_df[col].isna()) | (merged_df[col] == default)

        empty_count = empty_mask.sum()
        empty_pct = empty_count / total_urls * 100 if total_urls > 0 else 0

        report.sparse_columns[col] = empty_pct

        if empty_pct == 100:
            report.empty_columns.append(col)
            report.add_issue(
                category='Empty Column',
                severity='High',
                description=f'Column "{col}" has no data (100% empty/default). Source: {source}',
                affected_count=empty_count
            )
            logger.warning(f"Column '{col}' is 100% empty/default")
        elif empty_pct >= 90:
            report.add_issue(
                category='Sparse Data',
                severity='High',
                description=f'Column "{col}" is {empty_pct:.1f}% empty/default. Source: {source}',
                affected_count=empty_count
            )
            logger.warning(f"Column '{col}' is {empty_pct:.1f}% empty/default")
        elif empty_pct >= 70:
            report.add_issue(
                category='Sparse Data',
                severity='Medium',
                description=f'Column "{col}" is {empty_pct:.1f}% empty/default. Source: {source}',
                affected_count=empty_count
            )

    # =========================================================================
    # 3. Check for specific data quality issues
    # =========================================================================

    # Check for URLs with traffic but no ranking data
    if 'sessions' in merged_df.columns and 'avg_position' in merged_df.columns:
        traffic_no_rank = merged_df[(merged_df['sessions'] > 0) & (merged_df['avg_position'] == 0)]
        if len(traffic_no_rank) > 0:
            report.add_issue(
                category='Data Inconsistency',
                severity='Medium',
                description=f'URLs with traffic but no GSC ranking data (avg_position=0).',
                affected_count=len(traffic_no_rank),
                examples=list(traffic_no_rank['url'].head(10))
            )

    # Check for URLs with rankings but no traffic
    if 'sessions' in merged_df.columns and 'avg_position' in merged_df.columns:
        rank_no_traffic = merged_df[(merged_df['avg_position'] > 0) & (merged_df['avg_position'] <= 10) & (merged_df['sessions'] == 0)]
        if len(rank_no_traffic) > 0:
            report.add_issue(
                category='Data Inconsistency',
                severity='Low',
                description=f'URLs ranking in top 10 but with 0 sessions (possible GA tracking issue).',
                affected_count=len(rank_no_traffic),
                examples=list(rank_no_traffic['url'].head(10))
            )

    # Check for pages with 0 word count
    if 'word_count' in merged_df.columns and 'status_code' in merged_df.columns:
        zero_word_count = merged_df[(merged_df['word_count'] == 0) & (merged_df['status_code'] == 200)]
        if len(zero_word_count) > 0:
            report.add_issue(
                category='Content Quality',
                severity='Medium',
                description=f'Live pages (200 status) with 0 word count.',
                affected_count=len(zero_word_count),
                examples=list(zero_word_count['url'].head(10))
            )

    # Check for missing page titles
    if 'page_title' in merged_df.columns and 'status_code' in merged_df.columns:
        no_title = merged_df[(merged_df['page_title'] == '') & (merged_df['status_code'] == 200)]
        if len(no_title) > 0:
            report.add_issue(
                category='SEO Quality',
                severity='Medium',
                description=f'Live pages missing page titles.',
                affected_count=len(no_title),
                examples=list(no_title['url'].head(10))
            )

    # Check for missing meta descriptions
    if 'meta_description' in merged_df.columns and 'status_code' in merged_df.columns:
        no_meta = merged_df[(merged_df['meta_description'] == '') & (merged_df['status_code'] == 200)]
        if len(no_meta) > 0:
            report.add_issue(
                category='SEO Quality',
                severity='Low',
                description=f'Live pages missing meta descriptions.',
                affected_count=len(no_meta),
                examples=list(no_meta['url'].head(10))
            )

    # =========================================================================
    # 4. Generate summary
    # =========================================================================

    high_issues = sum(1 for i in report.issues if i['severity'] == 'High')
    medium_issues = sum(1 for i in report.issues if i['severity'] == 'Medium')
    low_issues = sum(1 for i in report.issues if i['severity'] == 'Low')

    logger.info("-" * 60)
    logger.info(f"Data Quality Summary:")
    logger.info(f"  Total URLs: {total_urls}")
    logger.info(f"  Issues Found: {len(report.issues)} ({high_issues} High, {medium_issues} Medium, {low_issues} Low)")
    logger.info(f"  Empty Columns: {len(report.empty_columns)}")
    if report.warnings:
        for warning in report.warnings:
            logger.info(f"  Warning: {warning}")
    logger.info("=" * 60)

    return report


def generate_data_quality_sheets(report: DataQualityReport) -> Dict[str, pd.DataFrame]:
    """
    Generate DataFrames for data quality reporting in Excel.

    Args:
        report: DataQualityReport with validation results

    Returns:
        Dictionary of DataFrames for Excel sheets
    """
    sheets = {}

    # Issues sheet
    sheets['issues'] = report.to_dataframe()

    # Coverage summary
    coverage_data = []
    for source, count in report.data_source_coverage.items():
        stats = report.url_match_stats.get(source, {})
        coverage_data.append({
            'Data Source': source,
            'URLs Matched': count,
            'Total Source URLs': stats.get('total_source_urls', 'N/A'),
            'Match Rate (%)': f"{stats.get('match_rate', 0):.1f}" if stats else 'N/A',
            'Unmatched URLs': stats.get('unmatched', 0)
        })
    sheets['coverage'] = pd.DataFrame(coverage_data)

    # Column sparsity
    sparsity_data = []
    for col, pct in sorted(report.sparse_columns.items(), key=lambda x: -x[1]):
        sparsity_data.append({
            'Column': col,
            'Empty/Default (%)': f"{pct:.1f}",
            'Status': 'Critical' if pct >= 90 else ('Warning' if pct >= 70 else 'OK')
        })
    sheets['column_quality'] = pd.DataFrame(sparsity_data)

    return sheets


# =============================================================================
# SUMMARY GENERATION
# =============================================================================

def generate_summary(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """
    Generate summary statistics from the processed DataFrame.

    Args:
        df: Processed URL-level DataFrame with actions and priorities

    Returns:
        Dictionary containing summary DataFrames
    """
    logger.info("Generating summary statistics...")

    summaries = {}

    # Count technical actions
    tech_actions_flat = []
    for actions in df['technical_actions']:
        if actions:
            for action in str(actions).split(', '):
                action = action.strip()
                if action:
                    tech_actions_flat.append(action)

    tech_counts = pd.Series(tech_actions_flat).value_counts().reset_index()
    tech_counts.columns = ['Technical Action', 'Count']
    summaries['technical_actions'] = tech_counts

    # Count content actions
    content_actions_flat = []
    for actions in df['content_actions']:
        if actions:
            for action in str(actions).split(', '):
                action = action.strip()
                if action:
                    content_actions_flat.append(action)

    content_counts = pd.Series(content_actions_flat).value_counts().reset_index()
    content_counts.columns = ['Content Action', 'Count']
    summaries['content_actions'] = content_counts

    # Count by priority
    priority_counts = df['priority'].value_counts().reset_index()
    priority_counts.columns = ['Priority', 'Count']
    # Ensure proper order
    priority_order = ['High', 'Medium', 'Low']
    priority_counts['sort_order'] = priority_counts['Priority'].apply(
        lambda x: priority_order.index(x) if x in priority_order else 99
    )
    priority_counts = priority_counts.sort_values('sort_order').drop('sort_order', axis=1)
    summaries['priority'] = priority_counts

    # Count by page type
    page_type_counts = df['page_type'].value_counts().reset_index()
    page_type_counts.columns = ['Page Type', 'Count']
    summaries['page_type'] = page_type_counts

    # Count by status code
    status_counts = df['status_code'].value_counts().reset_index()
    status_counts.columns = ['Status Code', 'Count']
    status_counts = status_counts.sort_values('Status Code')
    summaries['status_code'] = status_counts

    logger.info(f"Generated {len(summaries)} summary tables")
    return summaries


# =============================================================================
# ANALYTICAL INSIGHT SHEETS
# =============================================================================

def build_content_to_optimize_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the 'Content to Optimize' analytical sheet.

    Identifies pages with value but underperforming that could benefit from
    optimization. These are pages that have search visibility but aren't
    converting that visibility into clicks efficiently.

    Selection Criteria:
    1. High impressions but low CTR (impressions >= 100 and CTR < 5%)
    2. Ranking positions 5-20 (striking distance to page 1)
    3. On-page issues (missing meta description, thin content, etc.)
    4. Status code 200 (live pages only)
    5. Indexable pages only

    Args:
        df: Processed URL-level DataFrame with all metrics

    Returns:
        DataFrame with content optimization opportunities, sorted by priority
    """
    # Filter to live, indexable pages
    mask = (
        (df['status_code'] == 200) &
        (df['indexable'] == True)
    )

    candidates = df[mask].copy()

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Avg Position', 'Impressions', 'Clicks', 'CTR (%)',
            'Word Count', 'Has Meta Description', 'Sessions', 'Optimization Signals', 'Priority Score'
        ])

    # Calculate optimization signals for each page
    optimization_signals = []
    priority_scores = []

    for _, row in candidates.iterrows():
        signals = []
        score = 0

        # Signal 1: High impressions, low CTR (CTR below 5% with significant impressions)
        if row['impressions'] >= 100 and row['ctr'] < 0.05:
            signals.append('Low CTR despite impressions')
            score += 3
        elif row['impressions'] >= 50 and row['ctr'] < 0.03:
            signals.append('Very low CTR')
            score += 4

        # Signal 2: Striking distance (position 5-20)
        if 5 <= row['avg_position'] <= 10:
            signals.append('Position 5-10 (near page 1)')
            score += 5
        elif 11 <= row['avg_position'] <= 20:
            signals.append('Position 11-20 (striking distance)')
            score += 3

        # Signal 3: Missing or thin meta description
        meta_desc = str(row.get('meta_description', '')).strip()
        if not meta_desc:
            signals.append('Missing meta description')
            score += 2
        elif len(meta_desc) < 70:
            signals.append('Short meta description')
            score += 1

        # Signal 4: Thin content (below 1000 words for substantial pages)
        if row['word_count'] < 500:
            signals.append('Very thin content (<500 words)')
            score += 3
        elif row['word_count'] < 1000:
            signals.append('Thin content (<1000 words)')
            score += 2

        # Signal 5: Has traffic but could have more
        if row['sessions'] > 0 and row['avg_position'] > 5:
            signals.append('Has traffic, room to improve')
            score += 1

        # Signal 6: High impressions indicates search demand
        if row['impressions'] >= 500:
            signals.append('High search demand')
            score += 2
        elif row['impressions'] >= 100:
            signals.append('Moderate search demand')
            score += 1

        optimization_signals.append('; '.join(signals) if signals else 'No specific signals')
        priority_scores.append(score)

    candidates['Optimization Signals'] = optimization_signals
    candidates['Priority Score'] = priority_scores

    # Filter to only pages with at least one optimization signal (score > 0)
    candidates = candidates[candidates['Priority Score'] > 0]

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Avg Position', 'Impressions', 'Clicks', 'CTR (%)',
            'Word Count', 'Has Meta Description', 'Sessions', 'Optimization Signals', 'Priority Score'
        ])

    # Sort by priority score descending and cap at top 20
    candidates = candidates.sort_values('Priority Score', ascending=False).head(20)

    # Build output DataFrame with selected columns
    result = pd.DataFrame({
        'URL': candidates['url'],
        'Page Type': candidates['page_type'],
        'Avg Position': candidates['avg_position'].round(1),
        'Impressions': candidates['impressions'].astype(int),
        'Clicks': candidates['clicks'].astype(int),
        'CTR (%)': (candidates['ctr'] * 100).round(2),
        'Word Count': candidates['word_count'].astype(int),
        'Has Meta Description': candidates['meta_description'].apply(lambda x: 'Yes' if str(x).strip() else 'No'),
        'Sessions': candidates['sessions'].astype(int),
        'Optimization Signals': candidates['Optimization Signals'],
        'Priority Score': candidates['Priority Score']
    })

    return result


def build_thin_content_sheet(df: pd.DataFrame, thin_threshold: int = 1000) -> pd.DataFrame:
    """
    Build the 'Thin Content Opportunities' analytical sheet.

    Identifies pages lacking depth that could benefit from content expansion.
    These are indexable pages with low word counts that have some search
    visibility or traffic potential.

    Selection Criteria:
    1. Word count below threshold (default 1000 words)
    2. Page is indexable and status 200
    3. Has some value signal (impressions, sessions, or backlinks)
    4. Not already flagged for deletion

    Args:
        df: Processed URL-level DataFrame with all metrics
        thin_threshold: Word count threshold for thin content (default 1000)

    Returns:
        DataFrame with thin content opportunities, sorted by potential value
    """
    # Filter to live, indexable pages with thin content
    mask = (
        (df['status_code'] == 200) &
        (df['indexable'] == True) &
        (df['word_count'] < thin_threshold)
    )

    candidates = df[mask].copy()

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Word Count', 'Content Gap', 'Sessions',
            'Impressions', 'Avg Position', 'Referring Domains', 'Opportunity Type', 'Value Score'
        ])

    # Calculate value score and opportunity type
    opportunity_types = []
    value_scores = []

    for _, row in candidates.iterrows():
        opp_type = []
        score = 0

        # Value from traffic
        if row['sessions'] >= 10:
            opp_type.append('Has significant traffic')
            score += 5
        elif row['sessions'] > 0:
            opp_type.append('Has some traffic')
            score += 2

        # Value from search visibility
        if row['impressions'] >= 100:
            opp_type.append('High search visibility')
            score += 4
        elif row['impressions'] > 0:
            opp_type.append('Some search visibility')
            score += 1

        # Value from rankings (already ranking = easier to improve)
        if 0 < row['avg_position'] <= 20:
            opp_type.append('Already ranking (pos <= 20)')
            score += 3
        elif 20 < row['avg_position'] <= 50:
            opp_type.append('Has rankings (pos 20-50)')
            score += 1

        # Value from backlinks (has authority)
        if row['referring_domains'] >= 5:
            opp_type.append('Strong backlink profile')
            score += 4
        elif row['referring_domains'] > 0:
            opp_type.append('Has backlinks')
            score += 2

        # Strategic page types get bonus
        if row['page_type'] in ['Service', 'Local Lander', 'Home']:
            opp_type.append(f'Strategic page ({row["page_type"]})')
            score += 3

        # Very thin content penalty/opportunity marker
        if row['word_count'] < 300:
            opp_type.append('Severely thin (<300 words)')
        elif row['word_count'] < 500:
            opp_type.append('Very thin (<500 words)')

        opportunity_types.append('; '.join(opp_type) if opp_type else 'Low priority')
        value_scores.append(score)

    candidates['Opportunity Type'] = opportunity_types
    candidates['Value Score'] = value_scores

    # Filter to pages with some value signal (score > 0)
    candidates = candidates[candidates['Value Score'] > 0]

    if len(candidates) == 0:
        return pd.DataFrame(columns=[
            'URL', 'Page Type', 'Word Count', 'Content Gap', 'Sessions',
            'Impressions', 'Avg Position', 'Referring Domains', 'Opportunity Type', 'Value Score'
        ])

    # Sort by value score descending
    candidates = candidates.sort_values('Value Score', ascending=False)

    # Calculate content gap (words needed to reach threshold)
    content_gap = thin_threshold - candidates['word_count']

    # Build output DataFrame
    result = pd.DataFrame({
        'URL': candidates['url'],
        'Page Type': candidates['page_type'],
        'Word Count': candidates['word_count'].astype(int),
        'Content Gap': content_gap.astype(int),
        'Sessions': candidates['sessions'].astype(int),
        'Impressions': candidates['impressions'].astype(int),
        'Avg Position': candidates['avg_position'].round(1),
        'Referring Domains': candidates['referring_domains'].astype(int),
        'Opportunity Type': candidates['Opportunity Type'],
        'Value Score': candidates['Value Score']
    })

    return result


# =============================================================================
# BUSINESS RELEVANCE & FEASIBILITY FILTERING
# =============================================================================
# These functions filter and score topics to ensure only realistic, sensible,
# on-brand opportunities are surfaced. A topic must pass all three gates:
# 1. Search Demand Gate (handled by clustering)
# 2. SEO Feasibility Gate (handled by action recommendation)
# 3. Business Relevance Gate (NEW - this module)

# Exclusion terms - topics containing these are automatically excluded
NAVIGATIONAL_TERMS = {
    'login', 'signin', 'sign in', 'signup', 'sign up', 'logout', 'log out',
    'register', 'forgot password', 'reset password', 'my account', 'dashboard',
    'admin', 'portal', 'intranet', 'employee', 'staff only'
}

POLICY_LEGAL_TERMS = {
    'privacy policy', 'terms of service', 'terms and conditions', 'cookie policy',
    'gdpr', 'disclaimer', 'legal notice', 'refund policy', 'return policy',
    'shipping policy', 'cancellation policy', 'accessibility statement'
}

JUNK_ARTIFACT_TERMS = {
    'utm_', 'utm=', '?p=', '?s=', 'index.php', 'page=', 'amp', '&amp',
    'session', 'sessionid', 'jsessionid', 'phpsessid', 'cfid', 'cftoken',
    'www.', 'http://', 'https://', '.html', '.php', '.aspx', '.jsp'
}

AUTHORITY_REQUIRED_TERMS = {
    # Medical - requires licensed practitioner
    'diagnosis', 'treatment plan', 'medical advice', 'prescription', 'dosage',
    'symptoms of', 'cure for', 'medication for', 'drug interaction',
    # Legal - requires licensed attorney
    'legal advice', 'attorney', 'lawsuit', 'sue for', 'liability for',
    'contract law', 'legal rights', 'file a claim',
    # Financial - requires licensed advisor
    'investment advice', 'tax advice', 'financial planning', 'retirement planning',
    'stock picks', 'crypto investment', 'guaranteed returns',
    # Government - requires official authority
    'official form', 'government application', 'visa application', 'passport renewal'
}

# Intent classification modifiers
TRANSACTIONAL_MODIFIERS = {
    'pricing', 'price', 'cost', 'quote', 'buy', 'purchase', 'order', 'hire',
    'service', 'services', 'provider', 'providers', 'company', 'companies',
    'contractor', 'contractors', 'professional', 'professionals', 'near me',
    'in my area', 'local', 'get', 'schedule', 'book', 'appointment'
}

COMMERCIAL_MODIFIERS = {
    'best', 'top', 'review', 'reviews', 'compare', 'comparison', 'vs', 'versus',
    'alternative', 'alternatives', 'software', 'tool', 'tools', 'solution',
    'solutions', 'platform', 'app', 'rated', 'recommended', 'affordable'
}

INFORMATIONAL_MODIFIERS = {
    'how', 'what', 'why', 'when', 'where', 'guide', 'tutorial', 'tips',
    'steps', 'process', 'explained', 'meaning', 'definition', 'example',
    'examples', 'learn', 'understand', 'basics', 'introduction', 'overview'
}


def calculate_business_relevance_score(
    topic_tokens: set,
    primary_query: str,
    site_context: dict = None
) -> tuple:
    """
    Calculate Business Relevance Score (0-100) for a topic cluster.

    Scoring components:
    - Site Theme Alignment (40%): Overlap with existing site content/themes
    - Commercial/Strategic Intent (30%): Transactional > Commercial > Informational
    - Content Feasibility (20%): Can the business reasonably create this content?
    - Brand Safety/Nonsense Filter (10%): No junk, artifacts, or malformed queries

    Args:
        topic_tokens: Set of tokens from the topic cluster
        primary_query: The primary/seed query for the topic
        site_context: Dict containing site theme data:
            - 'existing_urls': List of existing site URLs
            - 'existing_titles': List of existing page titles
            - 'top_queries': List of top-performing queries
            - 'site_categories': List of site category terms (services, products, etc.)
            - 'brand_terms': List of brand/company terms

    Returns:
        Tuple of (score: int 0-100, notes: str, excluded: bool, exclusion_reason: str)
    """
    if site_context is None:
        site_context = {}

    query_lower = primary_query.lower()
    notes = []
    exclusion_reason = ""

    # =========================================================================
    # HARD EXCLUSION CHECKS (Brand Safety / Nonsense Filter)
    # =========================================================================

    # Check for navigational intent (login, account pages)
    for term in NAVIGATIONAL_TERMS:
        if term in query_lower:
            return (0, "", True, f"Navigational intent: '{term}'")

    # Check for policy/legal pages
    for term in POLICY_LEGAL_TERMS:
        if term in query_lower:
            return (0, "", True, f"Policy/legal page: '{term}'")

    # Check for junk artifacts (URL parameters, session IDs)
    for term in JUNK_ARTIFACT_TERMS:
        if term in query_lower:
            return (0, "", True, f"Junk/artifact term: '{term}'")

    # Check for queries requiring professional authority
    for term in AUTHORITY_REQUIRED_TERMS:
        if term in query_lower:
            return (0, "", True, f"Requires professional authority: '{term}'")

    # Check for malformed queries (too short, all numbers, etc.)
    clean_query = ''.join(c for c in query_lower if c.isalnum() or c.isspace())
    if len(clean_query.strip()) < 5:
        return (0, "", True, "Query too short/malformed")

    if clean_query.replace(' ', '').isdigit():
        return (0, "", True, "Query is only numbers")

    # =========================================================================
    # COMPONENT 1: Site Theme Alignment (40%)
    # =========================================================================
    theme_score = 0
    max_theme_score = 40

    existing_urls = set(str(u).lower() for u in site_context.get('existing_urls', []))
    existing_titles = set(str(t).lower() for t in site_context.get('existing_titles', []))
    top_queries = set(str(q).lower() for q in site_context.get('top_queries', []))
    site_categories = set(str(c).lower() for c in site_context.get('site_categories', []))

    # Check token overlap with existing content
    all_site_text = ' '.join(existing_urls) + ' ' + ' '.join(existing_titles) + ' ' + ' '.join(top_queries)
    site_tokens = set(all_site_text.split())

    if topic_tokens and site_tokens:
        overlap = len(topic_tokens & site_tokens)
        overlap_ratio = overlap / len(topic_tokens) if topic_tokens else 0

        if overlap_ratio >= 0.6:
            theme_score = max_theme_score  # Strong alignment
        elif overlap_ratio >= 0.4:
            theme_score = int(max_theme_score * 0.75)  # Good alignment
        elif overlap_ratio >= 0.2:
            theme_score = int(max_theme_score * 0.5)  # Moderate alignment
        elif overlap_ratio > 0:
            theme_score = int(max_theme_score * 0.25)  # Weak alignment
        else:
            theme_score = 0
            notes.append("No site theme overlap")

    # Bonus for matching site categories
    for category in site_categories:
        if category in query_lower:
            theme_score = min(max_theme_score, theme_score + 10)
            break

    # =========================================================================
    # COMPONENT 2: Commercial/Strategic Intent (30%)
    # =========================================================================
    intent_score = 0
    max_intent_score = 30

    # Check for transactional intent (highest value)
    transactional_matches = sum(1 for term in TRANSACTIONAL_MODIFIERS if term in query_lower)
    commercial_matches = sum(1 for term in COMMERCIAL_MODIFIERS if term in query_lower)
    informational_matches = sum(1 for term in INFORMATIONAL_MODIFIERS if term in query_lower)

    if transactional_matches > 0:
        intent_score = max_intent_score  # Full score for transactional
    elif commercial_matches > 0:
        intent_score = int(max_intent_score * 0.7)  # 70% for commercial
    elif informational_matches > 0:
        # Informational allowed only if supports commercial area
        if theme_score >= max_theme_score * 0.5:
            intent_score = int(max_intent_score * 0.4)  # 40% if supports site themes
        else:
            intent_score = int(max_intent_score * 0.15)  # 15% standalone informational
            notes.append("Informational intent with weak commercial tie-in")
    else:
        intent_score = int(max_intent_score * 0.3)  # Neutral intent

    # =========================================================================
    # COMPONENT 3: Content Feasibility (20%)
    # =========================================================================
    feasibility_score = 20  # Start with full score, deduct for issues
    max_feasibility_score = 20

    # Check for topics that would require expertise the business likely doesn't have
    questionable_expertise = {
        'diy', 'homemade', 'home remedy', 'self-diagnose', 'without professional',
        'instead of doctor', 'instead of lawyer', 'free legal', 'free medical'
    }

    for term in questionable_expertise:
        if term in query_lower:
            feasibility_score = int(feasibility_score * 0.5)
            notes.append(f"Questionable expertise required: '{term}'")
            break

    # Topics with competitor brand names (unless comparison content)
    brand_terms = set(str(b).lower() for b in site_context.get('brand_terms', []))
    competitor_indicators = ['competitor', 'vs', 'versus', 'compare', 'alternative']

    # If query contains what looks like a competitor name but isn't a comparison
    # This is a heuristic - proper nouns that aren't in brand_terms
    words = query_lower.split()
    potential_brands = [w for w in words if w.isalpha() and len(w) > 3 and w not in brand_terms]

    is_comparison = any(ind in query_lower for ind in competitor_indicators)
    if not is_comparison and potential_brands:
        # Don't penalize too hard, just note it
        pass  # Could add logic here if needed

    # =========================================================================
    # COMPONENT 4: Brand Safety / Nonsense Filter (10%)
    # =========================================================================
    # Already handled in hard exclusions above, remaining 10% is baseline
    safety_score = 10

    # Slight penalty for very long queries (often spam/keyword stuffing)
    if len(query_lower.split()) > 8:
        safety_score = int(safety_score * 0.7)
        notes.append("Query unusually long")

    # Penalty for repeated words (spam indicator)
    word_list = query_lower.split()
    if len(word_list) > len(set(word_list)) + 2:
        safety_score = int(safety_score * 0.5)
        notes.append("Repeated words in query")

    # =========================================================================
    # CALCULATE FINAL SCORE
    # =========================================================================
    total_score = theme_score + intent_score + feasibility_score + safety_score

    # Normalize to 0-100
    total_score = max(0, min(100, total_score))

    # Check if score meets minimum threshold
    if total_score < 50:
        exclusion_reason = f"Low business relevance score ({total_score}/100)"
        return (total_score, '; '.join(notes) if notes else '', True, exclusion_reason)

    return (total_score, '; '.join(notes) if notes else '', False, '')


def build_site_context_from_df(df: pd.DataFrame, gsc_df: pd.DataFrame = None) -> dict:
    """
    Build site context dictionary from crawl and GSC data for business relevance scoring.

    Args:
        df: URL-level DataFrame with crawl data
        gsc_df: Optional GSC query-level data

    Returns:
        Dict with site context for relevance scoring
    """
    context = {
        'existing_urls': [],
        'existing_titles': [],
        'top_queries': [],
        'site_categories': [],
        'brand_terms': []
    }

    if df is not None and len(df) > 0:
        # Extract URLs
        if 'url' in df.columns:
            context['existing_urls'] = df['url'].dropna().tolist()

        # Extract page titles
        if 'page_title' in df.columns:
            context['existing_titles'] = df['page_title'].dropna().tolist()

        # Extract page types as categories
        if 'page_type' in df.columns:
            categories = df['page_type'].dropna().unique().tolist()
            # Also extract common terms from page types
            for cat in categories:
                context['site_categories'].extend(str(cat).lower().split())
            context['site_categories'] = list(set(context['site_categories']))

    if gsc_df is not None and len(gsc_df) > 0:
        # Get top queries by impressions
        query_col = 'query' if 'query' in gsc_df.columns else 'primary_keyword'
        if query_col in gsc_df.columns and 'impressions' in gsc_df.columns:
            top_queries_df = gsc_df.nlargest(100, 'impressions')
            context['top_queries'] = top_queries_df[query_col].dropna().tolist()
        elif query_col in gsc_df.columns:
            context['top_queries'] = gsc_df[query_col].dropna().head(100).tolist()

    # Extract potential brand terms from URL patterns (e.g., domain name)
    if context['existing_urls']:
        from urllib.parse import urlparse
        try:
            sample_url = context['existing_urls'][0]
            parsed = urlparse(sample_url)
            domain_parts = parsed.netloc.replace('www.', '').split('.')
            if domain_parts:
                context['brand_terms'] = [domain_parts[0]]
        except:
            pass

    return context


def filter_topics_by_business_relevance(
    topics_df: pd.DataFrame,
    site_context: dict = None,
    min_score: int = 50,
    max_topics: int = 15
) -> tuple:
    """
    Filter and score topics by business relevance.

    Args:
        topics_df: DataFrame with topic opportunities
        site_context: Site context for relevance scoring
        min_score: Minimum business relevance score to include (default: 50)
        max_topics: Maximum number of topics to return (default: 15)

    Returns:
        Tuple of (filtered_df, excluded_topics_list)
    """
    if topics_df is None or len(topics_df) == 0:
        return topics_df, []

    if site_context is None:
        site_context = {}

    # Calculate business relevance for each topic
    relevance_scores = []
    feasibility_notes = []
    excluded_topics = []

    for _, row in topics_df.iterrows():
        primary_query = str(row.get('Primary Keyword', '')).strip()
        secondary_keywords = str(row.get('Secondary Keywords', '')).strip()

        # Combine tokens from primary and secondary keywords
        all_keywords = primary_query + ' ' + secondary_keywords
        tokens = set(all_keywords.lower().split())

        score, notes, excluded, reason = calculate_business_relevance_score(
            topic_tokens=tokens,
            primary_query=primary_query,
            site_context=site_context
        )

        if excluded:
            excluded_topics.append({
                'Topic': row.get('Suggested Topic', primary_query),
                'Primary Keyword': primary_query,
                'Score': score,
                'Reason': reason
            })
            relevance_scores.append(0)
            feasibility_notes.append(reason)
        else:
            relevance_scores.append(score)
            feasibility_notes.append(notes if notes else '')

    # Add new columns
    topics_df = topics_df.copy()
    topics_df['Business Relevance Score'] = relevance_scores
    topics_df['Feasibility Notes'] = feasibility_notes

    # Filter out excluded topics (score = 0 or below threshold)
    mask = topics_df['Business Relevance Score'] >= min_score
    filtered_df = topics_df[mask].copy()

    # Sort by Priority Score first, then Business Relevance Score
    if len(filtered_df) > 0:
        filtered_df = filtered_df.sort_values(
            ['Priority Score', 'Business Relevance Score'],
            ascending=[False, False]
        )

        # Cap to max_topics
        filtered_df = filtered_df.head(max_topics)

    # Log filtering stats
    logger.info(f"[BUSINESS RELEVANCE] Topics before filtering: {len(topics_df)}")
    logger.info(f"[BUSINESS RELEVANCE] Topics after filtering: {len(filtered_df)}")
    logger.info(f"[BUSINESS RELEVANCE] Topics excluded: {len(excluded_topics)}")

    if excluded_topics:
        exclusion_reasons = {}
        for ex in excluded_topics:
            reason = ex['Reason'].split(':')[0] if ':' in ex['Reason'] else ex['Reason']
            exclusion_reasons[reason] = exclusion_reasons.get(reason, 0) + 1
        logger.info(f"[BUSINESS RELEVANCE] Exclusion breakdown: {exclusion_reasons}")

    return filtered_df, excluded_topics


def cluster_queries_into_topics(queries_df: pd.DataFrame, max_topics: int = 20, thin_urls: set = None, url_word_counts: dict = None) -> pd.DataFrame:
    """
    Cluster similar queries into topic groups for new content opportunities.

    Uses a simple word-overlap algorithm to group queries that share
    significant terms, creating topic clusters from individual queries.

    Implements Content Action Recommendation logic:
    - Create new page: No dominant page, content gap exists
    - Expand existing page: One page ranks but needs depth/optimization
    - Consolidate pages: Multiple pages cannibalize each other

    Args:
        queries_df: DataFrame with query-level data (query, impressions, clicks, avg_position, url)
        max_topics: Maximum number of topic clusters to return
        thin_urls: Set of URLs flagged as thin content (optional)
        url_word_counts: Dict mapping URL -> word count (optional, for expand detection)

    Returns:
        DataFrame with topic clusters including Recommended Action column
    """
    import re
    from collections import defaultdict

    if queries_df is None or len(queries_df) == 0:
        return pd.DataFrame()

    if thin_urls is None:
        thin_urls = set()

    if url_word_counts is None:
        url_word_counts = {}

    # Normalize and tokenize queries
    def tokenize(query):
        query = str(query).lower().strip()
        # Remove common stopwords and punctuation
        stopwords = {'a', 'an', 'the', 'in', 'on', 'at', 'to', 'for', 'of', 'and', 'or', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could', 'should', 'may', 'might', 'must', 'shall', 'can', 'need', 'what', 'which', 'who', 'whom', 'this', 'that', 'these', 'those', 'am', 'with', 'as', 'by', 'from', 'how', 'when', 'where', 'why', 'if', 'then', 'so', 'than', 'too', 'very', 'just', 'only', 'also', 'even', 'more', 'most', 'other', 'into', 'over', 'after', 'before', 'between', 'through', 'during', 'under', 'around', 'about', 'near'}
        words = re.findall(r'\b[a-z]{2,}\b', query)
        return [w for w in words if w not in stopwords]

    # Build query data with tokens
    query_data = []
    for _, row in queries_df.iterrows():
        query = str(row.get('query', row.get('primary_keyword', ''))).strip()
        if not query or query == 'nan':
            continue
        tokens = tokenize(query)
        if tokens:
            query_data.append({
                'query': query,
                'tokens': set(tokens),
                'impressions': int(row.get('impressions', 0)),
                'clicks': int(row.get('clicks', 0)),
                'avg_position': float(row.get('avg_position', 0)),
                'url': str(row.get('url', ''))
            })

    if not query_data:
        return pd.DataFrame()

    # Sort by impressions to prioritize high-volume queries as cluster seeds
    query_data.sort(key=lambda x: x['impressions'], reverse=True)

    # Cluster queries using greedy word-overlap approach
    clusters = []
    used_queries = set()

    for seed in query_data:
        if seed['query'] in used_queries:
            continue

        # Start new cluster with this seed
        cluster = {
            'primary_query': seed['query'],
            'queries': [seed],
            'all_tokens': seed['tokens'].copy(),
            'total_impressions': seed['impressions'],
            'total_clicks': seed['clicks'],
            'positions': [seed['avg_position']] if seed['avg_position'] > 0 else [],
            'urls': defaultdict(lambda: {'impressions': 0, 'positions': [], 'clicks': 0})
        }
        # Track URL-level metrics for cannibalization detection
        if seed['url']:
            cluster['urls'][seed['url']]['impressions'] += seed['impressions']
            cluster['urls'][seed['url']]['clicks'] += seed['clicks']
            if seed['avg_position'] > 0:
                cluster['urls'][seed['url']]['positions'].append(seed['avg_position'])
        used_queries.add(seed['query'])

        # Find similar queries (at least 50% token overlap with seed)
        for candidate in query_data:
            if candidate['query'] in used_queries:
                continue

            # Calculate Jaccard-like similarity
            overlap = len(seed['tokens'] & candidate['tokens'])
            min_len = min(len(seed['tokens']), len(candidate['tokens']))

            if min_len > 0 and overlap / min_len >= 0.5:
                cluster['queries'].append(candidate)
                cluster['all_tokens'].update(candidate['tokens'])
                cluster['total_impressions'] += candidate['impressions']
                cluster['total_clicks'] += candidate['clicks']
                if candidate['avg_position'] > 0:
                    cluster['positions'].append(candidate['avg_position'])
                # Track URL-level metrics
                if candidate['url']:
                    cluster['urls'][candidate['url']]['impressions'] += candidate['impressions']
                    cluster['urls'][candidate['url']]['clicks'] += candidate['clicks']
                    if candidate['avg_position'] > 0:
                        cluster['urls'][candidate['url']]['positions'].append(candidate['avg_position'])
                used_queries.add(candidate['query'])

        clusters.append(cluster)

    # Filter and score clusters with Content Action Recommendation logic
    topic_opportunities = []

    # DEBUG: Log clustering pipeline statistics
    logger.info(f"[CLUSTER DEBUG] Total clusters formed: {len(clusters)}")
    clusters_skipped_low_volume = 0
    clusters_with_dominant_page = 0
    clusters_processed = 0
    action_distribution = {'Create new page': 0, 'Expand existing page': 0, 'Consolidate pages': 0}

    for cluster in clusters:
        # Skip clusters with only 1 query and low impressions
        if len(cluster['queries']) == 1 and cluster['total_impressions'] < 100:
            clusters_skipped_low_volume += 1
            continue

        # Calculate weighted average position
        if cluster['positions']:
            weighted_pos = sum(
                q['avg_position'] * q['impressions']
                for q in cluster['queries']
                if q['avg_position'] > 0
            )
            total_weight = sum(
                q['impressions']
                for q in cluster['queries']
                if q['avg_position'] > 0
            )
            avg_position = weighted_pos / total_weight if total_weight > 0 else 0
        else:
            avg_position = 0

        # =====================================================================
        # URL-LEVEL METRICS ANALYSIS
        # =====================================================================
        url_metrics = cluster['urls']
        num_urls = len(url_metrics)

        # Calculate per-URL metrics with CTR
        url_analysis = []
        for url, metrics in url_metrics.items():
            url_impressions = metrics['impressions']
            url_clicks = metrics['clicks']
            url_positions = metrics['positions']
            url_avg_pos = sum(url_positions) / len(url_positions) if url_positions else 0
            url_ctr = (url_clicks / url_impressions * 100) if url_impressions > 0 else 0
            url_share = url_impressions / cluster['total_impressions'] if cluster['total_impressions'] > 0 else 0
            url_word_count = url_word_counts.get(url, 0)
            is_thin = url in thin_urls or url_word_count < 1000

            url_analysis.append({
                'url': url,
                'impressions': url_impressions,
                'clicks': url_clicks,
                'avg_position': url_avg_pos,
                'ctr': url_ctr,
                'share': url_share,
                'word_count': url_word_count,
                'is_thin': is_thin
            })

        # Sort by impressions (strongest first)
        url_analysis.sort(key=lambda x: x['impressions'], reverse=True)

        # Identify dominant page (if any)
        dominant_url = None
        dominant_data = None
        if url_analysis:
            top_url = url_analysis[0]
            if top_url['share'] >= 0.60 and top_url['avg_position'] <= 15:
                # Check if it meets dominant criteria (not thin)
                if not top_url['is_thin']:
                    dominant_url = top_url['url']
                    dominant_data = top_url

        # =====================================================================
        # CONTENT ACTION RECOMMENDATION LOGIC
        # =====================================================================
        recommended_action = ""
        primary_url = ""
        secondary_urls = ""
        reasoning = ""

        # Track clusters with dominant pages for debugging (but DO NOT exclude them)
        if dominant_url and dominant_data:
            clusters_with_dominant_page += 1
            logger.debug(f"[CLUSTER DEBUG] Topic '{cluster['primary_query']}' has dominant page {dominant_url} "
                        f"({dominant_data['share']:.0%} impressions at position {dominant_data['avg_position']:.1f}) - evaluating for action")
            # NOTE: Previously this would `continue` and exclude the cluster.
            # Now we flow through to assign an appropriate action (usually Expand existing page)
            # This ensures cannibalization detection can still happen for edge cases

        # DECISION LOGIC
        if num_urls == 0:
            # No pages at all - Create new page
            recommended_action = "Create new page"
            primary_url = ""
            secondary_urls = ""
            reasoning = f"Search demand exists ({cluster['total_impressions']:,} impressions) but no landing page currently serves these queries. This is a true content gap."

        elif num_urls == 1:
            # Single URL - check if it needs expansion or if we should create new
            single_url = url_analysis[0]

            if single_url['avg_position'] > 20:
                # Ranking very poorly - Create new page
                recommended_action = "Create new page"
                primary_url = ""
                secondary_urls = ""
                reasoning = f"Existing page ({single_url['url'][:60]}...) ranks poorly at position {single_url['avg_position']:.0f}. A dedicated, well-optimized page would better serve this topic."

            elif single_url['is_thin']:
                # Thin content - Expand
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page exists but has thin content ({single_url['word_count']} words). Expanding with comprehensive content at position {single_url['avg_position']:.0f} can capture more traffic."

            elif single_url['ctr'] < 2.0 and single_url['impressions'] >= 100:
                # Low CTR despite impressions - Expand/Optimize
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} with high impressions ({single_url['impressions']:,}) but very low CTR ({single_url['ctr']:.1f}%). Improve content depth and meta data to boost clicks."

            elif single_url['avg_position'] > 10:
                # Moderate ranking - Expand to improve
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} but not in top 10. Expanding content depth can improve rankings and capture more traffic."

            else:
                # Good ranking but included because not dominant - still suggest expansion
                recommended_action = "Expand existing page"
                primary_url = single_url['url']
                secondary_urls = ""
                reasoning = f"Page ranks at position {single_url['avg_position']:.0f} but doesn't fully capture the topic. Expanding with related subtopics can increase topical authority."

        elif num_urls == 2:
            # Two URLs - check for consolidation vs expansion
            url1, url2 = url_analysis[0], url_analysis[1]

            # Consolidation triggers (loosened for better cannibalization detection):
            # 1. Neither URL exceeds 60% dominance (true competition)
            # 2. Second URL has at least 12% share (meaningful competitor)
            # 3. URLs rank within 5 positions of each other (close competition)
            no_clear_dominant = url1['share'] < 0.60
            meaningful_second = url2['share'] >= 0.12
            close_rankings = abs(url1['avg_position'] - url2['avg_position']) <= 5 and url2['avg_position'] > 0

            if no_clear_dominant or (meaningful_second and close_rankings):
                # Both URLs competing - Consolidate
                recommended_action = "Consolidate pages"
                primary_url = url1['url']
                secondary_urls = url2['url']
                reasoning = f"Two pages are competing for the same keywords, diluting rankings. {url1['url'][:50]}... has {url1['share']:.0%} impressions at position {url1['avg_position']:.0f}; {url2['url'][:50]}... has {url2['share']:.0%}. Consolidate into one authoritative page."
            else:
                # One URL clearly dominant among the two - Expand it
                recommended_action = "Expand existing page"
                primary_url = url1['url']
                secondary_urls = ""
                reasoning = f"One page ({url1['url'][:50]}...) receives most impressions ({url1['share']:.0%}) at position {url1['avg_position']:.0f}. Expand this page to fully own the topic."

        else:
            # 3+ URLs - Consolidation needed
            # Select top URLs for consolidation (max 3)
            top_urls = url_analysis[:3]
            meaningful_urls = [u for u in top_urls if u['share'] >= 0.08]  # Lowered from 0.10 to catch more competition

            # Additional consolidation trigger: close rankings between top URLs
            close_ranking_urls = []
            if len(url_analysis) >= 2:
                best_pos = url_analysis[0]['avg_position']
                for u in url_analysis[1:3]:  # Check top 3
                    if u['avg_position'] > 0 and abs(u['avg_position'] - best_pos) <= 5:
                        close_ranking_urls.append(u)

            # Consolidate if: 2+ URLs have meaningful share OR 2+ URLs rank closely
            should_consolidate = len(meaningful_urls) >= 2 or (len(close_ranking_urls) >= 1 and url_analysis[0]['share'] < 0.70)

            if should_consolidate:
                # Build list of URLs to consolidate (prefer meaningful share, include close rankers)
                consolidate_urls = list(meaningful_urls) if len(meaningful_urls) >= 2 else [url_analysis[0]] + close_ranking_urls[:2]
                consolidate_urls = consolidate_urls[:3]  # Cap at 3

                recommended_action = "Consolidate pages"
                primary_url = consolidate_urls[0]['url']
                secondary_urls = ', '.join([u['url'] for u in consolidate_urls[1:]])
                url_summary = '; '.join([f"{u['url'][:40]}... ({u['share']:.0%})" for u in consolidate_urls])
                reasoning = f"Traffic is fragmented across {num_urls} pages causing keyword cannibalization. Top competing pages: {url_summary}. Consolidate into one authoritative hub."
            elif url_analysis[0]['avg_position'] > 20:
                # Fragmented but all ranking poorly - Create new
                recommended_action = "Create new page"
                primary_url = ""
                secondary_urls = ""
                reasoning = f"Traffic is split across {num_urls} weak pages (best position: {url_analysis[0]['avg_position']:.0f}). Creating a dedicated, comprehensive page would better serve this topic than consolidating poor performers."
            else:
                # One page somewhat dominant among many - Expand it
                recommended_action = "Expand existing page"
                primary_url = url_analysis[0]['url']
                secondary_urls = ""
                reasoning = f"Multiple pages exist but {url_analysis[0]['url'][:50]}... leads with {url_analysis[0]['share']:.0%} impressions. Expand this page to consolidate authority."

        # Generate suggested topic from tokens
        primary_query = cluster['primary_query']
        common_tokens = sorted(cluster['all_tokens'], key=lambda t: sum(1 for q in cluster['queries'] if t in q['tokens']), reverse=True)
        suggested_topic = ' '.join(common_tokens[:4]).title() if common_tokens else primary_query.title()

        # Determine page type suggestion based on query intent
        query_lower = primary_query.lower()
        intent_value = 0.5
        if any(word in query_lower for word in ['how to', 'guide', 'tutorial', 'tips', 'steps']):
            page_type = 'Guide / How-To'
            intent_value = 0.7
        elif any(word in query_lower for word in ['best', 'top', 'review', 'compare', 'vs']):
            page_type = 'Comparison / Review'
            intent_value = 0.9
        elif any(word in query_lower for word in ['what is', 'meaning', 'definition', 'explain']):
            page_type = 'Educational / Explainer'
            intent_value = 0.6
        elif any(word in query_lower for word in ['near me', 'in ', 'local']):
            page_type = 'Local Landing Page'
            intent_value = 1.0
        elif any(word in query_lower for word in ['cost', 'price', 'pricing', 'quote', 'estimate']):
            page_type = 'Service / Pricing Page'
            intent_value = 1.0
        elif any(word in query_lower for word in ['buy', 'order', 'purchase', 'shop']):
            page_type = 'Product / Service Page'
            intent_value = 1.0
        else:
            page_type = 'Blog Post / Article'
            intent_value = 0.5

        # Get secondary keywords
        secondary = [q['query'] for q in cluster['queries'] if q['query'] != primary_query][:8]

        # =====================================================================
        # PRIORITY SCORE CALCULATION
        # 40% Impressions, 25% Action urgency, 20% Position weakness, 15% Intent value
        # =====================================================================
        import math
        impressions_score = min(10, math.log10(max(1, cluster['total_impressions'])) * 2.5)

        # Action urgency score (25% weight)
        if recommended_action == "Create new page":
            action_urgency_score = 10 if num_urls == 0 else 7
        elif recommended_action == "Consolidate pages":
            action_urgency_score = 9  # High urgency - cannibalization hurts rankings
        else:  # Expand
            action_urgency_score = 5  # Lower urgency - existing page works somewhat

        # Position weakness score (20% weight)
        if avg_position == 0 or num_urls == 0:
            position_score = 10
        elif avg_position > 30:
            position_score = 9
        elif avg_position > 20:
            position_score = 7
        elif avg_position > 15:
            position_score = 5
        elif avg_position > 10:
            position_score = 3
        else:
            position_score = 1

        # Intent value score (15% weight)
        intent_score = intent_value * 10

        priority_score = round(
            (impressions_score * 0.40) +
            (action_urgency_score * 0.25) +
            (position_score * 0.20) +
            (intent_score * 0.15),
            1
        )

        # SAFETY: Ensure no cluster falls through without an action
        # This guarantees every qualifying cluster gets a recommendation
        if not recommended_action:
            # Default fallback - should rarely happen with above logic
            if num_urls == 0:
                recommended_action = "Create new page"
                reasoning = f"Search demand exists ({cluster['total_impressions']:,} impressions) with no current coverage. Create dedicated content."
            elif url_analysis:
                recommended_action = "Expand existing page"
                primary_url = url_analysis[0]['url']
                reasoning = f"Existing page could be expanded to better capture this topic's full potential."
            else:
                recommended_action = "Create new page"
                reasoning = f"Content gap identified - create new targeted content."
            logger.debug(f"[CLUSTER DEBUG] Applied fallback action '{recommended_action}' to topic '{cluster['primary_query']}'")

        # Track action distribution
        clusters_processed += 1
        if recommended_action in action_distribution:
            action_distribution[recommended_action] += 1

        topic_opportunities.append({
            'Suggested Topic': suggested_topic,
            'Primary Keyword': primary_query,
            'Secondary Keywords': ', '.join(secondary) if secondary else '',
            'Total Impressions': cluster['total_impressions'],
            'Avg Position': round(avg_position, 1) if avg_position > 0 else 'N/A',
            'Recommended Action': recommended_action,
            'Primary URL': primary_url,
            'Secondary URLs': secondary_urls,
            'Reasoning': reasoning,
            'Suggested Page Type': page_type,
            'Priority Score': priority_score
        })

    # DEBUG: Log pipeline summary
    logger.info(f"[CLUSTER DEBUG] Pipeline summary:")
    logger.info(f"  - Clusters formed: {len(clusters)}")
    logger.info(f"  - Skipped (low volume): {clusters_skipped_low_volume}")
    logger.info(f"  - With dominant page (now processed, not excluded): {clusters_with_dominant_page}")
    logger.info(f"  - Processed into opportunities: {clusters_processed}")
    logger.info(f"  - Action distribution: {action_distribution}")

    if not topic_opportunities:
        logger.warning("[CLUSTER DEBUG] No topic opportunities generated - all clusters were filtered or empty input")
        return pd.DataFrame()

    # Sort by priority score and limit to max_topics
    result = pd.DataFrame(topic_opportunities)
    result = result.sort_values('Priority Score', ascending=False).head(max_topics)

    return result


def build_new_content_opportunities_sheet(df: pd.DataFrame, gsc_df: Optional[pd.DataFrame] = None, thin_threshold: int = 1000) -> pd.DataFrame:
    """
    Build the 'New Content Opportunities' analytical sheet.

    Identifies content gaps as TOPICS (not URLs) - areas where there's search
    demand but no strong dedicated content exists.

    This sheet answers: "What should we write that doesn't exist on the site yet?"

    The approach:
    1. Analyze query-level GSC data (not aggregated URL data)
    2. Cluster similar queries into topic groups
    3. Detect cannibalization - exclude topics where a strong page already exists
    4. Flag cannibalization risks for topics with fragmented traffic
    5. Output suggested topics for new content creation

    Args:
        df: Processed URL-level DataFrame with all metrics
        gsc_df: Raw GSC query-level data (before URL aggregation)
        thin_threshold: Word count threshold for thin content identification

    Returns:
        DataFrame with new content TOPICS (not URLs), sorted by priority
    """
    # Define output columns for empty result (includes Content Action Recommendation columns and Business Relevance columns)
    empty_columns = [
        'Suggested Topic', 'Primary Keyword', 'Secondary Keywords', 'Total Impressions',
        'Avg Position', 'Recommended Action', 'Primary URL', 'Secondary URLs', 'Reasoning',
        'Suggested Page Type', 'Priority Score', 'Business Relevance Score', 'Feasibility Notes'
    ]

    # If no GSC data, return empty with message
    if gsc_df is None or len(gsc_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Build set of thin content URLs and URL word count mapping for action recommendations
    thin_urls = set()
    url_word_counts = {}
    if df is not None and len(df) > 0:
        word_count_col = None
        for col in ['word_count', 'Word Count', 'wordcount']:
            if col in df.columns:
                word_count_col = col
                break

        url_col = None
        for col in ['url', 'URL', 'Address']:
            if col in df.columns:
                url_col = col
                break

        if word_count_col and url_col:
            df_copy = df.copy()
            df_copy[word_count_col] = pd.to_numeric(df_copy[word_count_col], errors='coerce').fillna(0)
            thin_urls = set(df_copy[df_copy[word_count_col] < thin_threshold][url_col].tolist())
            # Build URL -> word count mapping for expand decision logic
            url_word_counts = dict(zip(df_copy[url_col], df_copy[word_count_col]))
            logger.debug(f"Identified {len(thin_urls)} thin content URLs and {len(url_word_counts)} URL word counts")

    # Ensure GSC data has required columns
    gsc_copy = gsc_df.copy()

    # Map column names if needed
    query_col = None
    for col in ['query', 'primary_keyword', 'Query', 'Keyword', 'Top queries']:
        if col in gsc_copy.columns:
            query_col = col
            break

    if query_col is None:
        # Try to use primary_keyword if available
        if 'primary_keyword' in gsc_copy.columns:
            gsc_copy['query'] = gsc_copy['primary_keyword']
        else:
            logger.warning("GSC data missing query column - cannot build topic-based opportunities")
            return pd.DataFrame(columns=empty_columns)
    elif query_col != 'query':
        gsc_copy['query'] = gsc_copy[query_col]

    # Ensure URL column exists for cannibalization detection
    url_col = None
    for col in ['url', 'URL', 'page', 'Page', 'landing_page']:
        if col in gsc_copy.columns:
            url_col = col
            break

    if url_col and url_col != 'url':
        gsc_copy['url'] = gsc_copy[url_col]
    elif url_col is None:
        gsc_copy['url'] = ''  # No URL data available

    # Ensure numeric columns
    for col in ['impressions', 'clicks', 'avg_position']:
        if col not in gsc_copy.columns:
            gsc_copy[col] = 0
        gsc_copy[col] = pd.to_numeric(gsc_copy[col], errors='coerce').fillna(0)

    # Filter to queries with meaningful impressions
    qualifying_queries = gsc_copy[gsc_copy['impressions'] >= 10].copy()

    if len(qualifying_queries) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Cluster queries into topics with Content Action Recommendations
    result = cluster_queries_into_topics(qualifying_queries, max_topics=20, thin_urls=thin_urls, url_word_counts=url_word_counts)

    if len(result) == 0:
        return pd.DataFrame(columns=empty_columns)

    # =========================================================================
    # BUSINESS RELEVANCE FILTERING (Gate 3)
    # =========================================================================
    # Build site context from crawl and GSC data
    site_context = build_site_context_from_df(df, gsc_df)

    # Apply business relevance filtering
    filtered_result, excluded_topics = filter_topics_by_business_relevance(
        topics_df=result,
        site_context=site_context,
        min_score=50,  # Minimum score threshold
        max_topics=15  # Cap to 15 topics max
    )

    # Log excluded topics for debugging
    if excluded_topics:
        logger.debug(f"[NEW CONTENT] Excluded {len(excluded_topics)} topics by business relevance filter")
        for ex in excluded_topics[:5]:  # Log first 5
            logger.debug(f"  - Excluded: {ex.get('Primary Keyword', 'N/A')} | Reason: {ex.get('Reason', 'N/A')}")

    # Ensure all expected columns exist
    for col in empty_columns:
        if col not in filtered_result.columns:
            filtered_result[col] = ''

    # Reorder columns to match expected output
    return filtered_result[empty_columns]


def build_redirect_merge_plan(new_content_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build the 'Redirect & Merge Plan' sheet from consolidation recommendations.

    Generates a 301 redirect mapping for every topic where the Recommended Action
    is "Consolidate pages". Each Secondary URL gets its own row mapping to the
    Primary URL.

    Args:
        new_content_df: DataFrame from build_new_content_opportunities_sheet
            containing Recommended Action, Primary URL, Secondary URLs columns

    Returns:
        DataFrame with columns: Topic, Recommended Action, Primary URL,
        Secondary URL, Redirect Type, Reason

    Primary URL Selection Validation:
    - Has highest total impressions OR clicks in the cluster
    - Has best (lowest) average position
    - Not flagged as thin content
    - Matches the dominant intent of the cluster
    """
    empty_columns = [
        'Topic', 'Recommended Action', 'Primary URL', 'Secondary URL',
        'Redirect Type', 'Reason'
    ]

    if new_content_df is None or len(new_content_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Filter for consolidation actions only
    consolidations = new_content_df[
        new_content_df['Recommended Action'] == 'Consolidate pages'
    ].copy()

    if len(consolidations) == 0:
        return pd.DataFrame(columns=empty_columns)

    redirect_rows = []

    for _, row in consolidations.iterrows():
        topic = row.get('Suggested Topic', '')
        primary_url = row.get('Primary URL', '')
        secondary_urls_str = row.get('Secondary URLs', '')
        original_reasoning = row.get('Reasoning', '')

        # Skip if no primary URL or secondary URLs
        if not primary_url or not secondary_urls_str:
            continue

        # Parse secondary URLs (comma-separated)
        secondary_urls = [
            url.strip() for url in secondary_urls_str.split(',')
            if url.strip() and url.strip() != primary_url
        ]

        # Skip if no valid secondary URLs after filtering
        if not secondary_urls:
            continue

        # Generate redirect entries for each secondary URL
        for secondary_url in secondary_urls:
            # Generate redirect-specific reasoning
            reason = (
                f"This page competes for the same keywords as the primary page "
                f"({primary_url[:50]}...) but has weaker rankings and engagement. "
                f"Redirecting consolidates authority and prevents keyword cannibalization."
            )

            redirect_rows.append({
                'Topic': topic,
                'Recommended Action': 'Consolidate pages',
                'Primary URL': primary_url,
                'Secondary URL': secondary_url,
                'Redirect Type': '301',
                'Reason': reason
            })

    if not redirect_rows:
        return pd.DataFrame(columns=empty_columns)

    result_df = pd.DataFrame(redirect_rows)

    # Remove any duplicate redirects (same Secondary URL)
    result_df = result_df.drop_duplicates(subset=['Secondary URL'], keep='first')

    # Ensure no self-redirects (Primary URL == Secondary URL)
    result_df = result_df[result_df['Primary URL'] != result_df['Secondary URL']]

    logger.info(f"Generated {len(result_df)} redirect entries from {len(consolidations)} consolidation topics")

    return result_df


def build_merge_playbooks(new_content_df: pd.DataFrame, gsc_df: Optional[pd.DataFrame] = None, url_word_counts: Optional[dict] = None) -> pd.DataFrame:
    """
    Build the 'Merge Playbooks' sheet with content-level merge instructions.

    For each "Consolidate pages" action, generates detailed content recommendations
    on what to keep from the Primary URL and what to move from Secondary URLs.

    Args:
        new_content_df: DataFrame from build_new_content_opportunities_sheet
            containing Recommended Action, Primary URL, Secondary URLs, etc.
        gsc_df: Raw GSC query-level data for analyzing per-URL query coverage
        url_word_counts: Dict mapping URL -> word count (optional)

    Returns:
        DataFrame with columns: Topic, Primary URL, Secondary URL, Keep This Content,
        Move These Sections, Retire This Page, Reasoning
    """
    empty_columns = [
        'Topic', 'Primary URL', 'Secondary URL', 'Keep This Content',
        'Move These Sections', 'Retire This Page', 'Reasoning'
    ]

    if new_content_df is None or len(new_content_df) == 0:
        return pd.DataFrame(columns=empty_columns)

    # Filter for consolidation actions only
    consolidations = new_content_df[
        new_content_df['Recommended Action'] == 'Consolidate pages'
    ].copy()

    if len(consolidations) == 0:
        return pd.DataFrame(columns=empty_columns)

    if url_word_counts is None:
        url_word_counts = {}

    # Build URL -> queries mapping from GSC data if available
    url_queries = {}  # url -> list of {query, impressions, clicks, position}
    if gsc_df is not None and len(gsc_df) > 0:
        gsc_copy = gsc_df.copy()

        # Find query column
        query_col = None
        for col in ['query', 'primary_keyword', 'Query', 'Keyword', 'Top queries']:
            if col in gsc_copy.columns:
                query_col = col
                break

        # Find URL column
        url_col = None
        for col in ['url', 'URL', 'page', 'Page', 'landing_page']:
            if col in gsc_copy.columns:
                url_col = col
                break

        if query_col and url_col:
            for _, row in gsc_copy.iterrows():
                url = str(row.get(url_col, '')).strip()
                query = str(row.get(query_col, '')).strip()
                if url and query and query != 'nan':
                    if url not in url_queries:
                        url_queries[url] = []
                    url_queries[url].append({
                        'query': query,
                        'impressions': int(row.get('impressions', 0)),
                        'clicks': int(row.get('clicks', 0)),
                        'position': float(row.get('avg_position', row.get('position', 0)))
                    })

    playbook_rows = []

    for _, row in consolidations.iterrows():
        topic = row.get('Suggested Topic', '')
        primary_keyword = row.get('Primary Keyword', '')
        primary_url = row.get('Primary URL', '')
        secondary_urls_str = row.get('Secondary URLs', '')
        original_reasoning = row.get('Reasoning', '')

        # Skip if no primary URL or secondary URLs
        if not primary_url or not secondary_urls_str:
            continue

        # Parse secondary URLs (comma-separated)
        secondary_urls = [
            url.strip() for url in secondary_urls_str.split(',')
            if url.strip() and url.strip() != primary_url
        ]

        if not secondary_urls:
            continue

        # Get Primary URL's queries
        primary_queries = url_queries.get(primary_url, [])
        primary_query_set = {q['query'].lower() for q in primary_queries}
        primary_word_count = url_word_counts.get(primary_url, 0)

        # Calculate total impressions for Primary URL
        primary_total_impressions = sum(q['impressions'] for q in primary_queries)

        # Build "Keep This Content" for Primary URL
        keep_content_bullets = []

        if primary_queries:
            # Group by position ranges
            top_position_queries = [q for q in primary_queries if q['position'] > 0 and q['position'] <= 5]
            strong_position_queries = [q for q in primary_queries if q['position'] > 5 and q['position'] <= 10]

            # Check for queries ranking 1-5
            if top_position_queries:
                top_query_examples = sorted(top_position_queries, key=lambda x: x['impressions'], reverse=True)[:3]
                query_names = ', '.join([f"'{q['query']}'" for q in top_query_examples])
                keep_content_bullets.append(f"Top-ranking content (positions 1-5): {query_names}")

            # Check for high-impression query clusters
            high_impression_queries = [q for q in primary_queries if q['impressions'] >= 50]
            if high_impression_queries:
                total_high_imp = sum(q['impressions'] for q in high_impression_queries)
                if primary_total_impressions > 0:
                    share = total_high_imp / primary_total_impressions
                    if share >= 0.20:
                        keep_content_bullets.append(f"High-traffic sections ({share:.0%} of impressions)")

            # Check for strong position queries (6-10)
            if strong_position_queries:
                keep_content_bullets.append(f"Content ranking positions 6-10 ({len(strong_position_queries)} queries)")

        # Add word count context
        if primary_word_count >= 1000:
            keep_content_bullets.append(f"Comprehensive content ({primary_word_count:,} words)")
        elif primary_word_count > 0:
            keep_content_bullets.append(f"Existing content ({primary_word_count:,} words)")

        if not keep_content_bullets:
            keep_content_bullets.append("Primary page structure and core content")

        keep_content = '\n'.join([f"• {b}" for b in keep_content_bullets])

        # Generate one row per secondary URL
        for secondary_url in secondary_urls:
            # Get Secondary URL's queries
            secondary_queries = url_queries.get(secondary_url, [])
            secondary_word_count = url_word_counts.get(secondary_url, 0)

            # Build "Move These Sections" for Secondary URL
            move_sections_bullets = []

            if secondary_queries:
                # Find unique queries not covered by Primary
                unique_queries = [
                    q for q in secondary_queries
                    if q['query'].lower() not in primary_query_set
                ]

                # Find queries where Secondary ranks better (>3 positions better)
                better_ranking_queries = []
                for sq in secondary_queries:
                    sq_query_lower = sq['query'].lower()
                    for pq in primary_queries:
                        if pq['query'].lower() == sq_query_lower:
                            if pq['position'] > 0 and sq['position'] > 0:
                                if pq['position'] - sq['position'] > 3:
                                    better_ranking_queries.append(sq)
                            break

                # Add unique query coverage
                if unique_queries:
                    unique_examples = sorted(unique_queries, key=lambda x: x['impressions'], reverse=True)[:3]
                    unique_names = ', '.join([f"'{q['query']}'" for q in unique_examples])
                    move_sections_bullets.append(f"Unique content not on Primary: {unique_names}")

                # Add better-ranking content
                if better_ranking_queries:
                    better_examples = sorted(better_ranking_queries, key=lambda x: x['impressions'], reverse=True)[:2]
                    better_names = ', '.join([f"'{q['query']}'" for q in better_examples])
                    move_sections_bullets.append(f"Content ranking better than Primary: {better_names}")

                # Check for high-value queries on Secondary
                high_value_secondary = [q for q in secondary_queries if q['impressions'] >= 30]
                if high_value_secondary and not move_sections_bullets:
                    hv_examples = sorted(high_value_secondary, key=lambda x: x['impressions'], reverse=True)[:2]
                    hv_names = ', '.join([f"'{q['query']}'" for q in hv_examples])
                    move_sections_bullets.append(f"High-impression queries: {hv_names}")

            # Check if Secondary has more content depth
            if secondary_word_count > primary_word_count and secondary_word_count > 500:
                diff = secondary_word_count - primary_word_count
                move_sections_bullets.append(f"Additional content depth ({diff:,} more words) - review for valuable sections")

            if not move_sections_bullets:
                move_sections_bullets.append("Review page for any unique angles or examples to preserve")

            move_sections = '\n'.join([f"• {b}" for b in move_sections_bullets])

            # Generate reasoning for this specific merge
            reasoning = (
                f"Merging '{secondary_url}' into '{primary_url}' eliminates keyword cannibalization "
                f"for the '{topic}' topic. Primary URL has stronger overall authority; "
                f"Secondary URL's unique content should be integrated before redirect."
            )

            playbook_rows.append({
                'Topic': topic,
                'Primary URL': primary_url,
                'Secondary URL': secondary_url,
                'Keep This Content': keep_content,
                'Move These Sections': move_sections,
                'Retire This Page': 'Yes',
                'Reasoning': reasoning
            })

    if not playbook_rows:
        return pd.DataFrame(columns=empty_columns)

    result_df = pd.DataFrame(playbook_rows)

    logger.info(f"Generated {len(result_df)} merge playbook entries from {len(consolidations)} consolidation topics")

    return result_df


def write_analytical_sheets(workbook, df: pd.DataFrame, thin_threshold: int = 1000, gsc_df: Optional[pd.DataFrame] = None) -> None:
    """
    Write all analytical insight sheets to the workbook.

    Creates five sheets:
    - Content to Optimize: Top 20 pages with value but underperforming
    - Thin Content Opportunities: Pages lacking depth
    - New Content Opportunities: Topic-based content gaps (not URLs)
    - Redirect & Merge Plan: 301 redirect mapping for consolidation actions
    - Merge Playbooks: Content-level merge instructions for consolidation

    Args:
        workbook: openpyxl Workbook object
        df: Processed URL-level DataFrame
        thin_threshold: Word count threshold for thin content
        gsc_df: Raw GSC query-level data for topic-based analysis
    """
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Define header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Sheet 1: Content to Optimize
    content_df = build_content_to_optimize_sheet(df)
    ws1 = workbook.create_sheet(title='Content to Optimize')

    # Write header
    if len(content_df) > 0:
        for col_idx, col_name in enumerate(content_df.columns, start=1):
            cell = ws1.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, row in enumerate(content_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws1.cell(row=row_idx, column=col_idx, value=value)

        # Set column widths
        ws1.column_dimensions['A'].width = 60  # URL
        ws1.column_dimensions['B'].width = 15  # Page Type
        ws1.column_dimensions['C'].width = 12  # Avg Position
        ws1.column_dimensions['D'].width = 12  # Impressions
        ws1.column_dimensions['E'].width = 10  # Clicks
        ws1.column_dimensions['F'].width = 10  # CTR
        ws1.column_dimensions['G'].width = 12  # Word Count
        ws1.column_dimensions['H'].width = 18  # Has Meta Desc
        ws1.column_dimensions['I'].width = 10  # Sessions
        ws1.column_dimensions['J'].width = 50  # Optimization Signals
        ws1.column_dimensions['K'].width = 14  # Priority Score
    else:
        ws1.cell(row=1, column=1, value='No content optimization opportunities found')

    ws1.freeze_panes = 'A2'
    logger.info(f"Wrote Content to Optimize sheet: {len(content_df)} rows")

    # Sheet 2: Thin Content Opportunities
    thin_df = build_thin_content_sheet(df, thin_threshold)
    ws2 = workbook.create_sheet(title='Thin Content Opportunities')

    if len(thin_df) > 0:
        for col_idx, col_name in enumerate(thin_df.columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(thin_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws2.cell(row=row_idx, column=col_idx, value=value)

        ws2.column_dimensions['A'].width = 60  # URL
        ws2.column_dimensions['B'].width = 15  # Page Type
        ws2.column_dimensions['C'].width = 12  # Word Count
        ws2.column_dimensions['D'].width = 12  # Content Gap
        ws2.column_dimensions['E'].width = 10  # Sessions
        ws2.column_dimensions['F'].width = 12  # Impressions
        ws2.column_dimensions['G'].width = 12  # Avg Position
        ws2.column_dimensions['H'].width = 16  # Referring Domains
        ws2.column_dimensions['I'].width = 50  # Opportunity Type
        ws2.column_dimensions['J'].width = 12  # Value Score
    else:
        ws2.cell(row=1, column=1, value='No thin content opportunities found')

    ws2.freeze_panes = 'A2'
    logger.info(f"Wrote Thin Content Opportunities sheet: {len(thin_df)} rows")

    # Sheet 3: New Content Opportunities (Topic-based, not URL-based)
    # Pass thin_threshold for cannibalization detection
    new_content_df = build_new_content_opportunities_sheet(df, gsc_df, thin_threshold)
    ws3 = workbook.create_sheet(title='New Content Opportunities')

    if len(new_content_df) > 0:
        for col_idx, col_name in enumerate(new_content_df.columns, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(new_content_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws3.cell(row=row_idx, column=col_idx, value=value)

        # Column widths for topic-based structure with Content Action Recommendations
        ws3.column_dimensions['A'].width = 35  # Suggested Topic
        ws3.column_dimensions['B'].width = 35  # Primary Keyword
        ws3.column_dimensions['C'].width = 60  # Secondary Keywords
        ws3.column_dimensions['D'].width = 16  # Total Impressions
        ws3.column_dimensions['E'].width = 12  # Avg Position
        ws3.column_dimensions['F'].width = 22  # Recommended Action
        ws3.column_dimensions['G'].width = 60  # Primary URL
        ws3.column_dimensions['H'].width = 80  # Secondary URLs
        ws3.column_dimensions['I'].width = 100  # Reasoning
        ws3.column_dimensions['J'].width = 22  # Suggested Page Type
        ws3.column_dimensions['K'].width = 14  # Priority Score
    else:
        ws3.cell(row=1, column=1, value='No new content opportunities found (GSC query data required)')

    ws3.freeze_panes = 'A2'
    logger.info(f"Wrote New Content Opportunities sheet: {len(new_content_df)} topics")

    # Sheet 4: Redirect & Merge Plan (for consolidation actions)
    redirect_df = build_redirect_merge_plan(new_content_df)
    ws4 = workbook.create_sheet(title='Redirect & Merge Plan')

    if len(redirect_df) > 0:
        for col_idx, col_name in enumerate(redirect_df.columns, start=1):
            cell = ws4.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(redirect_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws4.cell(row=row_idx, column=col_idx, value=value)

        # Column widths for redirect plan
        ws4.column_dimensions['A'].width = 35  # Topic
        ws4.column_dimensions['B'].width = 20  # Recommended Action
        ws4.column_dimensions['C'].width = 70  # Primary URL (destination)
        ws4.column_dimensions['D'].width = 70  # Secondary URL (source)
        ws4.column_dimensions['E'].width = 14  # Redirect Type
        ws4.column_dimensions['F'].width = 100  # Reason
    else:
        ws4.cell(row=1, column=1, value='No consolidation redirects needed (no "Consolidate pages" actions found)')

    ws4.freeze_panes = 'A2'
    logger.info(f"Wrote Redirect & Merge Plan sheet: {len(redirect_df)} redirects")

    # Sheet 5: Merge Playbooks (content-level merge instructions)
    # Build url_word_counts for merge playbooks
    url_word_counts = {}
    if df is not None and len(df) > 0:
        word_count_col = None
        for col in ['word_count', 'Word Count', 'wordcount']:
            if col in df.columns:
                word_count_col = col
                break

        url_col = None
        for col in ['url', 'URL', 'Address']:
            if col in df.columns:
                url_col = col
                break

        if word_count_col and url_col:
            df_copy = df.copy()
            df_copy[word_count_col] = pd.to_numeric(df_copy[word_count_col], errors='coerce').fillna(0)
            url_word_counts = dict(zip(df_copy[url_col], df_copy[word_count_col]))

    playbook_df = build_merge_playbooks(new_content_df, gsc_df, url_word_counts)
    ws5 = workbook.create_sheet(title='Merge Playbooks')

    if len(playbook_df) > 0:
        for col_idx, col_name in enumerate(playbook_df.columns, start=1):
            cell = ws5.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(playbook_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws5.cell(row=row_idx, column=col_idx, value=value)

        # Column widths for merge playbooks
        ws5.column_dimensions['A'].width = 35  # Topic
        ws5.column_dimensions['B'].width = 70  # Primary URL
        ws5.column_dimensions['C'].width = 70  # Secondary URL
        ws5.column_dimensions['D'].width = 60  # Keep This Content
        ws5.column_dimensions['E'].width = 60  # Move These Sections
        ws5.column_dimensions['F'].width = 18  # Retire This Page
        ws5.column_dimensions['G'].width = 100  # Reasoning
    else:
        ws5.cell(row=1, column=1, value='No merge playbooks needed (no "Consolidate pages" actions found)')

    ws5.freeze_panes = 'A2'
    logger.info(f"Wrote Merge Playbooks sheet: {len(playbook_df)} entries")

    # GUARD: Log warning if GSC data exists but all related sheets are empty
    # This indicates a potential logic issue that should be investigated
    gsc_has_data = gsc_df is not None and len(gsc_df) > 0
    new_content_empty = len(new_content_df) == 0
    redirect_empty = len(redirect_df) == 0
    playbook_empty = len(playbook_df) == 0

    if gsc_has_data and new_content_empty:
        logger.warning("[GUARD] GSC data exists but New Content Opportunities is empty - check clustering logic")
        logger.warning(f"[GUARD] GSC data has {len(gsc_df)} rows")

    # Count action types in new_content_df
    consolidate_count = 0
    if len(new_content_df) > 0 and 'Recommended Action' in new_content_df.columns:
        consolidate_count = len(new_content_df[new_content_df['Recommended Action'] == 'Consolidate pages'])
        create_count = len(new_content_df[new_content_df['Recommended Action'] == 'Create new page'])
        expand_count = len(new_content_df[new_content_df['Recommended Action'] == 'Expand existing page'])
        logger.info(f"[GUARD] Action summary: Create={create_count}, Expand={expand_count}, Consolidate={consolidate_count}")

    # If consolidate actions exist but downstream sheets are empty, log an error
    if consolidate_count > 0 and redirect_empty:
        logger.error(f"[GUARD] {consolidate_count} consolidation actions exist but Redirect & Merge Plan is empty - check downstream logic")

    if consolidate_count > 0 and playbook_empty:
        logger.error(f"[GUARD] {consolidate_count} consolidation actions exist but Merge Playbooks is empty - check downstream logic")


def write_excel(
    df: pd.DataFrame,
    summaries: Dict[str, pd.DataFrame],
    output_path: str,
    quality_report: Optional[DataQualityReport] = None,
    thin_content_threshold: int = 1000,
    gsc_df: Optional[pd.DataFrame] = None
) -> None:
    """
    Write the processed data and summaries to an Excel file.

    Creates nine sheets:
    - Aggregation: Full dataset with all columns
    - Actions: Filtered view with key columns, sorted by priority
    - Summary: Statistical summaries
    - Data Quality: Data quality issues and diagnostics (if provided)
    - Content to Optimize: Top 20 pages with value but underperforming
    - Thin Content Opportunities: Pages lacking depth
    - New Content Opportunities: Topic-based content gaps (requires GSC query data)
    - Redirect & Merge Plan: 301 redirect mapping for consolidation actions
    - Merge Playbooks: Content-level merge instructions for consolidation

    Args:
        df: Processed URL-level DataFrame
        summaries: Dictionary of summary DataFrames
        output_path: Path to output Excel file
        quality_report: Optional DataQualityReport with validation results
        thin_content_threshold: Word count threshold for thin content analysis
        gsc_df: Raw GSC query-level data for topic-based New Content analysis
    """
    logger.info(f"Writing Excel report to: {output_path}")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Aggregation (full data)
        df.to_excel(writer, sheet_name='Aggregation', index=False)
        logger.info(f"Wrote Aggregation sheet: {len(df)} rows")

        # Sheet 2: Actions (filtered and sorted)
        action_columns = [
            'url', 'page_type', 'status_code', 'sessions', 'avg_position',
            'referring_domains', 'backlinks', 'word_count', 'inlinks',
            'technical_actions', 'content_actions', 'priority'
        ]

        # Only include columns that exist
        action_columns = [col for col in action_columns if col in df.columns]

        df_actions = df[action_columns].copy()

        # Sort by priority (High > Medium > Low), then by sessions descending
        priority_map = {'High': 0, 'Medium': 1, 'Low': 2}
        df_actions['priority_sort'] = df_actions['priority'].map(priority_map)
        df_actions = df_actions.sort_values(
            ['priority_sort', 'sessions'],
            ascending=[True, False]
        ).drop('priority_sort', axis=1)

        df_actions.to_excel(writer, sheet_name='Actions', index=False)
        logger.info(f"Wrote Actions sheet: {len(df_actions)} rows")

        # Sheet 3: Summary
        current_row = 0

        # Write each summary table with headers
        summary_order = [
            ('Priority Distribution', 'priority'),
            ('Technical Actions', 'technical_actions'),
            ('Content Actions', 'content_actions'),
            ('Page Types', 'page_type'),
            ('Status Codes', 'status_code'),
        ]

        for title, key in summary_order:
            if key in summaries:
                # Write title
                title_df = pd.DataFrame([[title]], columns=[''])
                title_df.to_excel(
                    writer,
                    sheet_name='Summary',
                    index=False,
                    header=False,
                    startrow=current_row
                )
                current_row += 1

                # Write data
                summaries[key].to_excel(
                    writer,
                    sheet_name='Summary',
                    index=False,
                    startrow=current_row
                )
                current_row += len(summaries[key]) + 3  # Add spacing

        logger.info("Wrote Summary sheet")

        # Sheet 4: Data Quality (if provided)
        if quality_report is not None:
            quality_sheets = generate_data_quality_sheets(quality_report)
            current_row = 0

            # Write header
            header_df = pd.DataFrame([['DATA QUALITY REPORT']], columns=[''])
            header_df.to_excel(
                writer,
                sheet_name='Data Quality',
                index=False,
                header=False,
                startrow=current_row
            )
            current_row += 2

            # Write Data Source Coverage
            title_df = pd.DataFrame([['Data Source Coverage']], columns=[''])
            title_df.to_excel(
                writer,
                sheet_name='Data Quality',
                index=False,
                header=False,
                startrow=current_row
            )
            current_row += 1

            if not quality_sheets['coverage'].empty:
                quality_sheets['coverage'].to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    startrow=current_row
                )
                current_row += len(quality_sheets['coverage']) + 3

            # Write Column Quality
            title_df = pd.DataFrame([['Column Data Quality']], columns=[''])
            title_df.to_excel(
                writer,
                sheet_name='Data Quality',
                index=False,
                header=False,
                startrow=current_row
            )
            current_row += 1

            if not quality_sheets['column_quality'].empty:
                quality_sheets['column_quality'].to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    startrow=current_row
                )
                current_row += len(quality_sheets['column_quality']) + 3

            # Write Issues
            title_df = pd.DataFrame([['Data Quality Issues']], columns=[''])
            title_df.to_excel(
                writer,
                sheet_name='Data Quality',
                index=False,
                header=False,
                startrow=current_row
            )
            current_row += 1

            if not quality_sheets['issues'].empty:
                quality_sheets['issues'].to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    startrow=current_row
                )
                current_row += len(quality_sheets['issues']) + 3
            else:
                no_issues_df = pd.DataFrame([['No data quality issues found!']], columns=[''])
                no_issues_df.to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    header=False,
                    startrow=current_row
                )

            # Write warnings if any
            if quality_report.warnings:
                current_row += 2
                title_df = pd.DataFrame([['Warnings']], columns=[''])
                title_df.to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    header=False,
                    startrow=current_row
                )
                current_row += 1
                warnings_df = pd.DataFrame([[w] for w in quality_report.warnings], columns=['Warning'])
                warnings_df.to_excel(
                    writer,
                    sheet_name='Data Quality',
                    index=False,
                    startrow=current_row
                )

            logger.info(f"Wrote Data Quality sheet: {len(quality_report.issues)} issues found")

        # Sheets 5-7: Analytical Insight Sheets
        write_analytical_sheets(writer.book, df, thin_content_threshold, gsc_df)

    logger.info(f"Excel report saved successfully: {output_path}")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main() -> int:
    """
    Main entry point for the WQA generator.

    Returns:
        Exit code (0 for success, 1 for error)
    """
    try:
        # Parse arguments
        args = parse_args()

        logger.info("=" * 60)
        logger.info("Website Quality Audit (WQA) Generator")
        logger.info("=" * 60)
        logger.info(f"Crawl file: {args.crawl}")
        logger.info(f"GA4 file: {args.ga or 'Not provided'}")
        logger.info(f"GSC file: {args.gsc or 'Not provided'}")
        logger.info(f"Backlinks file: {args.backlinks or 'Not provided'}")
        logger.info(f"Output file: {args.output}")
        logger.info("-" * 60)
        logger.info("Thresholds:")
        logger.info(f"  Low traffic: {args.low_traffic_threshold} sessions")
        logger.info(f"  Thin content: {args.thin_content_threshold} words")
        logger.info(f"  High rank max position: {args.high_rank_max_position}")
        logger.info(f"  Low CTR: {args.low_ctr_threshold * 100}%")
        logger.info("-" * 60)

        # Load data
        crawl_df = load_crawl_data(args.crawl)
        ga_df = load_ga_data(args.ga)
        gsc_df = load_gsc_data(args.gsc)
        backlink_df = load_backlink_data(args.backlinks)

        # Merge datasets
        df = merge_datasets(crawl_df, ga_df, gsc_df, backlink_df)

        # Classify page types
        logger.info("Classifying page types...")
        df['page_type'] = df['url'].apply(classify_page_type)

        page_type_dist = df['page_type'].value_counts()
        for ptype, count in page_type_dist.items():
            logger.info(f"  {ptype}: {count} URLs")

        # Assign actions
        logger.info("Applying rules engine to assign actions...")
        actions_result = df.apply(
            lambda row: assign_actions(
                row,
                low_traffic_threshold=args.low_traffic_threshold,
                thin_content_threshold=args.thin_content_threshold,
                high_rank_max_position=args.high_rank_max_position,
                low_ctr_threshold=args.low_ctr_threshold
            ),
            axis=1
        )

        # Unpack results
        df['technical_actions'] = actions_result.apply(lambda x: ', '.join(x[0]) if x[0] else '')
        df['content_actions'] = actions_result.apply(lambda x: ', '.join(x[1]) if x[1] else '')

        # Assign priority
        logger.info("Assigning priorities...")
        df['priority'] = df.apply(assign_priority, axis=1)

        priority_dist = df['priority'].value_counts()
        for priority, count in priority_dist.items():
            logger.info(f"  {priority}: {count} URLs")

        # Generate summary
        summaries = generate_summary(df)

        # Validate data quality
        quality_report = validate_data_quality(df, crawl_df, ga_df, gsc_df, backlink_df)

        # Write output
        write_excel(df, summaries, args.output, quality_report, args.thin_content_threshold, gsc_df)

        # Log summary of data quality issues
        high_issues = sum(1 for i in quality_report.issues if i['severity'] == 'High')
        if high_issues > 0:
            logger.warning(f"ATTENTION: {high_issues} high-severity data quality issues found!")
            logger.warning("Review the 'Data Quality' sheet in the Excel report for details.")

        logger.info("=" * 60)
        logger.info("WQA GENERATION COMPLETE")
        logger.info(f"Total URLs processed: {len(df)}")
        logger.info(f"Data Quality Issues: {len(quality_report.issues)} ({high_issues} High)")
        logger.info(f"Report saved to: {args.output}")
        logger.info("=" * 60)

        return 0

    except FileNotFoundError as e:
        logger.error(f"File not found: {e}")
        return 1
    except Exception as e:
        logger.error(f"Error generating WQA report: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())
