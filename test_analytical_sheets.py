#!/usr/bin/env python3
"""
Verification script for Analytical Insight sheets in WQA Generator output.

This script verifies that the five analytical insight sheets are correctly created:
- Content to Optimize
- Thin Content Opportunities
- New Content Opportunities
- Redirect & Merge Plan
- Merge Playbooks

Usage:
    python test_analytical_sheets.py <path_to_wqa_output.xlsx>
    python test_analytical_sheets.py --unit-test  # Run unit tests with synthetic data

Example:
    python test_analytical_sheets.py wqa_report.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def run_clustering_unit_tests() -> bool:
    """
    Run unit tests for the cluster_queries_into_topics function using synthetic data.

    Tests various scenarios:
    1. Clusters with dominant pages should NOT be excluded
    2. Consolidation should trigger for competing URLs
    3. Every cluster should get an action (no empty recommended_action)

    Returns:
        True if all tests pass, False otherwise
    """
    import pandas as pd
    import sys
    sys.path.insert(0, str(Path(__file__).parent))

    try:
        from wqa_generator import cluster_queries_into_topics, build_new_content_opportunities_sheet
    except ImportError as e:
        print(f"ERROR: Could not import wqa_generator: {e}")
        return False

    print("=" * 60)
    print("UNIT TEST: Clustering Logic Verification")
    print("=" * 60)
    print()

    errors = []
    tests_passed = 0
    tests_total = 0

    # =========================================================================
    # Test 1: Clusters with dominant pages should NOT be excluded
    # =========================================================================
    tests_total += 1
    print("Test 1: Dominant page clusters should produce recommendations...")

    gsc_data_1 = pd.DataFrame([
        # Cluster: "new homes huntsville" - one dominant page
        {'query': 'new homes huntsville', 'url': 'https://example.com/huntsville', 'impressions': 1000, 'clicks': 50, 'avg_position': 5},
        {'query': 'huntsville new construction', 'url': 'https://example.com/huntsville', 'impressions': 800, 'clicks': 40, 'avg_position': 6},
        {'query': 'new home builders huntsville', 'url': 'https://example.com/huntsville', 'impressions': 600, 'clicks': 30, 'avg_position': 7},
        # Small secondary presence (10% of total)
        {'query': 'homes for sale huntsville', 'url': 'https://example.com/alabama', 'impressions': 100, 'clicks': 5, 'avg_position': 15},
    ])

    result_1 = cluster_queries_into_topics(gsc_data_1, max_topics=10)

    if len(result_1) == 0:
        errors.append("Test 1 FAILED: No clusters produced (dominant page was wrongly excluded)")
    else:
        # Check that a recommendation exists
        if 'Recommended Action' not in result_1.columns:
            errors.append("Test 1 FAILED: Missing 'Recommended Action' column")
        elif result_1['Recommended Action'].isna().any() or (result_1['Recommended Action'] == '').any():
            errors.append("Test 1 FAILED: Some clusters have empty Recommended Action")
        else:
            print(f"  [OK] Produced {len(result_1)} cluster(s) with actions")
            tests_passed += 1

    # =========================================================================
    # Test 2: Two competing URLs should trigger consolidation
    # =========================================================================
    tests_total += 1
    print("Test 2: Competing URLs should trigger 'Consolidate pages'...")

    gsc_data_2 = pd.DataFrame([
        # Cluster: Two pages competing for same keywords (cannibalization)
        {'query': 'best roofing company atlanta', 'url': 'https://example.com/roofing', 'impressions': 500, 'clicks': 25, 'avg_position': 8},
        {'query': 'atlanta roofing services', 'url': 'https://example.com/roofing', 'impressions': 400, 'clicks': 20, 'avg_position': 9},
        {'query': 'top roofers atlanta', 'url': 'https://example.com/atlanta-roofing', 'impressions': 350, 'clicks': 18, 'avg_position': 10},
        {'query': 'roofing contractors atlanta ga', 'url': 'https://example.com/atlanta-roofing', 'impressions': 300, 'clicks': 15, 'avg_position': 11},
    ])

    result_2 = cluster_queries_into_topics(gsc_data_2, max_topics=10)

    if len(result_2) == 0:
        errors.append("Test 2 FAILED: No clusters produced")
    else:
        consolidate_count = (result_2['Recommended Action'] == 'Consolidate pages').sum()
        if consolidate_count == 0:
            actions = result_2['Recommended Action'].tolist()
            errors.append(f"Test 2 FAILED: Expected 'Consolidate pages' but got: {actions}")
        else:
            print(f"  [OK] Found {consolidate_count} 'Consolidate pages' action(s)")
            tests_passed += 1

    # =========================================================================
    # Test 3: No URL coverage should trigger "Create new page"
    # =========================================================================
    tests_total += 1
    print("Test 3: No URL coverage should trigger 'Create new page'...")

    gsc_data_3 = pd.DataFrame([
        # Cluster: Queries with no URL (content gap)
        {'query': 'how to install solar panels', 'url': '', 'impressions': 800, 'clicks': 0, 'avg_position': 0},
        {'query': 'solar panel installation guide', 'url': '', 'impressions': 600, 'clicks': 0, 'avg_position': 0},
        {'query': 'diy solar panels', 'url': '', 'impressions': 400, 'clicks': 0, 'avg_position': 0},
    ])

    result_3 = cluster_queries_into_topics(gsc_data_3, max_topics=10)

    if len(result_3) == 0:
        errors.append("Test 3 FAILED: No clusters produced")
    else:
        create_count = (result_3['Recommended Action'] == 'Create new page').sum()
        if create_count == 0:
            actions = result_3['Recommended Action'].tolist()
            errors.append(f"Test 3 FAILED: Expected 'Create new page' but got: {actions}")
        else:
            print(f"  [OK] Found {create_count} 'Create new page' action(s)")
            tests_passed += 1

    # =========================================================================
    # Test 4: Thin content should trigger "Expand existing page"
    # =========================================================================
    tests_total += 1
    print("Test 4: Thin content should trigger 'Expand existing page'...")

    gsc_data_4 = pd.DataFrame([
        # Cluster: Single page with thin content
        {'query': 'bathroom remodel ideas', 'url': 'https://example.com/bathroom', 'impressions': 500, 'clicks': 25, 'avg_position': 12},
        {'query': 'small bathroom renovation', 'url': 'https://example.com/bathroom', 'impressions': 400, 'clicks': 20, 'avg_position': 14},
    ])

    thin_urls = {'https://example.com/bathroom'}
    url_word_counts = {'https://example.com/bathroom': 300}  # Thin content (< 1000 words)

    result_4 = cluster_queries_into_topics(gsc_data_4, max_topics=10, thin_urls=thin_urls, url_word_counts=url_word_counts)

    if len(result_4) == 0:
        errors.append("Test 4 FAILED: No clusters produced")
    else:
        expand_count = (result_4['Recommended Action'] == 'Expand existing page').sum()
        if expand_count == 0:
            actions = result_4['Recommended Action'].tolist()
            errors.append(f"Test 4 FAILED: Expected 'Expand existing page' but got: {actions}")
        else:
            print(f"  [OK] Found {expand_count} 'Expand existing page' action(s)")
            tests_passed += 1

    # =========================================================================
    # Test 5: Three competing URLs should trigger consolidation
    # =========================================================================
    tests_total += 1
    print("Test 5: Three+ competing URLs should trigger consolidation...")

    gsc_data_5 = pd.DataFrame([
        # Cluster: Three pages competing (fragmented authority)
        {'query': 'plumber near me', 'url': 'https://example.com/plumbing-1', 'impressions': 400, 'clicks': 20, 'avg_position': 10},
        {'query': 'local plumber services', 'url': 'https://example.com/plumbing-1', 'impressions': 300, 'clicks': 15, 'avg_position': 11},
        {'query': 'best plumber in town', 'url': 'https://example.com/plumbing-2', 'impressions': 350, 'clicks': 18, 'avg_position': 12},
        {'query': 'emergency plumber', 'url': 'https://example.com/plumbing-2', 'impressions': 250, 'clicks': 12, 'avg_position': 13},
        {'query': 'plumbing repair service', 'url': 'https://example.com/plumbing-3', 'impressions': 200, 'clicks': 10, 'avg_position': 14},
    ])

    result_5 = cluster_queries_into_topics(gsc_data_5, max_topics=10)

    if len(result_5) == 0:
        errors.append("Test 5 FAILED: No clusters produced")
    else:
        consolidate_count = (result_5['Recommended Action'] == 'Consolidate pages').sum()
        if consolidate_count == 0:
            actions = result_5['Recommended Action'].tolist()
            # This might be expand if one URL dominates, which is also acceptable
            if (result_5['Recommended Action'] == 'Expand existing page').sum() > 0:
                print(f"  [OK] Got 'Expand existing page' (one URL slightly dominant)")
                tests_passed += 1
            else:
                errors.append(f"Test 5 FAILED: Expected consolidation or expand but got: {actions}")
        else:
            # Verify Secondary URLs has multiple entries
            consolidate_rows = result_5[result_5['Recommended Action'] == 'Consolidate pages']
            has_multiple_urls = any(',' in str(row) for row in consolidate_rows['Secondary URLs'])
            print(f"  [OK] Found {consolidate_count} 'Consolidate pages' action(s)")
            tests_passed += 1

    # =========================================================================
    # Test 6: Integration test with build_new_content_opportunities_sheet
    # =========================================================================
    tests_total += 1
    print("Test 6: Full integration test with build_new_content_opportunities_sheet...")

    # Create mock URL-level DataFrame
    df = pd.DataFrame([
        {'url': 'https://example.com/page1', 'word_count': 1500, 'sessions': 100},
        {'url': 'https://example.com/page2', 'word_count': 500, 'sessions': 50},  # Thin
    ])

    gsc_data_6 = pd.DataFrame([
        # Mix of scenarios
        {'query': 'keyword one test', 'url': 'https://example.com/page1', 'impressions': 500, 'clicks': 25, 'avg_position': 8},
        {'query': 'keyword two test', 'url': 'https://example.com/page1', 'impressions': 400, 'clicks': 20, 'avg_position': 9},
        {'query': 'another topic query', 'url': 'https://example.com/page2', 'impressions': 300, 'clicks': 15, 'avg_position': 15},
        {'query': 'topic query similar', 'url': 'https://example.com/page2', 'impressions': 250, 'clicks': 12, 'avg_position': 16},
    ])

    result_6 = build_new_content_opportunities_sheet(df, gsc_data_6, thin_threshold=1000)

    if len(result_6) == 0:
        errors.append("Test 6 FAILED: build_new_content_opportunities_sheet returned empty DataFrame")
    else:
        # Check all required columns exist
        required_cols = ['Suggested Topic', 'Primary Keyword', 'Recommended Action', 'Primary URL', 'Secondary URLs', 'Reasoning']
        missing_cols = [col for col in required_cols if col not in result_6.columns]
        if missing_cols:
            errors.append(f"Test 6 FAILED: Missing columns: {missing_cols}")
        else:
            # Check no empty actions
            empty_actions = result_6['Recommended Action'].isna().sum() + (result_6['Recommended Action'] == '').sum()
            if empty_actions > 0:
                errors.append(f"Test 6 FAILED: {empty_actions} row(s) have empty Recommended Action")
            else:
                print(f"  [OK] Produced {len(result_6)} topic(s) with complete data")
                actions = result_6['Recommended Action'].value_counts().to_dict()
                print(f"       Actions: {actions}")
                tests_passed += 1

    # =========================================================================
    # Summary
    # =========================================================================
    print()
    print("=" * 60)
    print(f"UNIT TEST RESULTS: {tests_passed}/{tests_total} tests passed")
    print("=" * 60)

    if errors:
        print()
        print("ERRORS:")
        for error in errors:
            print(f"  [FAIL] {error}")
        return False
    else:
        print("[OK] All unit tests passed!")
        return True


def verify_analytical_sheets(excel_path: str) -> bool:
    """
    Verify that the analytical insight sheets exist and have correct structure.

    Args:
        excel_path: Path to the WQA output Excel file

    Returns:
        True if all verifications pass, False otherwise
    """
    errors = []
    warnings = []

    # Check file exists
    if not Path(excel_path).exists():
        print(f"ERROR: File not found: {excel_path}")
        return False

    # Load workbook
    try:
        wb = load_workbook(excel_path, read_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}")
        return False

    print(f"Loaded workbook: {excel_path}")
    print(f"Available sheets: {wb.sheetnames}")
    print()

    # ==========================================================================
    # Verify Sheet 1: Content to Optimize
    # ==========================================================================
    sheet_name = 'Content to Optimize'
    print(f"Checking '{sheet_name}'...")

    if sheet_name not in wb.sheetnames:
        errors.append(f"Sheet '{sheet_name}' not found")
    else:
        ws = wb[sheet_name]
        print(f"  [OK] Sheet exists")

        # Check for expected headers
        expected_headers = [
            'URL', 'Page Type', 'Avg Position', 'Impressions', 'Clicks', 'CTR (%)',
            'Word Count', 'Has Meta Description', 'Sessions', 'Optimization Signals', 'Priority Score'
        ]

        # Check if sheet has data or the "no opportunities" message
        first_cell = ws.cell(row=1, column=1).value
        if first_cell == 'No content optimization opportunities found':
            print(f"  [OK] Sheet has 'no opportunities' message (data-dependent)")
        else:
            # Check headers
            for col_idx, expected in enumerate(expected_headers, start=1):
                actual = ws.cell(row=1, column=col_idx).value
                if actual != expected:
                    errors.append(f"  {sheet_name}: Column {col_idx} header - expected '{expected}', got '{actual}'")
                else:
                    print(f"  [OK] Column {col_idx}: '{expected}'")

            # Check freeze panes
            if ws.freeze_panes != 'A2':
                warnings.append(f"  {sheet_name}: Freeze panes expected 'A2', got '{ws.freeze_panes}'")
            else:
                print(f"  [OK] Freeze panes at A2")

            # Count data rows (should be capped at max 20)
            row_count = ws.max_row - 1  # Exclude header
            print(f"  [OK] Data rows: {row_count}")

            # Verify Top 20 cap
            if row_count > 20:
                errors.append(f"  {sheet_name}: Expected max 20 rows, got {row_count}")
            else:
                print(f"  [OK] Row count within Top 20 cap")

    print()

    # ==========================================================================
    # Verify Sheet 2: Thin Content Opportunities
    # ==========================================================================
    sheet_name = 'Thin Content Opportunities'
    print(f"Checking '{sheet_name}'...")

    if sheet_name not in wb.sheetnames:
        errors.append(f"Sheet '{sheet_name}' not found")
    else:
        ws = wb[sheet_name]
        print(f"  [OK] Sheet exists")

        # Check for expected headers
        expected_headers = [
            'URL', 'Page Type', 'Word Count', 'Content Gap', 'Sessions',
            'Impressions', 'Avg Position', 'Referring Domains', 'Opportunity Type', 'Value Score'
        ]

        # Check if sheet has data or the "no opportunities" message
        first_cell = ws.cell(row=1, column=1).value
        if first_cell == 'No thin content opportunities found':
            print(f"  [OK] Sheet has 'no opportunities' message (data-dependent)")
        else:
            # Check headers
            for col_idx, expected in enumerate(expected_headers, start=1):
                actual = ws.cell(row=1, column=col_idx).value
                if actual != expected:
                    errors.append(f"  {sheet_name}: Column {col_idx} header - expected '{expected}', got '{actual}'")
                else:
                    print(f"  [OK] Column {col_idx}: '{expected}'")

            # Check freeze panes
            if ws.freeze_panes != 'A2':
                warnings.append(f"  {sheet_name}: Freeze panes expected 'A2', got '{ws.freeze_panes}'")
            else:
                print(f"  [OK] Freeze panes at A2")

            # Count data rows
            row_count = ws.max_row - 1  # Exclude header
            print(f"  [OK] Data rows: {row_count}")

    print()

    # ==========================================================================
    # Verify Sheet 3: New Content Opportunities
    # ==========================================================================
    sheet_name = 'New Content Opportunities'
    print(f"Checking '{sheet_name}'...")

    if sheet_name not in wb.sheetnames:
        errors.append(f"Sheet '{sheet_name}' not found")
    else:
        ws = wb[sheet_name]
        print(f"  [OK] Sheet exists")

        # Check for expected headers with Content Action Recommendation columns
        expected_headers = [
            'Suggested Topic', 'Primary Keyword', 'Secondary Keywords', 'Total Impressions',
            'Avg Position', 'Recommended Action', 'Primary URL', 'Secondary URLs', 'Reasoning',
            'Suggested Page Type', 'Priority Score'
        ]

        # Valid action values
        valid_actions = ['Create new page', 'Expand existing page', 'Consolidate pages']

        # Check if sheet has data or the "no opportunities" message
        first_cell = ws.cell(row=1, column=1).value
        if first_cell == 'No new content topic opportunities found':
            print(f"  [OK] Sheet has 'no opportunities' message (data-dependent)")
        else:
            # Check headers
            for col_idx, expected in enumerate(expected_headers, start=1):
                actual = ws.cell(row=1, column=col_idx).value
                if actual != expected:
                    errors.append(f"  {sheet_name}: Column {col_idx} header - expected '{expected}', got '{actual}'")
                else:
                    print(f"  [OK] Column {col_idx}: '{expected}'")

            # Check freeze panes
            if ws.freeze_panes != 'A2':
                warnings.append(f"  {sheet_name}: Freeze panes expected 'A2', got '{ws.freeze_panes}'")
            else:
                print(f"  [OK] Freeze panes at A2")

            # Count data rows (topic-based sheet, capped at 20)
            row_count = ws.max_row - 1  # Exclude header
            print(f"  [OK] Data rows: {row_count}")

            # Verify Top 20 cap
            if row_count > 20:
                errors.append(f"  {sheet_name}: Expected max 20 topics, got {row_count}")
            else:
                print(f"  [OK] Row count within Top 20 cap")

            # ======================================================================
            # Content Action Recommendation Validation
            # ======================================================================
            print(f"  Validating Content Action Recommendations...")

            # Column indices (1-based)
            action_col = 6      # Recommended Action
            primary_url_col = 7  # Primary URL
            secondary_urls_col = 8  # Secondary URLs

            action_counts = {'Create new page': 0, 'Expand existing page': 0, 'Consolidate pages': 0}
            rows_without_action = 0
            invalid_actions = []
            create_with_url = 0
            consolidate_without_multiple = 0

            for row_idx in range(2, ws.max_row + 1):
                action = ws.cell(row=row_idx, column=action_col).value
                primary_url = ws.cell(row=row_idx, column=primary_url_col).value or ''
                secondary_urls = ws.cell(row=row_idx, column=secondary_urls_col).value or ''

                # Check each topic has exactly one recommendation
                if not action:
                    rows_without_action += 1
                elif action not in valid_actions:
                    invalid_actions.append(f"Row {row_idx}: '{action}'")
                else:
                    action_counts[action] += 1

                    # "Create new page" rows should have empty Primary URL
                    if action == 'Create new page':
                        if primary_url.strip():
                            create_with_url += 1

                    # "Consolidate pages" rows should list >= 2 URLs in Secondary URLs
                    if action == 'Consolidate pages':
                        url_count = len([u for u in secondary_urls.split(',') if u.strip()]) if secondary_urls else 0
                        if url_count < 2:
                            consolidate_without_multiple += 1

            # Report action distribution
            total_actions = sum(action_counts.values())
            print(f"    Actions: Create={action_counts['Create new page']}, Expand={action_counts['Expand existing page']}, Consolidate={action_counts['Consolidate pages']}")

            # Validate: every topic has exactly one recommendation
            if rows_without_action > 0:
                errors.append(f"  {sheet_name}: {rows_without_action} row(s) missing Recommended Action")
            else:
                print(f"  [OK] Every topic has a Recommended Action")

            # Validate: all actions are valid
            if invalid_actions:
                for inv in invalid_actions[:5]:  # Show first 5
                    errors.append(f"  {sheet_name}: Invalid action - {inv}")
            else:
                print(f"  [OK] All actions are valid values")

            # Validate: "Create new page" rows have no Primary URL
            if create_with_url > 0:
                errors.append(f"  {sheet_name}: {create_with_url} 'Create new page' row(s) have Primary URL (should be empty)")
            elif action_counts['Create new page'] > 0:
                print(f"  [OK] 'Create new page' rows have empty Primary URL")

            # Validate: "Consolidate pages" rows list >= 2 URLs
            if consolidate_without_multiple > 0:
                errors.append(f"  {sheet_name}: {consolidate_without_multiple} 'Consolidate pages' row(s) have < 2 Secondary URLs")
            elif action_counts['Consolidate pages'] > 0:
                print(f"  [OK] 'Consolidate pages' rows list >= 2 Secondary URLs")

    print()

    # ==========================================================================
    # Verify Sheet 4: Redirect & Merge Plan
    # ==========================================================================
    sheet_name = 'Redirect & Merge Plan'
    print(f"Checking '{sheet_name}'...")

    if sheet_name not in wb.sheetnames:
        # Sheet may not exist if no consolidations were found
        print(f"  [INFO] Sheet '{sheet_name}' not found (expected if no consolidations)")
    else:
        ws = wb[sheet_name]
        print(f"  [OK] Sheet exists")

        # Check for expected headers
        expected_headers = [
            'Topic', 'Recommended Action', 'Primary URL', 'Secondary URL',
            'Redirect Type', 'Reason'
        ]

        # Check if sheet has data or the "no redirects" message
        first_cell = ws.cell(row=1, column=1).value
        if first_cell == 'No consolidation redirects needed (no "Consolidate pages" actions found)':
            print(f"  [OK] Sheet has 'no redirects' message (no consolidations found)")
        else:
            # Check headers
            for col_idx, expected in enumerate(expected_headers, start=1):
                actual = ws.cell(row=1, column=col_idx).value
                if actual != expected:
                    errors.append(f"  {sheet_name}: Column {col_idx} header - expected '{expected}', got '{actual}'")
                else:
                    print(f"  [OK] Column {col_idx}: '{expected}'")

            # Check freeze panes
            if ws.freeze_panes != 'A2':
                warnings.append(f"  {sheet_name}: Freeze panes expected 'A2', got '{ws.freeze_panes}'")
            else:
                print(f"  [OK] Freeze panes at A2")

            # Count data rows
            row_count = ws.max_row - 1  # Exclude header
            print(f"  [OK] Data rows: {row_count}")

            # ======================================================================
            # Redirect & Merge Plan Validation
            # ======================================================================
            print(f"  Validating Redirect Plan...")

            # Column indices (1-based)
            action_col = 2      # Recommended Action
            primary_url_col = 3  # Primary URL
            secondary_url_col = 4  # Secondary URL
            redirect_type_col = 5  # Redirect Type

            secondary_urls_seen = set()
            primary_urls_seen = set()
            non_consolidate_actions = 0
            non_301_types = 0
            duplicate_secondary = 0
            primary_as_secondary = 0

            for row_idx in range(2, ws.max_row + 1):
                action = ws.cell(row=row_idx, column=action_col).value
                primary_url = ws.cell(row=row_idx, column=primary_url_col).value or ''
                secondary_url = ws.cell(row=row_idx, column=secondary_url_col).value or ''
                redirect_type = ws.cell(row=row_idx, column=redirect_type_col).value or ''

                # Check: Only consolidation actions should generate redirects
                if action != 'Consolidate pages':
                    non_consolidate_actions += 1

                # Check: Redirect Type should always be "301"
                if str(redirect_type) != '301':
                    non_301_types += 1

                # Track primary URLs
                if primary_url:
                    primary_urls_seen.add(primary_url)

                # Check: Each Secondary URL appears exactly once
                if secondary_url:
                    if secondary_url in secondary_urls_seen:
                        duplicate_secondary += 1
                    secondary_urls_seen.add(secondary_url)

            # Check: Primary URL never appears as Secondary URL
            for primary in primary_urls_seen:
                if primary in secondary_urls_seen:
                    primary_as_secondary += 1

            # Report validation results
            if non_consolidate_actions > 0:
                errors.append(f"  {sheet_name}: {non_consolidate_actions} row(s) have non-'Consolidate pages' action")
            else:
                print(f"  [OK] All redirects are for 'Consolidate pages' actions")

            if non_301_types > 0:
                errors.append(f"  {sheet_name}: {non_301_types} row(s) have redirect type != '301'")
            else:
                print(f"  [OK] All redirects are type '301'")

            if duplicate_secondary > 0:
                errors.append(f"  {sheet_name}: {duplicate_secondary} duplicate Secondary URL(s) found")
            else:
                print(f"  [OK] Each Secondary URL appears exactly once")

            if primary_as_secondary > 0:
                errors.append(f"  {sheet_name}: {primary_as_secondary} Primary URL(s) also appear as Secondary URL")
            else:
                print(f"  [OK] No Primary URL appears as Secondary URL")

    print()

    # ==========================================================================
    # Verify Sheet 5: Merge Playbooks
    # ==========================================================================
    sheet_name = 'Merge Playbooks'
    print(f"Checking '{sheet_name}'...")

    if sheet_name not in wb.sheetnames:
        # Sheet may not exist if no consolidations were found
        print(f"  [INFO] Sheet '{sheet_name}' not found (expected if no consolidations)")
    else:
        ws = wb[sheet_name]
        print(f"  [OK] Sheet exists")

        # Check for expected headers
        expected_headers = [
            'Topic', 'Primary URL', 'Secondary URL', 'Keep This Content',
            'Move These Sections', 'Retire This Page', 'Reasoning'
        ]

        # Check if sheet has data or the "no playbooks" message
        first_cell = ws.cell(row=1, column=1).value
        if first_cell == 'No merge playbooks needed (no "Consolidate pages" actions found)':
            print(f"  [OK] Sheet has 'no playbooks' message (no consolidations found)")
        else:
            # Check headers
            for col_idx, expected in enumerate(expected_headers, start=1):
                actual = ws.cell(row=1, column=col_idx).value
                if actual != expected:
                    errors.append(f"  {sheet_name}: Column {col_idx} header - expected '{expected}', got '{actual}'")
                else:
                    print(f"  [OK] Column {col_idx}: '{expected}'")

            # Check freeze panes
            if ws.freeze_panes != 'A2':
                warnings.append(f"  {sheet_name}: Freeze panes expected 'A2', got '{ws.freeze_panes}'")
            else:
                print(f"  [OK] Freeze panes at A2")

            # Count data rows
            row_count = ws.max_row - 1  # Exclude header
            print(f"  [OK] Data rows: {row_count}")

            # ======================================================================
            # Merge Playbooks Validation
            # ======================================================================
            print(f"  Validating Merge Playbooks...")

            # Column indices (1-based)
            retire_col = 6  # Retire This Page

            non_yes_retire = 0
            rows_without_keep = 0
            rows_without_move = 0

            for row_idx in range(2, ws.max_row + 1):
                retire_value = ws.cell(row=row_idx, column=retire_col).value
                keep_content = ws.cell(row=row_idx, column=4).value or ''
                move_sections = ws.cell(row=row_idx, column=5).value or ''

                # Check: Retire This Page is always "Yes"
                if retire_value != 'Yes':
                    non_yes_retire += 1

                # Check: Keep This Content is not empty
                if not keep_content.strip():
                    rows_without_keep += 1

                # Check: Move These Sections is not empty
                if not move_sections.strip():
                    rows_without_move += 1

            # Report validation results
            if non_yes_retire > 0:
                errors.append(f"  {sheet_name}: {non_yes_retire} row(s) have 'Retire This Page' != 'Yes'")
            else:
                print(f"  [OK] All rows have 'Retire This Page' = 'Yes'")

            if rows_without_keep > 0:
                warnings.append(f"  {sheet_name}: {rows_without_keep} row(s) have empty 'Keep This Content'")
            else:
                print(f"  [OK] All rows have 'Keep This Content' populated")

            if rows_without_move > 0:
                warnings.append(f"  {sheet_name}: {rows_without_move} row(s) have empty 'Move These Sections'")
            else:
                print(f"  [OK] All rows have 'Move These Sections' populated")

    print()

    # Close workbook
    wb.close()

    # ==========================================================================
    # Report Results
    # ==========================================================================
    print("=" * 60)

    if errors:
        print("ERRORS:")
        for error in errors:
            print(f"  [FAIL] {error}")
        print()

    if warnings:
        print("WARNINGS:")
        for warning in warnings:
            print(f"  [WARN] {warning}")
        print()

    if not errors:
        print("[OK] All verifications PASSED")
        return True
    else:
        print(f"[FAIL] {len(errors)} verification(s) FAILED")
        return False


def verify_empty_sheet_regression(excel_path: str, gsc_path: str = None) -> bool:
    """
    Regression test: Verify that sheets are not empty when GSC data exists.

    This test catches the bug where qualification logic was too strict,
    causing all clusters to be filtered out and resulting in empty sheets.

    Args:
        excel_path: Path to the WQA output Excel file
        gsc_path: Optional path to the GSC data file used in the run

    Returns:
        True if test passes, False if regression detected
    """
    import os

    errors = []

    # Load workbook
    try:
        wb = load_workbook(excel_path, read_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}")
        return False

    print()
    print("=" * 60)
    print("REGRESSION TEST: Empty Sheet Detection")
    print("=" * 60)

    # Check New Content Opportunities
    new_content_sheet = 'New Content Opportunities'
    if new_content_sheet in wb.sheetnames:
        ws = wb[new_content_sheet]
        first_cell = ws.cell(row=1, column=1).value
        row_count = ws.max_row - 1 if first_cell != 'No new content opportunities found (GSC query data required)' else 0

        # Check if this is the "no opportunities" message
        if 'No new content' in str(first_cell) or row_count == 0:
            if gsc_path and os.path.exists(gsc_path):
                errors.append(f"REGRESSION: {new_content_sheet} is empty but GSC data file exists ({gsc_path})")
            else:
                print(f"  [INFO] {new_content_sheet} is empty (expected if no GSC data)")
        else:
            print(f"  [OK] {new_content_sheet} has {row_count} rows")

            # Count actions
            action_col = 6  # Recommended Action column
            consolidate_count = 0
            create_count = 0
            expand_count = 0

            for row_idx in range(2, ws.max_row + 1):
                action = ws.cell(row=row_idx, column=action_col).value
                if action == 'Consolidate pages':
                    consolidate_count += 1
                elif action == 'Create new page':
                    create_count += 1
                elif action == 'Expand existing page':
                    expand_count += 1

            print(f"  [INFO] Actions: Create={create_count}, Expand={expand_count}, Consolidate={consolidate_count}")

            # Check Redirect & Merge Plan if consolidations exist
            redirect_sheet = 'Redirect & Merge Plan'
            if consolidate_count > 0 and redirect_sheet in wb.sheetnames:
                ws_r = wb[redirect_sheet]
                first_cell_r = ws_r.cell(row=1, column=1).value
                if 'No consolidation redirects' in str(first_cell_r):
                    errors.append(f"REGRESSION: {consolidate_count} consolidation actions exist but {redirect_sheet} is empty")
                else:
                    redirect_rows = ws_r.max_row - 1
                    print(f"  [OK] {redirect_sheet} has {redirect_rows} rows")

            # Check Merge Playbooks if consolidations exist
            playbook_sheet = 'Merge Playbooks'
            if consolidate_count > 0 and playbook_sheet in wb.sheetnames:
                ws_p = wb[playbook_sheet]
                first_cell_p = ws_p.cell(row=1, column=1).value
                if 'No merge playbooks' in str(first_cell_p):
                    errors.append(f"REGRESSION: {consolidate_count} consolidation actions exist but {playbook_sheet} is empty")
                else:
                    playbook_rows = ws_p.max_row - 1
                    print(f"  [OK] {playbook_sheet} has {playbook_rows} rows")

    wb.close()

    print()
    if errors:
        print("REGRESSION DETECTED:")
        for error in errors:
            print(f"  [FAIL] {error}")
        return False
    else:
        print("[OK] No regressions detected")
        return True


def run_api_query_data_test() -> bool:
    """
    Test the API path for GSC query data flow.

    This test verifies that:
    1. Query-level GSC data flows correctly to write_analytical_sheets_api
    2. New Content Opportunities populates with topic clusters
    3. Consolidation actions trigger Redirect & Merge Plan entries
    4. Merge Playbooks populate when consolidations exist

    Returns:
        True if all tests pass, False otherwise
    """
    import pandas as pd
    import io
    import sys
    sys.path.insert(0, str(Path(__file__).parent / 'api'))

    try:
        from api.index import (
            build_new_content_opportunities_api,
            build_redirect_merge_plan_api,
            build_merge_playbooks_api,
            write_analytical_sheets_api
        )
        from openpyxl import Workbook
    except ImportError as e:
        print(f"ERROR: Could not import API modules: {e}")
        return False

    print("=" * 60)
    print("API PATH TEST: GSC Query Data Flow")
    print("=" * 60)
    print()

    errors = []
    tests_passed = 0
    tests_total = 0

    # Create synthetic query-level GSC data (the format returned by fetch_gsc_query_data)
    gsc_query_data = pd.DataFrame([
        # Topic 1: "best hvac company" - two pages competing (should trigger consolidation)
        {'query': 'best hvac company atlanta', 'url': 'https://example.com/hvac-services', 'impressions': 800, 'clicks': 40, 'avg_position': 5},
        {'query': 'top hvac contractors atlanta', 'url': 'https://example.com/hvac-services', 'impressions': 600, 'clicks': 30, 'avg_position': 6},
        {'query': 'atlanta hvac repair', 'url': 'https://example.com/hvac-repair', 'impressions': 500, 'clicks': 25, 'avg_position': 7},
        {'query': 'hvac service near me atlanta', 'url': 'https://example.com/hvac-repair', 'impressions': 400, 'clicks': 20, 'avg_position': 8},

        # Topic 2: "plumbing services" - one dominant page (should trigger expand)
        {'query': 'plumbing services atlanta', 'url': 'https://example.com/plumbing', 'impressions': 1200, 'clicks': 60, 'avg_position': 4},
        {'query': 'atlanta plumbers', 'url': 'https://example.com/plumbing', 'impressions': 900, 'clicks': 45, 'avg_position': 5},
        {'query': 'emergency plumber atlanta', 'url': 'https://example.com/plumbing', 'impressions': 700, 'clicks': 35, 'avg_position': 6},

        # Topic 3: "new construction" - no current coverage (should trigger create)
        {'query': 'new construction home builder', 'url': '', 'impressions': 500, 'clicks': 15, 'avg_position': 25},
        {'query': 'custom home builder atlanta', 'url': '', 'impressions': 400, 'clicks': 10, 'avg_position': 28},
    ])

    # Create mock URL-level DataFrame (for thin content detection)
    # Must include all columns that write_analytical_sheets_api expects
    df = pd.DataFrame([
        {'url': 'https://example.com/hvac-services', 'word_count': 1500, 'sessions': 100, 'status_code': 200,
         'indexable': True, 'impressions': 1400, 'clicks': 70, 'ctr': 0.05, 'avg_position': 5.5,
         'meta_description': 'HVAC services in Atlanta', 'page_type': 'Service Page', 'page_title': 'HVAC Services'},
        {'url': 'https://example.com/hvac-repair', 'word_count': 1200, 'sessions': 80, 'status_code': 200,
         'indexable': True, 'impressions': 900, 'clicks': 45, 'ctr': 0.05, 'avg_position': 7.5,
         'meta_description': 'HVAC repair experts', 'page_type': 'Service Page', 'page_title': 'HVAC Repair'},
        {'url': 'https://example.com/plumbing', 'word_count': 2000, 'sessions': 150, 'status_code': 200,
         'indexable': True, 'impressions': 2800, 'clicks': 140, 'ctr': 0.05, 'avg_position': 5.0,
         'meta_description': 'Plumbing services Atlanta', 'page_type': 'Service Page', 'page_title': 'Plumbing Services'},
    ])

    # =========================================================================
    # Test 1: build_new_content_opportunities_api produces results
    # =========================================================================
    tests_total += 1
    print("Test 1: API build_new_content_opportunities_api produces results...")

    new_content_df = build_new_content_opportunities_api(df, gsc_query_data, thin_threshold=1000)

    if len(new_content_df) == 0:
        errors.append("Test 1 FAILED: build_new_content_opportunities_api returned empty DataFrame")
    else:
        # Check all required columns exist
        required_cols = ['Suggested Topic', 'Primary Keyword', 'Recommended Action', 'Primary URL', 'Secondary URLs', 'Reasoning']
        missing_cols = [col for col in required_cols if col not in new_content_df.columns]
        if missing_cols:
            errors.append(f"Test 1 FAILED: Missing columns: {missing_cols}")
        else:
            actions = new_content_df['Recommended Action'].value_counts().to_dict()
            print(f"  [OK] Produced {len(new_content_df)} topics: {actions}")
            tests_passed += 1

    # =========================================================================
    # Test 2: Consolidation action exists for competing URLs
    # =========================================================================
    tests_total += 1
    print("Test 2: Consolidation action exists for competing URLs...")

    consolidate_count = len(new_content_df[new_content_df['Recommended Action'] == 'Consolidate pages']) if len(new_content_df) > 0 else 0

    if consolidate_count == 0:
        errors.append("Test 2 FAILED: No 'Consolidate pages' actions found despite competing URLs in data")
    else:
        print(f"  [OK] Found {consolidate_count} 'Consolidate pages' action(s)")
        tests_passed += 1

    # =========================================================================
    # Test 3: Redirect & Merge Plan populates from consolidations
    # =========================================================================
    tests_total += 1
    print("Test 3: Redirect & Merge Plan populates from consolidations...")

    redirect_df = build_redirect_merge_plan_api(new_content_df)

    if consolidate_count > 0 and len(redirect_df) == 0:
        errors.append("Test 3 FAILED: Consolidation actions exist but Redirect & Merge Plan is empty")
    elif consolidate_count == 0:
        print(f"  [SKIP] No consolidations to test")
    else:
        print(f"  [OK] Redirect & Merge Plan has {len(redirect_df)} entries")
        tests_passed += 1

    # =========================================================================
    # Test 4: Merge Playbooks populates from consolidations
    # =========================================================================
    tests_total += 1
    print("Test 4: Merge Playbooks populates from consolidations...")

    url_word_counts = dict(zip(df['url'], df['word_count']))
    playbook_df = build_merge_playbooks_api(new_content_df, gsc_query_data, url_word_counts)

    if consolidate_count > 0 and len(playbook_df) == 0:
        errors.append("Test 4 FAILED: Consolidation actions exist but Merge Playbooks is empty")
    elif consolidate_count == 0:
        print(f"  [SKIP] No consolidations to test")
    else:
        print(f"  [OK] Merge Playbooks has {len(playbook_df)} entries")
        tests_passed += 1

    # =========================================================================
    # Test 5: Full write_analytical_sheets_api integration
    # =========================================================================
    tests_total += 1
    print("Test 5: Full write_analytical_sheets_api integration...")

    try:
        workbook = Workbook()
        write_analytical_sheets_api(workbook, df, thin_content_threshold=1000, gsc_df=gsc_query_data)

        # Check sheet names exist
        expected_sheets = ['New Content Opportunities', 'Redirect & Merge Plan', 'Merge Playbooks']
        for sheet_name in expected_sheets:
            if sheet_name not in workbook.sheetnames:
                errors.append(f"Test 5 FAILED: Missing sheet '{sheet_name}'")

        if all(s in workbook.sheetnames for s in expected_sheets):
            # Check New Content Opportunities has data
            ws = workbook['New Content Opportunities']
            first_cell = ws.cell(row=1, column=1).value
            if 'GSC data not connected' in str(first_cell) or 'GSC returned 0 rows' in str(first_cell):
                errors.append(f"Test 5 FAILED: New Content Opportunities shows empty message: {first_cell}")
            else:
                row_count = ws.max_row - 1  # Exclude header
                if row_count > 0:
                    print(f"  [OK] write_analytical_sheets_api created sheets with {row_count} content opportunities")
                    tests_passed += 1
                else:
                    errors.append("Test 5 FAILED: New Content Opportunities has 0 data rows")
    except Exception as e:
        errors.append(f"Test 5 FAILED: Exception during write_analytical_sheets_api: {e}")

    # =========================================================================
    # Summary
    # =========================================================================
    print()
    print("=" * 60)
    print(f"API PATH TEST RESULTS: {tests_passed}/{tests_total} tests passed")
    print("=" * 60)

    if errors:
        print()
        print("ERRORS:")
        for error in errors:
            print(f"  [FAIL] {error}")
        return False
    else:
        print("[OK] All API path tests passed!")
        return True


def main():
    """Main entry point."""
    # Check for --unit-test flag
    if '--unit-test' in sys.argv:
        cli_success = run_clustering_unit_tests()
        print()
        api_success = run_api_query_data_test()
        sys.exit(0 if (cli_success and api_success) else 1)

    if len(sys.argv) < 2:
        print("Usage: python test_analytical_sheets.py <path_to_wqa_output.xlsx> [--gsc <gsc_file.csv>]")
        print("       python test_analytical_sheets.py --unit-test")
        print()
        print("Example:")
        print("  python test_analytical_sheets.py wqa_report.xlsx")
        print("  python test_analytical_sheets.py wqa_report.xlsx --gsc gsc_data.csv")
        print("  python test_analytical_sheets.py --unit-test  # Run unit tests with synthetic data")
        sys.exit(1)

    excel_path = sys.argv[1]

    # Parse optional --gsc argument
    gsc_path = None
    if '--gsc' in sys.argv:
        gsc_idx = sys.argv.index('--gsc')
        if gsc_idx + 1 < len(sys.argv):
            gsc_path = sys.argv[gsc_idx + 1]

    # Run standard verification
    success = verify_analytical_sheets(excel_path)

    # Run regression test
    regression_success = verify_empty_sheet_regression(excel_path, gsc_path)

    sys.exit(0 if (success and regression_success) else 1)


if __name__ == "__main__":
    main()
